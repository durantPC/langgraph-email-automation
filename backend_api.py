"""
邮件自动化系统 - 后端API
提供前端所需的所有接口
"""
import os
import json
import re
import asyncio
import threading
import time
import csv
import io
import uuid
from threading import Lock
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from typing import Optional, List
from urllib.parse import unquote, quote
from typing import Dict, Set

# 尝试导入openpyxl，如果没有则使用CSV
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_DATE_DATETIME  # pyright: ignore[reportMissingModuleSource]
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl未安装，将使用CSV格式导出。要使用XLSX格式，请运行: pip install openpyxl")
from fastapi import FastAPI, HTTPException, BackgroundTasks, WebSocket, WebSocketDisconnect, Header, Depends, Request, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel, Field
from dotenv import load_dotenv
from typing import Optional
from contextlib import asynccontextmanager

# 加载环境变量
load_dotenv()

# 导入邮件工具
from src.tools.QQEmailTools import QQEmailToolsClass
from src.tools.EmailUrgencyDetector import analyze_email_urgency

# 保存主事件循环引用，在线程中使用 run_coroutine_threadsafe 推送
websocket_event_loop: Optional[asyncio.AbstractEventLoop] = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期管理"""
    # Startup
    global websocket_event_loop
    try:
        websocket_event_loop = asyncio.get_event_loop()
       
    except Exception as e:
        print(f"⚠️ [WS] 无法保存事件循环: {e}")
    
    yield
    
    # Shutdown (如果需要清理资源，可以在这里添加)
    print("🔄 [应用] 正在关闭...")

app = FastAPI(
    title="邮件自动化系统 API",
    version="1.0.0",
    description="基于AI代理和RAG的客户支持邮件自动化系统后端API",
    lifespan=lifespan
)

# CORS配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== 线程池管理 ====================

# 主线程池：用于常规操作（邮件获取、索引构建等）
# 固定大小，确保其他API请求不受影响
thread_pool = ThreadPoolExecutor(max_workers=4, thread_name_prefix="email_processor")

# 单封邮件处理线程池：专门用于单封邮件处理，独立于主线程池
# 根据用户设置的 singleEmailConcurrency 动态调整
# 避免单封邮件处理占用主线程池，影响其他API请求
single_email_thread_pool = None
single_email_thread_pool_lock = Lock()

def get_or_create_single_email_thread_pool(concurrency: int = 4):
    """
    获取或创建单封邮件处理专用的线程池
    这个线程池独立于主线程池，不会影响其他API请求
    
    线程池大小应该 >= concurrency，但不超过20（避免资源耗尽）
    """
    global single_email_thread_pool
    with single_email_thread_pool_lock:
        # 计算需要的线程数：concurrency + 1（预留一些线程）
        required_workers = min(max(concurrency, 2), 20)  # 最小2，最大20
        
        # 如果还没有创建，或者当前线程池大小不够，创建新的线程池
        if single_email_thread_pool is None or single_email_thread_pool._max_workers < required_workers:
            if single_email_thread_pool is not None:
                print(f"🔄 [单封邮件线程池] 调整线程池大小: {single_email_thread_pool._max_workers} -> {required_workers} (concurrency={concurrency})")
                # 关闭旧线程池（不等待，因为新任务会使用新线程池）
                single_email_thread_pool.shutdown(wait=False)
            else:
                print(f"🔄 [单封邮件线程池] 创建单封邮件处理线程池: {required_workers} 个工作线程 (concurrency={concurrency})")
            
            # 创建新线程池
            single_email_thread_pool = ThreadPoolExecutor(
                max_workers=required_workers, 
                thread_name_prefix="single_email_processor"
            )
        
        return single_email_thread_pool

# 批量处理线程池：专门用于批量邮件处理，根据用户设置的 batchSize 动态调整
# 独立线程池，不会影响其他操作
batch_thread_pool = None
batch_thread_pool_lock = Lock()

def get_or_create_batch_thread_pool(batch_size: int = 4):
    """
    获取或创建批量处理专用的线程池
    这个线程池独立于主线程池，不会影响其他API请求
    
    线程池大小应该 >= batch_size，但不超过30（避免资源耗尽）
    """
    global batch_thread_pool
    with batch_thread_pool_lock:
        # 计算需要的线程数：batch_size + 2（预留一些线程）
        required_workers = min(max(batch_size, 4), 30)  # 最小4，最大30
        
        # 如果还没有创建，或者当前线程池大小不够，创建新的线程池
        if batch_thread_pool is None or batch_thread_pool._max_workers < required_workers:
            if batch_thread_pool is not None:
                print(f"🔄 [批量线程池] 调整线程池大小: {batch_thread_pool._max_workers} -> {required_workers} (batch_size={batch_size})")
                # 关闭旧线程池（不等待，因为新任务会使用新线程池）
                batch_thread_pool.shutdown(wait=False)
            else:
                print(f"🔄 [批量线程池] 创建批量处理线程池: {required_workers} 个工作线程 (batch_size={batch_size})")
            
            # 创建新线程池
            batch_thread_pool = ThreadPoolExecutor(
                max_workers=required_workers, 
                thread_name_prefix="batch_email_processor"
            )
        
        return batch_thread_pool

# 摘要生成线程池：专门用于异步生成邮件摘要，避免阻塞其他操作
# 固定大小，限制并发摘要生成的数量，防止资源耗尽
# 每个摘要生成任务内部会并发生成 body_summary 和 reply_summary
summary_generation_pool = ThreadPoolExecutor(
    max_workers=15,  # 最多同时生成 15 封邮件的摘要（每封邮件内部并发生成2个摘要）
    thread_name_prefix="summary_generator"
)


# 用户级锁字典，用于保护同一用户的并发操作（如邮件处理、状态更新等）
user_locks: dict[str, Lock] = {}

def get_user_lock(username: str) -> Lock:
    """获取用户的锁，如果不存在则创建"""
    if username not in user_locks:
        user_locks[username] = Lock()
    return user_locks[username]

# RAG测试取消标志字典（按用户存储）
rag_test_cancelled: dict[str, threading.Event] = {}

def get_rag_cancel_flag(username: str) -> threading.Event:
    """获取用户的RAG测试取消标志"""
    if username not in rag_test_cancelled:
        rag_test_cancelled[username] = threading.Event()
    return rag_test_cancelled[username]

def clear_rag_cancel_flag(username: str):
    """清除用户的RAG测试取消标志"""
    if username in rag_test_cancelled:
        rag_test_cancelled[username].clear()

# ==================== 发送速率控制 ====================

# 发送速率控制配置
SEND_RATE_LIMIT = {
    "per_hour": 20,        # 每小时最多20封
    "per_half_hour": 10,   # 每半小时最多10封
    "interval": 30         # 每封邮件之间至少间隔30秒（便于测试）
}

# 发送速率控制（按用户）
send_rate_control = {}  # {username: {"last_send_time": timestamp, "send_count_hour": count, "send_count_half_hour": count, "reset_time_hour": timestamp, "reset_time_half_hour": timestamp}}
send_rate_lock = Lock()

def check_send_rate_limit(username: str) -> tuple[bool, str]:
    """
    检查是否达到发送速率限制
    返回 (是否可以发送, 提示信息)
    """
    with send_rate_lock:
        now = time.time()
        
        if username not in send_rate_control:
            send_rate_control[username] = {
                "last_send_time": 0,
                "send_count_hour": 0,
                "send_count_half_hour": 0,
                "reset_time_hour": now + 3600,      # 1小时后重置
                "reset_time_half_hour": now + 1800  # 30分钟后重置
            }
            print(f"✅ [速率限制] 用户 {username} 首次发送，允许发送")
            return True, ""
        
        control = send_rate_control[username]
        
        # 检查每小时限制
        if now > control["reset_time_hour"]:
            control["send_count_hour"] = 0
            control["reset_time_hour"] = now + 3600
        
        # 检查每半小时限制
        if now > control["reset_time_half_hour"]:
            control["send_count_half_hour"] = 0
            control["reset_time_half_hour"] = now + 1800
        
        # 检查是否达到每小时限制
        if control["send_count_hour"] >= SEND_RATE_LIMIT["per_hour"]:
            remaining_time = control["reset_time_hour"] - now
            remaining_minutes = int(remaining_time / 60)
            msg = f"已达到每小时发送限制（{SEND_RATE_LIMIT['per_hour']}封），请等待 {remaining_minutes} 分钟后重试"
            print(f"⏸️ [速率限制] 用户 {username}: {msg}")
            return False, msg
        
        # 检查是否达到每半小时限制
        if control["send_count_half_hour"] >= SEND_RATE_LIMIT["per_half_hour"]:
            remaining_time = control["reset_time_half_hour"] - now
            remaining_minutes = int(remaining_time / 60)
            msg = f"已达到每半小时发送限制（{SEND_RATE_LIMIT['per_half_hour']}封），请等待 {remaining_minutes} 分钟后重试"
            print(f"⏸️ [速率限制] 用户 {username}: {msg}")
            return False, msg
        
        # 检查发送间隔（至少30秒）
        # 注意：如果 last_send_time 为 0，说明是首次发送，允许发送
        if control["last_send_time"] > 0:
            time_since_last = now - control["last_send_time"]
            print(f"🔍 [速率限制] 用户 {username} 距离上次发送: {int(time_since_last)} 秒 (需要间隔 {SEND_RATE_LIMIT['interval']} 秒)")
            if time_since_last < SEND_RATE_LIMIT["interval"]:
                wait_time = SEND_RATE_LIMIT["interval"] - time_since_last
                wait_seconds = int(wait_time)
                msg = f"发送间隔不足，请等待 {wait_seconds} 秒后重试（每封邮件需间隔30秒）"
                print(f"⏸️ [速率限制] 用户 {username}: {msg}")
                return False, msg
            print(f"✅ [速率限制] 用户 {username} 检查通过，允许发送 (距离上次发送 {int(time_since_last)} 秒)")
        else:
            print(f"✅ [速率限制] 用户 {username} 首次发送，允许发送")
        
        return True, ""

def update_send_rate_limit(username: str):
    """更新发送速率限制计数"""
    with send_rate_lock:
        now = time.time()
        if username not in send_rate_control:
            send_rate_control[username] = {
                "last_send_time": now,
                "send_count_hour": 1,
                "send_count_half_hour": 1,
                "reset_time_hour": now + 3600,
                "reset_time_half_hour": now + 1800
            }
        else:
            control = send_rate_control[username]
            old_time = control["last_send_time"]
            control["last_send_time"] = now
            control["send_count_hour"] += 1
            control["send_count_half_hour"] += 1
            old_time_str = datetime.fromtimestamp(old_time).strftime('%H:%M:%S') if old_time > 0 else '首次'
            print(f"📝 [速率限制] 用户 {username} 更新计数: {control['send_count_hour']}/20 (小时), {control['send_count_half_hour']}/10 (半小时), 上次发送: {old_time_str}, 当前时间: {datetime.fromtimestamp(now).strftime('%H:%M:%S')}")

def send_processed_emails_with_rate_limit(username: str):
    """
    在后台发送所有已处理的邮件（受速率限制）
    当用户开启自动发送并保存设置时调用
    """
    try:
        print(f"📧 [自动发送] 开始检查用户 {username} 的已处理邮件...")
        user_state = get_user_state(username, check_auto_start=False)
        
        # 使用用户锁保护数据读取（避免读取到不一致的数据）
        user_lock = get_user_lock(username)
        with user_lock:
            # 调试：打印所有邮件的状态
            print(f"🔍 [自动发送] 调试信息：缓存中共有 {len(user_state.emails_cache)} 封邮件")
            for idx, email in enumerate(user_state.emails_cache):
                status = email.get('status', '未知')
                has_reply = bool(email.get('reply'))
                subject = email.get('subject', '无主题')[:30]
                print(f"  [{idx+1}] 状态: {status}, 有回复: {has_reply}, 主题: {subject}")
            
            # 查找所有已处理且有回复内容的邮件（复制一份，避免在锁外访问）
            processed_emails = [
                e.copy() for e in user_state.emails_cache 
                if e.get('status') == 'processed' and e.get('reply')
            ]
        
        if not processed_emails:
            print(f"📧 [自动发送] 用户 {username} 没有需要发送的已处理邮件")
            # 调试：检查是否有状态为 processed 但没有 reply 的邮件
            with user_lock:
                processed_without_reply = [
                    e for e in user_state.emails_cache 
                    if e.get('status') == 'processed' and not e.get('reply')
                ]
            if processed_without_reply:
                print(f"⚠️ [自动发送] 发现 {len(processed_without_reply)} 封已处理但无回复内容的邮件（无法发送）")
            return
        
        print(f"📧 [自动发送] 检测到 {len(processed_emails)} 封已处理邮件，开始自动发送（受速率限制）...")
        
        # 导入 nodes（需要在这里导入，因为是在后台任务中执行）
        from src.nodes import Nodes
        # 获取用户设置中的模板配置
        user_settings = get_user_settings(username)
        nodes = Nodes(
            signature=user_settings.get("signature"),
            greeting=user_settings.get("greeting"),
            closing=user_settings.get("closing")
        )
        
        # 创建邮件对象类
        class EmailObj:
            def __init__(self, data):
                # 确保 sender 不为空，如果为空则抛出错误
                sender = data.get('sender', '').strip()
                if not sender:
                    print(f"❌ [自动发送] 错误：邮件数据中缺少发件人地址")
                    print(f"   邮件数据: {data}")
                    raise ValueError(f"邮件数据中缺少发件人地址: {data.get('subject', '无主题')}")
                
                # 如果 sender 包含 < >，提取邮箱地址部分（与 fetch_unanswered_emails 中的逻辑一致）
                if '<' in sender and '>' in sender:
                    try:
                        sender = sender.split('<')[1].split('>')[0].strip()
                    except (IndexError, AttributeError):
                        print(f"⚠️ [自动发送] 警告：无法从发件人地址中提取邮箱，使用原始值: {sender}")
                
                # 验证邮箱地址格式
                if '@' not in sender:
                    print(f"❌ [自动发送] 错误：发件人地址格式无效（缺少@符号）: {sender}")
                    raise ValueError(f"无效的发件人地址格式（缺少@符号）: {sender}")
                
                self.sender = sender
                self.subject = data.get('subject', '')
                self.messageId = data.get('messageId', '')
                self.references = data.get('references', '')
                self.imap_id = data.get('imap_id', b'')
                
                print(f"📧 [自动发送] 创建邮件对象: 发件人={self.sender}, 主题={self.subject[:30]}")
        
        # 遍历已处理的邮件，使用速率限制发送
        sent_count = 0
        failed_count = 0
        limited_count = 0
        
        for email in processed_emails:
            email_obj = EmailObj(email)
            result, message = send_reply_with_rate_limit(
                username,
                nodes.email_tools,
                email_obj,
                email.get('reply', ''),
                email
            )
            
            if result:
                sent_count += 1
                sender_name = email.get('sender', '').split('@')[0] if '@' in email.get('sender', '') else email.get('sender', '未知')
                print(f"✓ [自动发送] 成功发送: {email.get('subject', '')} -> {sender_name}")
            else:
                # 判断是速率限制还是发送失败
                if "限制" in message or "间隔" in message or "等待" in message:
                    limited_count += 1
                    print(f"⏸️ [自动发送] 达到速率限制，暂停发送: {email.get('subject', '')} - {message}")
                    # 如果是间隔限制（每封邮件之间需要间隔），继续尝试下一封
                    # 如果是数量限制（每小时/每半小时上限），停止发送
                    if "间隔" in message or "等待" in message:
                        # 间隔限制：继续尝试下一封（虽然可能也会被限制，但至少会尝试）
                        print(f"💡 [自动发送] 间隔限制，继续尝试下一封邮件...")
                        continue
                    else:
                        # 数量限制：停止发送剩余邮件（等待下次触发）
                        print(f"💡 [自动发送] 数量限制，停止发送剩余邮件，等待下次检查...")
                        break
                else:
                    failed_count += 1
                    print(f"⚠️ [自动发送] 发送失败: {email.get('subject', '')} - {message}")
        
        # 计算剩余待发送的邮件数（不包括因速率限制暂停的，因为它们仍然需要发送）
        remaining_count = len(processed_emails) - sent_count - failed_count
        print(f"📊 [自动发送] 完成: 成功 {sent_count} 封, 失败 {failed_count} 封, 因速率限制暂停 {limited_count} 封, 剩余 {remaining_count} 封待发送")
        
    except Exception as e:
        print(f"❌ [自动发送] 发送已处理邮件时出错: {e}")
        import traceback
        traceback.print_exc()

def send_reply_with_rate_limit(username: str, email_tools, email_obj, reply_text: str, email_data: dict) -> tuple[bool, str]:
    """
    带速率限制的发送回复函数
    返回 (是否成功, 提示信息)
    """
    # 检查速率限制
    print(f"🔍 [速率限制] 开始检查用户 {username} 的发送限制...")
    can_send, message = check_send_rate_limit(username)
    if not can_send:
        print(f"⏸️ [速率限制] 用户 {username} 发送被限制: {message}")
        return False, message
    
    # 执行发送
    print(f"📧 [速率限制] 用户 {username} 开始发送邮件: {email_data.get('subject', '')}")
    try:
        result = email_tools.send_reply(email_obj, reply_text)
        if result:
            # 更新速率限制计数
            print(f"📝 [速率限制] 用户 {username} 发送成功，更新速率限制计数...")
            update_send_rate_limit(username)
            
            # 更新邮件状态和统计（确保与个人中心统计同步）
            user_state = get_user_state(username, check_auto_start=False)
            for email in user_state.emails_cache:
                if email.get('id') == email_data.get('id'):
                    email['status'] = 'sent'
                    email['reply'] = reply_text  # 保存回复内容
                    
                    # 发送成功后标记为已读
                    imap_id = email.get('imap_id')
                    if imap_id:
                        try:
                            email_tools.mark_email_as_read(imap_id)
                        except Exception as e:
                            print(f"⚠️ [自动发送] 标记已读失败: {e}")
                    
                    # 更新历史记录（如果已存在则更新，否则添加）
                    email_id = email.get('id', '')
                    email_subject = email.get('subject', '')
                    email_sender = email.get('sender', '')
                    
                    history_updated = False
                    for history_record in user_state.history:
                        # 匹配条件：ID相同，或者主题和发件人都相同
                        if (history_record.get('id') == email_id or 
                            (history_record.get('subject') == email_subject and 
                             history_record.get('sender') == email_sender)):
                            # 更新历史记录
                            history_record['reply'] = reply_text
                            history_record['status'] = 'sent'
                            history_record['processed_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            history_updated = True
                            break
                    
                    if not history_updated:
                        # 添加到历史记录
                        history_record = {
                            **email,
                            'reply': reply_text,
                            'status': 'sent',
                            'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        if not history_record.get('id'):
                            history_record['id'] = email_id
                        user_state.history.insert(0, history_record)
                    
                    # 更新统计
                    user_state.stats['sent'] = user_state.stats.get('sent', 0) + 1
                    
                    # 从缓存中移除邮件（与手动发送保持一致）
                    user_state.emails_cache.remove(email)
                    print(f"📧 [自动发送] 已从缓存中移除邮件: {email.get('subject', '')}")
                    
                    # 保存数据
                    save_user_email_data(username, user_state)
                    break
            
            # 获取当前发送计数（用于日志）
            with send_rate_lock:
                control = send_rate_control.get(username, {})
                count_hour = control.get("send_count_hour", 0)
                count_half_hour = control.get("send_count_half_hour", 0)
            
            print(f"✓ 自动发送回复成功: {email_data.get('subject', '')} (今日第 {count_hour} 封，半小时内第 {count_half_hour} 封)")
            return True, "发送成功"
        else:
            return False, "发送失败"
    except Exception as e:
        error_msg = f"发送时出错: {str(e)}"
        print(f"❌ {error_msg}")
        return False, error_msg

# ==================== 数据模型 ====================

class LoginRequest(BaseModel):
    username: str
    password: str

class RegisterRequest(BaseModel):
    username: str
    password: str
    email: str  # 必填项：QQ邮箱

class PreferencesRequest(BaseModel):
    theme: Optional[str] = "light"
    notification: Optional[bool] = True
    sound: Optional[bool] = False
    defaultPage: Optional[str] = "/dashboard"

class LoginResponse(BaseModel):
    token: str
    username: str

class EmailItem(BaseModel):
    id: str
    sender: str
    subject: str
    preview: str
    body: str
    time: str
    category: Optional[str] = None
    status: str = "pending"
    reply: Optional[str] = None

class ProcessEmailRequest(BaseModel):
    email_id: str

class SendReplyRequest(BaseModel):
    email_id: str
    reply: str

class UpdateReplyRequest(BaseModel):
    email_id: str
    reply: str

class MarkReadRequest(BaseModel):
    email_id: str

class SettingsModel(BaseModel):
    email: Optional[str] = None
    authCode: Optional[str] = None
    apiKey: Optional[str] = None  # 保留用于向后兼容，但不再在前端显示
    model: Optional[str] = None  # 保留用于向后兼容
    replyModel: Optional[str] = None  # 回复大模型
    embeddingModel: Optional[str] = None  # 嵌入大模型
    interval: Optional[int] = None
    autoProcess: Optional[bool] = None  # 监控运行时自动处理新邮件
    autoSend: Optional[bool] = None
    batchSize: Optional[int] = None  # 每批并发处理的邮件数量（1-30）
    singleEmailConcurrency: Optional[int] = None  # 单封邮件处理的并发数量（2-20）
    signature: Optional[str] = None
    greeting: Optional[str] = None
    closing: Optional[str] = None

class CustomModelModel(BaseModel):
    provider: str
    model: str
    apiKey: str
    type: str  # 'reply' 或 'embedding'
    apiBaseUrl: Optional[str] = None  # 自定义API base URL，如果为空则根据provider自动推断

class TestAIRequest(BaseModel):
    apiKey: Optional[str] = None  # 如果提供了自定义模型的API，则使用它；否则使用系统默认API
    replyModel: Optional[str] = None
    embeddingModel: Optional[str] = None
    replyApiBaseUrl: Optional[str] = None  # 回复模型的API base URL
    embeddingApiBaseUrl: Optional[str] = None  # 嵌入模型的API base URL

class TestEmailRequest(BaseModel):
    email: Optional[str] = None
    authCode: Optional[str] = None

class RAGTestRequest(BaseModel):
    question: str

# ==================== 辅助函数 ====================

def get_api_base_url(provider: str, custom_url: Optional[str] = None) -> str:
    """
    根据服务商获取API base URL
    
    Args:
        provider: 服务商名称
        custom_url: 自定义API base URL（优先使用）
    
    Returns:
        API base URL
    """
    # 如果提供了自定义URL，直接使用
    if custom_url:
        return custom_url
    
    # 预定义的服务商映射
    provider_mapping = {
        "硅基流动": "https://api.siliconflow.cn/v1",
        "OpenAI": "https://api.openai.com/v1",
        "Anthropic": "https://api.anthropic.com/v1",
        "DeepSeek": "https://api.deepseek.com/v1",
        "Moonshot": "https://api.moonshot.cn/v1",
        "智谱AI": "https://open.bigmodel.cn/api/paas/v4",
        "阿里云": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "腾讯云": "https://api.hunyuan.cloud.tencent.com/v1",
    }
    
    # 返回对应的URL，如果没有匹配则默认使用硅基流动
    return provider_mapping.get(provider, "https://api.siliconflow.cn/v1")

def auto_classify_email(subject, body):
    """根据邮件主题和内容自动分类"""
    text = (subject + ' ' + body).lower()
    
    # 投诉相关关键词
    if any(word in text for word in ['投诉', '不满', '差评', '退款', '问题严重', '态度差', '垃圾', '骗子']):
        return 'customer_complaint'
    
    # 反馈相关关键词
    if any(word in text for word in ['反馈', '建议', '意见', '希望', '改进', '体验']):
        return 'customer_feedback'
    
    # 产品咨询相关关键词
    if any(word in text for word in ['价格', '咨询', '了解', '产品', '功能', '服务', 'api', '接口', '如何', '怎么', '请问', '多少']):
        return 'product_enquiry'
    
    # 无关邮件
    if any(word in text for word in ['广告', '推广', '优惠券', '中奖', '抽奖', '促销', '特价']):
        return 'unrelated'
    
    # 默认为产品咨询
    return 'product_enquiry'

# ==================== 全局状态 ====================

class SystemState:
    def __init__(self, username: str = None):
        self.username = username  # 关联的用户名
        self.is_running = False
        self.auto_process = False  # 自动处理开关
        self.stop_processing = False  # 停止处理标志（用于终止批量处理）
        self.stopped_email_ids = set()  # 被终止的邮件ID集合
        self.monitor_thread = None
        self.last_check_time = None
        self.last_auto_send_check = None  # 上次检查自动发送的时间
        self.check_interval = 900  # 15分钟
        self.emails_cache = []
        self.history = []
        self.activities = []  # 最近操作记录
        self.stats = {
            "today_emails": 0,
            "processed": 0,
            "pending": 0,
            "failed": 0,
            "sent": 0  # 发送回复数
        }
    
    def add_activity(self, activity_type: str, content: str, icon: str = None):
        """添加操作记录"""
        activity = {
            "time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "type": activity_type,  # success, primary, info, warning, danger
            "icon": icon or "CircleCheck",
            "content": content
        }
        self.activities.insert(0, activity)  # 插入到开头
        # 只保留最近50条记录
        if len(self.activities) > 50:
            self.activities = self.activities[:50]
        
    def start_monitor(self):
        if not self.is_running:
            self.is_running = True
            print(f"🚀 [监控系统] 启动监控线程，用户: {self.username}")
            self.monitor_thread = threading.Thread(target=self._monitor_loop, daemon=True)
            self.monitor_thread.start()
            print(f"✅ [监控系统] 监控线程已启动")
            
            # 检查是否开启了自动发送，只有开启时才启动自动发送线程
            user_settings = get_user_settings(self.username)
            if user_settings.get("autoSend", False):
                print(f"🚀 [监控系统] 启动自动发送检查线程，用户: {self.username}")
                self.auto_send_thread = threading.Thread(target=self._auto_send_loop, daemon=True)
                self.auto_send_thread.start()
                print(f"✅ [监控系统] 自动发送检查线程已启动")
            else:
                print(f"ℹ️ [监控系统] 自动发送未开启，跳过启动自动发送线程")
        else:
            print(f"⚠️ [监控系统] 监控已在运行中，跳过启动")
            
    def stop_monitor(self):
        self.is_running = False
        
    def _monitor_loop(self):
        print(f"🔄 [监控循环] 监控循环已启动，用户: {self.username}, 检查间隔: {self.check_interval}秒")
        print(f"🔄 [监控循环] 初始 auto_process 状态: {self.auto_process}")
        while self.is_running:
            try:
                print(f"🔍 [监控循环] 开始检查邮件（用户: {self.username}, 自动处理: {'✅ 开启' if self.auto_process else '❌ 关闭'}）")
                new_emails_count = self._check_emails()
                
                # 如果有新邮件，通知前端刷新
                if new_emails_count > 0:
                    self._notify_frontend({
                        "type": "new_emails",
                        "message": f"检测到 {new_emails_count} 封新邮件",
                        "count": new_emails_count
                    })
                
                # 检查待处理邮件数量
                pending_count = len([e for e in self.emails_cache if e.get('status') == 'pending'])
                print(f"📊 [监控循环] 当前待处理邮件数: {pending_count}")
                
                # 如果开启了自动处理，处理所有待处理邮件
                if self.auto_process:
                    print(f"✅ [监控循环] 自动处理已开启，检查待处理邮件...")
                    if pending_count > 0:
                        # 在线程池中异步执行（不阻塞监控循环）
                        print(f"🚀 [自动处理] 发现 {pending_count} 封待处理邮件，提交到线程池异步处理")
                        
                        def auto_process_callback(future):
                            """自动处理完成后的回调函数"""
                            try:
                                result = future.result()
                                if result:
                                    self._notify_frontend({
                                        "type": "auto_process_complete",
                                        "message": result['message'],
                                        "processed": result['processed'],
                                        "skipped": result['skipped'],
                                        "failed": result['failed']
                                    })
                            except Exception as e:
                                print(f"❌ [自动处理] 处理错误: {e}")
                                import traceback
                                traceback.print_exc()
                        
                        # 提交到线程池异步执行（不阻塞监控循环）
                        future = thread_pool.submit(self._auto_process_emails_async)
                        future.add_done_callback(auto_process_callback)
                    else:
                        # 没有待处理邮件时也输出日志（方便调试）
                        print(f"ℹ️ [自动处理] 自动处理已开启，但当前没有待处理邮件（用户: {self.username}）")
                else:
                    print(f"❌ [监控循环] 自动处理已关闭，跳过自动处理（待处理邮件: {pending_count}）")
                
            except Exception as e:
                print(f"监控循环错误: {e}")
            time.sleep(self.check_interval)
    
    def _auto_send_loop(self):
        """独立的自动发送检查循环，每30秒检查一次"""
        print(f"🔄 [自动发送线程] 线程已启动，用户: {self.username}")
        loop_count = 0
        
        # 立即执行第一次检查（不等待30秒）
        try:
            loop_count += 1
            # 获取用户设置，检查是否开启了自动发送
            user_settings = get_user_settings(self.username)
            if user_settings.get("autoSend", False):
                print(f"🔄 [自动发送线程] 第 {loop_count} 次检查 - 自动发送已开启 (用户: {self.username}, 时间: {datetime.now().strftime('%H:%M:%S')})")
                # 检查并发送已处理的邮件（受速率限制）
                send_processed_emails_with_rate_limit(self.username)
            # 如果未开启，不输出日志（避免日志过多）
        except Exception as e:
            print(f"❌ [自动发送线程] 自动发送检查错误: {e}")
            import traceback
            traceback.print_exc()
        
        # 然后每30秒检查一次
        while self.is_running:
            try:
                loop_count += 1
                # 获取用户设置，检查是否开启了自动发送
                user_settings = get_user_settings(self.username)
                if user_settings.get("autoSend", False):
                    print(f"🔄 [自动发送线程] 第 {loop_count} 次检查 - 自动发送已开启 (用户: {self.username}, 时间: {datetime.now().strftime('%H:%M:%S')})")
                    # 检查并发送已处理的邮件（受速率限制）
                    send_processed_emails_with_rate_limit(self.username)
                # 如果未开启，不输出日志（避免日志过多）
            except Exception as e:
                print(f"❌ [自动发送线程] 自动发送检查错误: {e}")
                import traceback
                traceback.print_exc()
            # 每30秒检查一次（与速率限制的间隔一致）
            time.sleep(30)
    
    def _notify_frontend(self, message: dict):
        """通过 WebSocket 通知前端"""
        try:
            # 在新的事件循环中运行异步广播
            import asyncio
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(ws_manager.broadcast(message))
            loop.close()
        except Exception as e:
            print(f"WebSocket 通知失败: {e}")
    
    def _auto_process_emails_async(self):
        """自动处理所有待处理邮件（异步并发处理，与"处理全部"按钮逻辑一致）"""
        from src.nodes import Nodes
        from src.state import Email
        from concurrent.futures import as_completed
        
        # 重新获取用户状态（确保使用最新的数据）
        task_user_state = get_user_state(self.username)
        
        # 重置停止标志（确保之前的终止操作不会影响本次自动处理）
        user_lock = get_user_lock(self.username)
        with user_lock:
            task_user_state.stop_processing = False
            task_user_state.stopped_email_ids.clear()
            print(f"🔄 [自动处理] 重置停止标志，开始新的自动处理")
        
        pending_emails = [e for e in task_user_state.emails_cache if e.get('status') == 'pending']
        if not pending_emails:
            return None
        
        print(f"🚀 [自动处理] 开始处理 {len(pending_emails)} 封邮件，使用线程池并发处理")
        
        # 分类名称映射
        category_names = {
            'product_enquiry': '产品咨询',
            'customer_complaint': '客户投诉',
            'customer_feedback': '客户反馈',
            'unrelated': '无关邮件'
        }
        
        # 线程安全的计数器（使用锁保护）
        processed_count = 0
        failed_count = 0
        skipped_count = 0
        cancelled_count = 0  # 添加终止计数
        
        # 获取用户配置（所有邮件共享）
        try:
            email_address, auth_code = get_user_email_config(self.username)
            user_settings = get_user_settings(self.username)
            reply_model = user_settings.get("replyModel", user_settings.get("model", "moonshotai/Kimi-K2-Thinking"))
            embedding_model = user_settings.get("embeddingModel", "Qwen/Qwen3-Embedding-4B")
            models_config = get_models_config(self.username, reply_model, embedding_model)
            api_key = models_config["apiKey"]
            reply_api_base = models_config["replyApiBaseUrl"]
            embedding_api_base = models_config["embeddingApiBaseUrl"]
        except Exception as e:
            print(f"❌ [自动处理] 获取用户配置失败: {e}")
            import traceback
            traceback.print_exc()
            return {
                "message": f"自动处理失败: {str(e)}",
                "processed": 0,
                "skipped": 0,
                "failed": len(pending_emails)
            }
        
        def process_single_email(email):
            """处理单封邮件的函数（在线程池中并发执行）"""
            email_id = email.get('id', '')
            try:
                email['status'] = 'processing'
                print(f"📧 [自动处理] 开始处理邮件: {email.get('subject', '')[:50]}...")
                
                # 检查点1：处理开始前
                if task_user_state.stop_processing:
                    print(f"⏹️ [自动处理终止] 邮件 {email_id} 在处理开始前被终止")
                    with user_lock:
                        email['status'] = 'pending'
                        email['processing'] = False
                    # 发送WebSocket通知
                    self._notify_frontend({
                        "type": "email_process_stopped",
                        "email_id": email_id,
                        "message": "已终止处理"
                    })
                    return {'status': 'cancelled'}
                
                # 为每封邮件创建独立的Nodes实例（避免并发冲突）
                nodes = Nodes(
                    email_address=email_address, 
                    auth_code=auth_code, 
                    api_key=api_key,
                    reply_model=reply_model,
                    embedding_model=embedding_model,
                    signature=user_settings.get("signature"),
                    greeting=user_settings.get("greeting"),
                    closing=user_settings.get("closing"),
                    reply_api_base=reply_api_base,
                    embedding_api_base=embedding_api_base
                )
                
                # 创建Email对象
                email_obj = Email(
                    id=email.get('id', ''),
                    threadId=email.get('threadId', ''),
                    messageId=email.get('messageId', ''),
                    references=email.get('references', ''),
                    sender=email.get('sender', ''),
                    subject=email.get('subject', ''),
                    body=email.get('body', ''),
                    imap_id=email.get('imap_id', b'')
                )
                
                # 构建状态
                state = {
                    "emails": [email_obj],
                    "current_email": email_obj,
                    "email_category": None,
                    "rag_queries": [],
                    "retrieved_documents": "",
                    "generated_email": "",
                    "sendable": False,
                    "trials": 0,
                    "writer_messages": []
                }
                
                # 1. 分类邮件
                categorize_result = nodes.categorize_email(state)
                state.update(categorize_result)
                category = state.get('email_category', 'product_enquiry')
                category_label = category_names.get(category, category or '未分类')
                
                # 检查点2：分类后
                if task_user_state.stop_processing:
                    print(f"⏹️ [自动处理终止] 邮件 {email_id} 在分类后被终止")
                    with user_lock:
                        email['status'] = 'pending'
                        email['processing'] = False
                    # 发送WebSocket通知
                    self._notify_frontend({
                        "type": "email_process_stopped",
                        "email_id": email_id,
                        "message": "已终止处理"
                    })
                    return {'status': 'cancelled'}
                
                # 2. 检查是否是无关邮件
                if category == 'unrelated':
                    with user_lock:
                        email['status'] = 'skipped'
                        email['category'] = category
                        email['reply'] = '无关邮件，已跳过'
                        # 同步紧急程度信息（从Email对象获取）
                        if 'emails' in state and len(state['emails']) > 0:
                            email_obj = state['emails'][0]
                            if hasattr(email_obj, 'urgency_level'):
                                email['urgency_level'] = email_obj.urgency_level
                            if hasattr(email_obj, 'urgency_keywords'):
                                email['urgency_keywords'] = email_obj.urgency_keywords
                        task_user_state.stats['pending'] = max(0, task_user_state.stats['pending'] - 1)
                        task_user_state.history.insert(0, {
                            **email,
                            'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        })
                    
                    imap_id = email.get('imap_id')
                    if imap_id:
                        try:
                            nodes.email_tools.mark_email_as_read(imap_id)
                        except:
                            pass
                    
                    print(f"⏭️ [自动处理] 跳过无关邮件: {email.get('subject', '')[:50]}...")
                    
                    # 发送WebSocket通知
                    urgency_info = email.get('urgency_level', 'normal')
                    urgency_keywords = email.get('urgency_keywords', [])
                    self._notify_frontend({
                       "type": "email_process_complete",
                        "email_id": email_id,
                        "message": "无关邮件，已跳过",
                        "category": "unrelated",
                        "status": "skipped",
                        "reply": "无关邮件，已跳过",
                        "urgency_level": urgency_info,
                        "urgency_keywords": urgency_keywords
                    })
                    
                    # 生成摘要
                    email_body = email.get('body', '')
                    if email_body:
                        generate_email_summaries_async(self.username, email_id, email_body, '')
                    
                    return {'status': 'skipped'}
                
                # 3. RAG查询
                if category != 'unrelated':
                    # 检查点3：RAG查询前
                    if task_user_state.stop_processing:
                        print(f"⏹️ [自动处理终止] 邮件 {email_id} 在RAG查询前被终止")
                        with user_lock:
                            email['status'] = 'pending'
                            email['processing'] = False
                        # 发送WebSocket通知
                        self._notify_frontend({
                            "type": "email_process_stopped",
                            "email_id": email_id,
                            "message": "已终止处理"
                        })
                        return {'status': 'cancelled'}
                    
                    rag_query_result = nodes.construct_rag_queries(state)
                    state.update(rag_query_result)
                    
                    # 发送通知：显示生成的 RAG 查询问题
                    rag_queries = state.get('rag_queries', [])
                    if rag_queries:
                        self._notify_frontend({
                            "type": "rag_queries_generated",
                            "email_id": email_id,
                            "queries": rag_queries,
                            "count": len(rag_queries)
                        })
                    
                    rag_result = nodes.retrieve_from_rag(state)
                    state.update(rag_result)
                
                # 检查点4：RAG查询后
                if task_user_state.stop_processing:
                    print(f"⏹️ [自动处理终止] 邮件 {email_id} 在RAG查询后被终止")
                    with user_lock:
                        email['status'] = 'pending'
                        email['processing'] = False
                    # 发送WebSocket通知
                    self._notify_frontend({
                        "type": "email_process_stopped",
                        "email_id": email_id,
                        "message": "已终止处理"
                    })
                    return {'status': 'cancelled'}
                
                # 4. 编写回复邮件
                # 检查点5：开始编写回复前
                if task_user_state.stop_processing:
                    print(f"⏹️ [自动处理终止] 邮件 {email_id} 在开始编写回复前被终止")
                    with user_lock:
                        email['status'] = 'pending'
                        email['processing'] = False
                    # 发送WebSocket通知
                    self._notify_frontend({
                        "type": "email_process_stopped",
                        "email_id": email_id,
                        "message": "已终止处理"
                    })
                    return {'status': 'cancelled'}
                
                max_trials = 3
                for trial in range(max_trials):
                    # 检查点6：每次重试前
                    if task_user_state.stop_processing:
                        print(f"⏹️ [自动处理终止] 邮件 {email_id} 在编写回复前被终止（第{trial+1}次尝试）")
                        with user_lock:
                            email['status'] = 'pending'
                            email['processing'] = False
                        # 发送WebSocket通知
                        self._notify_frontend({
                            "type": "email_process_stopped",
                            "email_id": email_id,
                            "message": "已终止处理"
                        })
                        return {'status': 'cancelled'}
                    
                    write_result = nodes.write_draft_email(state)
                    state.update(write_result)
                    
                    # 检查点7：验证前
                    if task_user_state.stop_processing:
                        print(f"⏹️ [自动处理终止] 邮件 {email_id} 在验证前被终止（第{trial+1}次尝试）")
                        with user_lock:
                            email['status'] = 'pending'
                            email['processing'] = False
                        # 发送WebSocket通知
                        self._notify_frontend({
                            "type": "email_process_stopped",
                            "email_id": email_id,
                            "message": "已终止处理"
                        })
                        return {'status': 'cancelled'}
                    
                    verify_result = nodes.verify_generated_email(state)
                    state.update(verify_result)
                    
                    # 检查点8：验证后
                    if task_user_state.stop_processing:
                        print(f"⏹️ [自动处理终止] 邮件 {email_id} 在验证后被终止（第{trial+1}次尝试）")
                        with user_lock:
                            email['status'] = 'pending'
                            email['processing'] = False
                        # 发送WebSocket通知
                        self._notify_frontend({
                            "type": "email_process_stopped",
                            "email_id": email_id,
                            "message": "已终止处理"
                        })
                        return {'status': 'cancelled'}
                    
                    if state.get('sendable', False):
                        break
                
                # 5. 获取生成的回复
                generated_reply = state.get('generated_email', '')
                
                # 6. 检查是否自动发送
                auto_send = user_settings.get("autoSend", False)
                final_status = 'processed'
                if auto_send and generated_reply:
                    try:
                        result, message = send_reply_with_rate_limit(
                            self.username,
                            nodes.email_tools,
                            email_obj,
                            generated_reply,
                            email
                        )
                        if result:
                            final_status = 'sent'
                            sender_name = email.get('sender', '').split('@')[0] if '@' in email.get('sender', '') else email.get('sender', '未知')
                            with user_lock:
                                task_user_state.add_activity('primary', f'自动发送回复给: {sender_name}', 'Message')
                    except Exception as send_err:
                        print(f"❌ [自动处理] 自动发送回复时出错: {send_err}")
                
                # 7. 标记为已读
                imap_id = email.get('imap_id')
                if imap_id:
                    try:
                        nodes.email_tools.mark_email_as_read(imap_id)
                    except:
                        pass
                
                # 8. 使用锁保护状态更新
                with user_lock:
                    email['category'] = category
                    email['reply'] = generated_reply
                    email['status'] = final_status
                    email['rag_queries'] = state.get('rag_queries', [])  # 保存 RAG 查询问题
                    # 同步紧急程度信息（从Email对象获取）
                    if 'emails' in state and len(state['emails']) > 0:
                        email_obj = state['emails'][0]
                        if hasattr(email_obj, 'urgency_level'):
                            email['urgency_level'] = email_obj.urgency_level
                        if hasattr(email_obj, 'urgency_keywords'):
                            email['urgency_keywords'] = email_obj.urgency_keywords
                    task_user_state.stats['processed'] += 1
                    task_user_state.stats['pending'] = max(0, task_user_state.stats['pending'] - 1)
                    task_user_state.history.insert(0, {
                        **email,
                        'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    if not auto_send or not generated_reply or final_status != 'sent':
                        task_user_state.add_activity('success', f'处理了邮件: {category_label}', 'CircleCheck')
                
                print(f"✅ [自动处理] 邮件处理完成: {email.get('subject', '')[:50]}...")
                
                # 发送WebSocket通知
                urgency_info = email.get('urgency_level', 'normal')
                urgency_keywords = email.get('urgency_keywords', [])
                rag_queries = state.get('rag_queries', [])
                self._notify_frontend({
                    "type": "email_process_complete",
                    "email_id": email_id,
                    "message": f"{category_label} - 处理成功",
                    "category": category,
                    "status": final_status,
                    "reply": generated_reply,
                    "urgency_level": urgency_info,
                    "urgency_keywords": urgency_keywords,
                    "rag_queries": rag_queries
                })
                
                # 生成摘要
                email_body = email.get('body', '')
                if email_body or generated_reply:
                    generate_email_summaries_async(self.username, email_id, email_body, generated_reply or '')
                
                return {'status': 'processed'}
                
            except Exception as e:
                print(f"❌ [自动处理] 处理邮件错误: {email.get('subject', '')[:50]}... - {e}")
                import traceback
                traceback.print_exc()
                
                with user_lock:
                    email['status'] = 'failed'
                    task_user_state.stats['failed'] += 1
                    task_user_state.history.insert(0, {
                        **email,
                        'status': 'failed',
                        'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                
                # 发送WebSocket通知
                urgency_info = email.get('urgency_level', 'normal')
                urgency_keywords = email.get('urgency_keywords', [])
                self._notify_frontend({
                    "type": "email_process_complete",
                    "email_id": email_id,
                    "message": f"处理失败: {str(e)}",
                    "status": "failed",
                    "reply": None,
                    "urgency_level": urgency_info,
                    "urgency_keywords": urgency_keywords
                })
                
                return {'status': 'failed'}
        
        # 获取批量大小配置
        batch_size = user_settings.get("batchSize", 4)
        batch_size = max(1, min(30, int(batch_size)))
        
        # 使用独立的批量处理线程池
        batch_pool = get_or_create_batch_thread_pool(batch_size)
        
        total_batches = (len(pending_emails) + batch_size - 1) // batch_size
        print(f"📦 [自动处理] 将 {len(pending_emails)} 封邮件分成 {total_batches} 批，每批最多 {batch_size} 封")
        
        # 分批并发处理
        for batch_idx in range(total_batches):
            start_idx = batch_idx * batch_size
            end_idx = min(start_idx + batch_size, len(pending_emails))
            batch_emails = pending_emails[start_idx:end_idx]
            
            print(f"🔄 [自动处理] 处理第 {batch_idx + 1}/{total_batches} 批，包含 {len(batch_emails)} 封邮件")
            
            # 提交批次内的所有邮件到线程池
            future_to_email = {
                batch_pool.submit(process_single_email, email): email 
                for email in batch_emails
            }
            
            # 等待批次完成并收集结果
            for future in as_completed(future_to_email):
                try:
                    result = future.result()
                    with user_lock:
                        if result['status'] == 'processed':
                            processed_count += 1
                        elif result['status'] == 'skipped':
                            skipped_count += 1
                        elif result['status'] == 'cancelled':
                            cancelled_count += 1
                        elif result['status'] == 'failed':
                            failed_count += 1
                except Exception as e:
                    print(f"❌ [自动处理] 获取处理结果时出错: {e}")
                    with user_lock:
                        failed_count += 1
            
            print(f"✅ [自动处理] 第 {batch_idx + 1}/{total_batches} 批处理完成")
        
        # 保存数据
        with user_lock:
            save_user_email_data(self.username, task_user_state)
        
        print(f"🎉 [自动处理] 全部处理完成: {processed_count} 封成功, {skipped_count} 封跳过, {cancelled_count} 封终止, {failed_count} 封失败")
        
        # 发送完成通知（与批量处理保持一致）
        message = f"自动处理完成: {processed_count} 封成功, {skipped_count} 封跳过"
        if cancelled_count > 0:
            message += f", {cancelled_count} 封终止"
        if failed_count > 0:
            message += f", {failed_count} 封失败"
        
        self._notify_frontend({
            "type": "process_all_stopped",
            "message": message,
            "processed": processed_count,
            "skipped": skipped_count,
            "cancelled": cancelled_count,
            "failed": failed_count
        })
        
        return {
            "message": message,
            "processed": processed_count,
            "skipped": skipped_count,
            "cancelled": cancelled_count,
            "failed": failed_count
        }
    
    def _check_emails(self):
        """检查新邮件"""
        try:
            # 获取当前用户的邮箱配置
            if not self.username:
                raise ValueError("SystemState 未关联用户名，无法获取邮箱配置")
            email_address, auth_code = get_user_email_config(self.username)
            email_tools = QQEmailToolsClass(email_address=email_address, auth_code=auth_code)
            # 获取所有未读邮件（不限制数量，默认最多50封，但可以通过参数调整）
            emails = email_tools.fetch_unanswered_emails(max_results=100)
            self.last_check_time = datetime.now().isoformat()
            
            # 获取当前未读邮件的ID列表
            current_unread_ids = {email_data.get('id', '') for email_data in emails}
            
            # 移除缓存中已经在QQ邮箱中被标记为已读的邮件
            # 但保留已处理、已跳过、已发送的邮件（这些是我们主动标记已读的）
            emails_to_remove = []
            for cached_email in self.emails_cache:
                cached_id = cached_email.get('id', '')
                cached_status = cached_email.get('status', '')
                # 只移除状态为 pending 或 read 且不在未读列表中的邮件
                # 保留 processed、skipped、sent、failed 状态的邮件
                if cached_id not in current_unread_ids and cached_status in ['pending', 'read']:
                    emails_to_remove.append(cached_email)
            
            for email_to_remove in emails_to_remove:
                self.emails_cache.remove(email_to_remove)
                if email_to_remove.get('status') == 'pending':
                    self.stats['pending'] = max(0, self.stats['pending'] - 1)
                print(f"邮件已从QQ邮箱移除，同步清除缓存: {email_to_remove.get('subject', '')}")
            
            # 添加新邮件到缓存
            new_count = 0
            for email_data in emails:
                email_id = email_data.get('id', '')
                if not any(e.get('id') == email_id for e in self.emails_cache):
                    # 自动分类邮件
                    subject = email_data.get('subject', '')
                    body = email_data.get('body', '')
                    category = auto_classify_email(subject, body)
                    
                    # 检测邮件紧急程度
                    try:
                        urgency_level, urgency_keywords = analyze_email_urgency(subject, body)
                    except Exception as e:
                        print(f"⚠️ 紧急程度检测失败: {str(e)}")
                        urgency_level = 'low'
                        urgency_keywords = []
                    
                    # 使用邮件的实际接收时间（如果存在），否则使用当前时间
                    email_time = email_data.get('date', '') or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    # 如果时间格式不完整，补充秒数
                    if len(email_time) < 19:  # 'YYYY-MM-DD HH:MM:SS' 应该是19个字符
                        email_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    self.emails_cache.append({
                        **email_data,
                        'time': email_time,
                        'status': 'pending',
                        'category': category,
                        'reply': None,
                        'preview': body[:100] + '...',
                        'urgency_level': urgency_level,
                        'urgency_keywords': urgency_keywords
                    })
                    
                    # 判断是否是今天的邮件
                    email_date = email_time[:10] if len(email_time) >= 10 else ''
                    today = datetime.now().strftime('%Y-%m-%d')
                    is_today = email_date == today
                    
                    if is_today:
                        self.stats['today_emails'] += 1
                        print(f"自动检查：添加新邮件（今日）: {subject[:50]}... (时间: {email_time}, 日期: {email_date})")
                    else:
                        print(f"自动检查：添加新邮件（非今日）: {subject[:50]}... (时间: {email_time}, 日期: {email_date}, 今天: {today})")
                    
                    self.stats['pending'] += 1
                    new_count += 1
            
            return new_count
                    
        except Exception as e:
            print(f"检查邮件错误: {e}")
            return 0

# 全局状态实例（按用户隔离）
# 格式: {username: SystemState实例}
user_states: dict[str, SystemState] = {}

def get_user_email_data_file(username: str, reload: bool = False) -> str:
    """获取用户邮件数据文件路径（使用user_id而不是username）
    
    @param username: 用户名
    @param reload: 是否强制重新加载数据（默认False，使用内存中的数据）
    """
    user_id = get_user_id_by_username(username, reload=reload)
    if user_id:
        return os.path.join(USER_DATA_DIR, f"user_email_data_{user_id}.json")
    # 兼容旧数据：如果找不到user_id，使用username（向后兼容）
    return os.path.join(USER_DATA_DIR, f"user_email_data_{username}.json")

def load_user_email_data(username: str) -> dict:
    """从文件加载用户的邮件数据（emails_cache, history, activities, stats）
    
    注意：此函数会先尝试通过映射关系找到实际用户名，然后使用user_id加载数据文件
    这样确保即使用户名改变，只要user_id不变，数据就能正确加载
    """
    # 首先检查用户名是否已迁移，获取当前有效的用户名
    actual_username = get_current_username(username)
    if actual_username != username:
        print(f"信息: load_user_email_data 检测到用户名 {username} 已迁移到 {actual_username}，使用新用户名")
        username = actual_username
    
    # 首先尝试使用user_id命名的文件（新格式）
    # 注意：这里使用 reload=False，优先使用内存中的数据，避免覆盖正在进行的修改
    data_file = get_user_email_data_file(username, reload=False)
    print(f"🔍 [加载数据] 用户 {username}，尝试加载文件: {data_file}")
    
    if os.path.exists(data_file):
        try:
            with open(data_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            print(f"✓ [加载数据] 成功加载用户 {username} 的数据文件: {data_file}")
            print(f"   邮件数: {len(data.get('emails_cache', []))}, 历史记录数: {len(data.get('history', []))}")
            return {
                "emails_cache": data.get("emails_cache", []),
                "history": data.get("history", []),
                "activities": data.get("activities", []),
                "stats": data.get("stats", {
                    "today_emails": 0,
                    "processed": 0,
                    "pending": 0,
                    "failed": 0
                }),
                "last_check_time": data.get("last_check_time"),
                "is_running": data.get("is_running", False),
                "auto_process": data.get("auto_process", False),
                "check_interval": data.get("check_interval", 900)
            }
        except Exception as e:
            print(f"❌ [加载数据] 加载用户 {username} 邮件数据失败: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    # 兼容旧数据：如果新格式文件不存在，尝试使用username命名的旧文件
    # 注意：这里需要检查旧用户名和新用户名对应的旧文件
    old_data_files = [
        os.path.join(USER_DATA_DIR, f"user_email_data_{username}.json"),  # 当前用户名对应的旧文件
    ]
    
    # 如果用户名已迁移，也检查旧用户名对应的文件
    if actual_username != username:
        old_data_files.append(os.path.join(USER_DATA_DIR, f"user_email_data_{username}.json"))
    
    # 也检查根目录的旧文件（向后兼容）
    old_data_files.extend([
        f"user_email_data_{username}.json",
        f"user_email_data_{actual_username}.json" if actual_username != username else None
    ])
    old_data_files = [f for f in old_data_files if f is not None]
    
    for old_data_file in old_data_files:
        if os.path.exists(old_data_file):
            try:
                print(f"🔍 [加载数据] 检测到旧格式数据文件 {old_data_file}，正在迁移...")
                with open(old_data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # 获取user_id并迁移到新格式
                user_id = get_user_id_by_username(username, reload=True)  # 迁移时需要重新加载
                if user_id:
                    new_data_file = os.path.join(USER_DATA_DIR, f"user_email_data_{user_id}.json")
                    # 保存到新格式文件
                    with open(new_data_file, 'w', encoding='utf-8') as f:
                        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
                    # 删除旧文件
                    try:
                        os.remove(old_data_file)
                        print(f"✓ [加载数据] 数据文件已从旧格式迁移到新格式: {old_data_file} -> {new_data_file}")
                    except Exception as e:
                        print(f"⚠️ [加载数据] 删除旧数据文件失败: {e}")
                
                return {
                    "emails_cache": data.get("emails_cache", []),
                    "history": data.get("history", []),
                    "activities": data.get("activities", []),
                    "stats": data.get("stats", {
                        "today_emails": 0,
                        "processed": 0,
                        "pending": 0,
                        "failed": 0
                    }),
                    "last_check_time": data.get("last_check_time"),
                    "is_running": data.get("is_running", False),
                    "auto_process": data.get("auto_process", False),
                    "check_interval": data.get("check_interval", 900)
                }
            except Exception as e:
                print(f"❌ [加载数据] 加载用户 {username} 旧格式邮件数据失败: {e}")
                continue
    
    print(f"⚠️ [加载数据] 用户 {username} 的数据文件不存在")
    return None

def save_user_email_data(username: str, user_state: SystemState):
    """保存用户的邮件数据到文件
    注意：此函数会通过用户名找到对应的user_id，然后使用user_id命名文件
    这样确保即使用户名改变，只要user_id不变，数据文件就不会改变
    """
    # 先获取实际用户名（处理用户名映射）
    actual_username = get_current_username(username, reload=True)
    if actual_username != username:
        print(f"信息: save_user_email_data 检测到用户名 {username} 已迁移到 {actual_username}，使用新用户名")
        username = actual_username
    
    data_file = get_user_email_data_file(username, reload=False)
    print(f"💾 [保存数据] 用户 {username}，保存到文件: {data_file}")
    try:
        # 准备要保存的数据（排除不能序列化的对象，如线程）
        data = {
            "emails_cache": user_state.emails_cache,
            "history": user_state.history,
            "activities": user_state.activities,
            "stats": user_state.stats,
            "last_check_time": user_state.last_check_time,
            "is_running": False,  # 不保存运行状态，重启后需要重新启动
            "auto_process": user_state.auto_process,
            "check_interval": user_state.check_interval,
            "last_save_time": datetime.now().isoformat()
        }
        with open(data_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        print(f"保存用户 {username} 邮件数据失败: {e}")

# 自动保存装饰器（用于在关键操作后自动保存）
def auto_save_email_data(func):
    """装饰器：在函数执行后自动保存用户邮件数据"""
    async def wrapper(*args, **kwargs):
        result = await func(*args, **kwargs)
        # 尝试从参数中获取 current_username
        current_username = None
        for arg in args:
            if isinstance(arg, str) and arg in user_states:
                current_username = arg
                break
        if 'current_username' in kwargs:
            current_username = kwargs['current_username']
        # 如果找到了用户名，保存数据
        if current_username and current_username in user_states:
            save_user_email_data(current_username, user_states[current_username])
        return result
    return wrapper

def get_user_state(username: str, check_auto_start: bool = True) -> SystemState:
    """
    获取指定用户的系统状态，如果不存在则创建，并从文件加载数据
    
    @param username: 用户名（可能是旧用户名或新用户名）
    @param check_auto_start: 是否检查 autoStart 设置并自动启动监控（默认 True）
    """
    # 首先检查用户名是否已迁移，获取当前有效的用户名
    actual_username = get_current_username(username)
    
    # 如果用户名已迁移，使用新用户名
    if actual_username != username:
        print(f"信息: get_user_state 检测到用户名 {username} 已迁移到 {actual_username}，使用新用户名")
        username = actual_username
    
    if username not in user_states:
        user_state = SystemState(username=username)
        # 尝试从文件加载之前保存的数据
        saved_data = load_user_email_data(username)
        if saved_data:
            user_state.emails_cache = saved_data.get("emails_cache", [])
            user_state.history = saved_data.get("history", [])
            user_state.activities = saved_data.get("activities", [])
            user_state.stats = saved_data.get("stats", {
                "today_emails": 0,
                "processed": 0,
                "pending": 0,
                "failed": 0,
                "sent": 0  # 发送回复数
            })
            # 确保有sent字段（兼容旧数据）
            if "sent" not in user_state.stats:
                user_state.stats["sent"] = 0
            user_state.last_check_time = saved_data.get("last_check_time")
            user_state.auto_process = saved_data.get("auto_process", False)
            user_state.check_interval = saved_data.get("check_interval", 900)
            print(f"已加载用户 {username} 的邮件数据: {len(user_state.emails_cache)} 封邮件, {len(user_state.history)} 条历史记录")
        else:
            print(f"警告: 用户 {username} 的邮件数据文件不存在或为空，使用空数据")
        user_states[username] = user_state
    
    return user_states[username]

def check_and_start_monitor_if_needed(username: str):
    """
    检查用户的 autoStart 设置，如果为 True 且系统未运行，自动启动监控
    这个函数可以在多个地方调用，确保开启"自动运行"后，系统会自动启动监控
    """
    try:
        # 获取用户设置
        user_settings = get_user_settings(username)
        auto_start = user_settings.get("autoStart", False)
        print(f"[自动启动检查] 用户 {username} 的 autoStart 设置: {auto_start}")
        
        if auto_start:
            # 使用 check_auto_start=False 避免无限递归
            user_state = get_user_state(username, check_auto_start=False)
            print(f"[自动启动检查] 用户 {username} 的监控状态: is_running={user_state.is_running}")
            if not user_state.is_running:
                try:
                    # 检查邮箱配置是否完整
                    get_user_email_config(username)
                    # 保存当前的 auto_process 值，确保不会被修改
                    original_auto_process = user_state.auto_process
                    # 使用 start_monitor 方法启动监控
                    user_state.start_monitor()
                    # 确保 auto_process 的值没有被意外修改
                    if user_state.auto_process != original_auto_process:
                        print(f"警告：auto_process 值被意外修改，恢复原值")
                        user_state.auto_process = original_auto_process
                    # 记录操作
                    user_state.add_activity('success', '自动启动了邮件监控（autoStart=True）', 'VideoPlay')
                    print(f"用户 {username} 的监控已自动启动（autoStart=True，在检查时触发）")
                except ValueError as e:
                    # 邮箱配置不完整，不自动启动
                    print(f"用户 {username} 的邮箱配置不完整，无法自动启动监控: {e}")
                except Exception as e:
                    print(f"自动启动监控失败: {e}")
    except Exception as e:
        print(f"检查并启动监控时出错: {e}")

def get_user_email_config(username: str) -> tuple[str, str]:
    """
    获取用户的邮箱配置（邮箱地址和授权码）
    
    @param username: 用户名
    @return: (email_address, auth_code) 元组
    @raises: ValueError 如果用户没有配置邮箱
    """
    global user_data
    user_data = load_user_data()
    
    if username not in user_data:
        raise ValueError(f"用户 {username} 不存在")
    
    user_info = user_data[username]
    email = user_info.get("email", "")
    email_auth_code = user_info.get("emailAuthCode", "")
    
    if not email:
        raise ValueError("您尚未配置邮箱地址。请前往【系统设置】页面配置QQ邮箱地址和授权码。")
    
    if not email_auth_code:
        raise ValueError(f"您已配置邮箱地址（{email}），但尚未配置邮箱授权码。请前往【系统设置】页面配置QQ邮箱授权码。授权码获取方法：登录QQ邮箱网页版 -> 设置 -> 账户 -> 开启IMAP/SMTP服务 -> 生成授权码。")
    
    return (email, email_auth_code)

# 为了向后兼容，保留一个默认的system_state（用于没有用户上下文的场景）
system_state = SystemState()

# ==================== WebSocket API ====================

# ==================== 认证API ====================

# 简单的用户数据存储（实际应用中应该使用数据库）
USER_DATA_DIR = "data/users"  # 用户数据目录
# 确保用户数据目录存在
os.makedirs(USER_DATA_DIR, exist_ok=True)
USER_DATA_FILE = os.path.join(USER_DATA_DIR, "user_data.json")
USERNAME_MAPPING_FILE = os.path.join(USER_DATA_DIR, "username_mapping.json")  # 记录用户名迁移映射：{旧用户名: 新用户名}

def get_user_id_by_username(username: str, reload: bool = False) -> Optional[str]:
    """通过用户名获取用户的唯一ID（UUID）
    
    @param username: 用户名（可能是旧用户名或新用户名）
    @param reload: 是否强制重新加载数据（默认False，使用内存中的数据）
    """
    global user_data
    
    # 只有在需要时才重新加载数据，避免覆盖内存中的修改
    if reload or 'user_data' not in globals() or not user_data:
        user_data = load_user_data()
    
    # 首先检查用户名是否直接存在
    if username not in user_data:
        # 检查是否有映射关系（用户名已迁移）
        actual_username = get_current_username(username)
        if actual_username != username and actual_username in user_data:
            print(f"信息: get_user_id_by_username 检测到用户名 {username} 已迁移到 {actual_username}")
            username = actual_username
        else:
            print(f"警告: get_user_id_by_username 无法找到用户 {username}，且无映射关系")
            return None
    
    user_info = user_data[username]
    
    # 如果用户没有user_id，为其生成一个（兼容旧数据）
    if "user_id" not in user_info:
        user_info["user_id"] = str(uuid.uuid4())
        save_user_data(user_data)
        print(f"为用户 {username} 生成新的 user_id: {user_info['user_id']}")
    
    return user_info["user_id"]

def get_username_by_user_id(user_id: str) -> Optional[str]:
    """通过用户ID获取用户名"""
    global user_data
    user_data = load_user_data()
    
    for username, user_info in user_data.items():
        if user_info.get("user_id") == user_id:
            return username
    
    return None

def load_user_data():
    """从文件加载用户数据"""
    if os.path.exists(USER_DATA_FILE):
        try:
            with open(USER_DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 验证数据格式：应该是字典
                if isinstance(data, dict):
                    # 修复或跳过有问题的用户数据，而不是直接返回默认数据
                    valid_data = {}
                    needs_save = False
                    
                    for username, user_info in data.items():
                        if not isinstance(user_info, dict):
                            print(f"警告: {USER_DATA_FILE} 用户 {username} 数据格式不正确，跳过该用户")
                            continue
                        
                        # 修复缺失的字段
                        if "password" not in user_info:
                            print(f"警告: {USER_DATA_FILE} 用户 {username} 缺少密码字段，使用默认密码")
                            user_info["password"] = "123456"  # 默认密码，用户需要修改
                            needs_save = True
                        
                        # 确保有必要的字段
                        if "devices" not in user_info:
                            user_info["devices"] = []
                            needs_save = True
                        
                        if "preferences" not in user_info:
                            user_info["preferences"] = {
                                "theme": "light",
                                "notification": True,
                                "sound": False,
                                "defaultPage": "/dashboard"
                            }
                            needs_save = True
                        
                        # 为旧用户生成user_id（兼容旧数据）
                        if "user_id" not in user_info:
                            user_info["user_id"] = str(uuid.uuid4())
                            needs_save = True
                            print(f"为用户 {username} 生成新的 user_id: {user_info['user_id']}")
                        
                        valid_data[username] = user_info
                    
                    # 如果有修复，保存回文件
                    if needs_save:
                        print(f"修复用户数据，保存到 {USER_DATA_FILE}")
                        save_user_data(valid_data)
                    
                    # 如果修复后没有有效用户，返回默认数据
                    if not valid_data:
                        print(f"警告: {USER_DATA_FILE} 没有有效用户数据，使用默认数据")
                        return {
                            "admin": {
                                "password": "admin123",
                                "devices": [],
                                "preferences": {
                                    "theme": "light",
                                    "notification": True,
                                    "sound": False,
                                    "defaultPage": "/dashboard"
                                }
                            }
                        }
                    
                    return valid_data
                else:
                    print(f"警告: {USER_DATA_FILE} 数据格式不正确（不是字典），使用默认数据")
        except json.JSONDecodeError as e:
            print(f"警告: {USER_DATA_FILE} JSON解析失败: {e}，使用默认数据")
        except Exception as e:
            print(f"警告: 加载 {USER_DATA_FILE} 失败: {e}，使用默认数据")
    # 默认数据
    return {
        "admin": {
            "password": "admin123",  # 实际应用中应该加密存储
            "devices": [],
            "preferences": {
                "theme": "light",
                "notification": True,
                "sound": False,
                "defaultPage": "/dashboard"
            }
        }
    }

def save_user_data(data):
    """保存用户数据到文件"""
    try:
        # 先写入临时文件，然后重命名，确保原子性
        temp_file = USER_DATA_FILE + ".tmp"
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        # 原子性地替换原文件
        if os.path.exists(USER_DATA_FILE):
            os.replace(temp_file, USER_DATA_FILE)
        else:
            os.rename(temp_file, USER_DATA_FILE)
        print(f"✓ 用户数据已成功保存到 {USER_DATA_FILE}")
    except Exception as e:
        print(f"❌ 保存用户数据失败: {e}")
        import traceback
        traceback.print_exc()
        # 如果保存失败，尝试删除临时文件
        temp_file = USER_DATA_FILE + ".tmp"
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass
        raise  # 重新抛出异常，让调用者知道保存失败

def load_username_mapping():
    """加载用户名映射关系"""
    if os.path.exists(USERNAME_MAPPING_FILE):
        try:
            with open(USERNAME_MAPPING_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载用户名映射失败: {e}")
    return {}

def save_username_mapping(mapping):
    """保存用户名映射关系"""
    try:
        with open(USERNAME_MAPPING_FILE, 'w', encoding='utf-8') as f:
            json.dump(mapping, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"保存用户名映射失败: {e}")

def get_current_username(username, reload: bool = True):
    """根据用户名（可能是旧用户名）获取当前用户名
    如果用户名已迁移，返回新用户名；否则返回原用户名
    
    @param username: 用户名（可能是旧用户名）
    @param reload: 是否重新加载数据（默认True，确保使用最新数据）
    """
    global user_data
    
    # 重新加载数据，确保使用最新数据（特别是在用户名刚修改后）
    if reload:
        user_data = load_user_data()
    
    # 先检查用户名是否直接存在
    if username in user_data:
        return username
    
    # 如果不存在，检查映射关系
    mapping = load_username_mapping()
    # 查找映射链（可能有多层迁移：A -> B -> C）
    current = username
    visited = set()  # 防止循环引用
    
    while current in mapping and current not in visited:
        visited.add(current)
        current = mapping[current]
        # 如果新用户名存在，返回它
        if current in user_data:
            return current
    
    # 如果映射链最终指向的用户名也不存在，返回原用户名
    return username

# 加载用户数据
user_data = load_user_data()

# 辅助函数：从请求头获取当前用户名
def get_username_from_request(authorization: Optional[str] = Header(None, alias="Authorization")) -> str:
    """从请求头获取当前登录用户名"""
    global user_data
    user_data = load_user_data()
    
    # 如果token中包含用户名信息，解析它
    # token格式: "token-{timestamp}-{username}" 或 "token-{timestamp}"
    if authorization:
        # 移除 "Bearer " 前缀（不区分大小写）
        token = authorization.replace("Bearer ", "").replace("bearer ", "").strip()
        
        # 尝试从token中解析用户名（如果token格式包含用户名）
        if token.startswith("token-") and "-" in token:
            parts = token.split("-", 2)  # 只分割前两部分，保留用户名部分（可能包含-）
            if len(parts) >= 3:
                # token格式: token-timestamp-username
                potential_username = parts[2]  # 第三部分就是用户名
                if potential_username in user_data:
                    return potential_username
                else:
                    # 如果token中的用户名不存在，检查是否有映射关系（用户名已迁移）
                    current_username = get_current_username(potential_username)
                    if current_username != potential_username and current_username in user_data:
                        print(f"信息: token中的用户名 {potential_username} 已迁移到 {current_username}")
                        return current_username
                    else:
                        print(f"警告: token中的用户名 {potential_username} 不在用户数据中，且无映射关系")
    
    # 如果无法从token解析，查找当前登录用户（有当前设备的）
    # 优先返回有当前设备的用户
    for username, user_info in user_data.items():
        devices = user_info.get("devices", [])
        # 如果有当前设备，说明是当前登录用户
        if devices and any(d.get("current", False) for d in devices):
            return username
    
    # 如果没有找到，返回user_data中的第一个用户（如果有的话）
    if user_data:
        first_username = list(user_data.keys())[0]
        print(f"警告: 无法从token解析用户名，使用第一个用户: {first_username}")
        return first_username
    
    # 最后返回admin（默认用户）
    print("警告: 没有找到任何用户，使用默认用户: admin")
    return "admin"

import asyncio
from fastapi import WebSocket, WebSocketDisconnect, Query
from typing import Optional, Dict, Set

# WebSocket 连接管理器（按用户名分组连接）
class ConnectionManager:
    def __init__(self):
        # username -> set(WebSocket)
        self.active_connections: Dict[str, Set[WebSocket]] = {}
        self._lock = threading.Lock()

    async def connect(self, websocket: WebSocket, token: Optional[str] = None) -> str:
        # 接受连接后解析用户名并加入映射
        await websocket.accept()
        try:
            username = get_username_from_request(token) if token else get_username_from_request()
        except Exception:
            # 回退为默认用户名（避免抛出错误阻断连接）
            username = "admin"
        with self._lock:
            conns = self.active_connections.setdefault(username, set())
            conns.add(websocket)
        print(f"🔌 [WS] 用户 {username} 已连接 (当前连接数: {len(self.active_connections.get(username, []))})")
        return username

    def disconnect(self, websocket: WebSocket):
        with self._lock:
            for user, conns in list(self.active_connections.items()):
                if websocket in conns:
                    conns.remove(websocket)
                    if len(conns) == 0:
                        del self.active_connections[user]
                    print(f"🔌 [WS] 断开连接: {user}")
                    return

    async def send_message_to_user(self, username: str, message: dict):
        """向指定用户的所有连接发送 JSON 消息"""
        conns = []
        with self._lock:
            if username in self.active_connections:
                conns = list(self.active_connections[username])
        
        if not conns:
            print(f"⚠️ [WS] 用户 {username} 没有活跃的 WebSocket 连接")
            return
        
        sent_count = 0
        for ws in conns:
            try:
                await ws.send_json(message)
                sent_count += 1
            except Exception as e:
                print(f"⚠️ [WS] 发送消息失败给 {username}: {e}")
        
        print(f"✅ [WS] 已向用户 {username} 的 {sent_count}/{len(conns)} 个连接发送消息: {message.get('type', 'unknown')}")
    
    async def broadcast(self, message: dict):
        """向所有连接的客户端广播消息（兼容旧代码）"""
        print(f"[WebSocket广播] 开始广播消息: {message.get('type', 'unknown')}")
        all_conns = []
        with self._lock:
            for conns in self.active_connections.values():
                all_conns.extend(list(conns))
        
        print(f"  - 当前连接数: {len(all_conns)}")
        sent_count = 0
        for ws in all_conns:
            try:
                await ws.send_json(message)
                sent_count += 1
            except Exception as e:
                print(f"[WebSocket广播] 发送失败: {e}")
        
        print(f"[WebSocket广播] 广播完成: 成功发送到 {sent_count}/{len(all_conns)} 个连接")


# 全局 manager，用于在其他模块/线程中推送
ws_manager = ConnectionManager()
# ==================== WebSocket 连接 ====================

@app.websocket("/api/ws")
async def websocket_endpoint(websocket: WebSocket, token: Optional[str] = Query(None)):
    """
    简单的 WebSocket 入口，使用 query 参数 token 解析用户名（与 HTTP 的 token 解析一致）
    前端连接示例: ws://localhost:8000/api/ws?token=token-xxxxx-username
    """
    username = await ws_manager.connect(websocket, token)
    try:
        while True:
            # 保持简单协议：客户端可以发送心跳或任意文本（服务器目前不需要处理）
            await websocket.receive_text()
    except WebSocketDisconnect:
        ws_manager.disconnect(websocket)
    except Exception as e:
        print(f"⚠️ [WS] 连接异常: {e}")
        ws_manager.disconnect(websocket)

class ChangePasswordRequest(BaseModel):
    oldPassword: str
    newPassword: str

class ForgotPasswordRequest(BaseModel):
    username: str
    email: str

class ResetPasswordRequest(BaseModel):
    username: str
    email: str
    newPassword: str

class UpdateProfileRequest(BaseModel):
    username: str

class DeviceInfo(BaseModel):
    device: str
    browser: str
    ip: str
    time: str
    current: bool = False

@app.get("/api/auth/check-username")
async def check_username(username: str):
    """检查用户名是否已存在"""
    username = username.strip()
    
    # 重新加载用户数据
    global user_data
    user_data = load_user_data()
    
    # 基本格式验证
    if len(username) < 2:
        return {"available": False, "message": "用户名长度至少2位"}
    
    if len(username) > 20:
        return {"available": False, "message": "用户名长度不能超过20位"}
    
    # 验证用户名格式（字母、数字、下划线、中文）
    if not re.match(r'^[a-zA-Z0-9_\u4e00-\u9fa5]+$', username):
        return {"available": False, "message": "用户名只能包含字母、数字、下划线和中文"}
    
    # 检查用户名是否已存在
    if username in user_data:
        return {"available": False, "message": "用户名已存在"}
    
    return {"available": True, "message": "用户名可用"}

@app.post("/api/auth/register")
async def register(request: RegisterRequest):
    """用户注册"""
    username = request.username.strip()
    password = request.password
    email = request.email.strip() if request.email else ""
    
    # 重新加载用户数据
    global user_data
    user_data = load_user_data()
    
    # 验证用户名
    if len(username) < 2:
        raise HTTPException(status_code=400, detail="用户名长度至少2位")
    
    if len(username) > 20:
        raise HTTPException(status_code=400, detail="用户名长度不能超过20位")
    
    # 验证用户名格式（字母、数字、下划线、中文）
    if not re.match(r'^[a-zA-Z0-9_\u4e00-\u9fa5]+$', username):
        raise HTTPException(status_code=400, detail="用户名只能包含字母、数字、下划线和中文")
    
    # 验证密码
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="密码长度至少6位")
    
    if len(password) > 50:
        raise HTTPException(status_code=400, detail="密码长度不能超过50位")
    
    # 验证邮箱（可选，但如果有邮箱则必须是QQ邮箱）
    if email:
        # 验证邮箱格式
        email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not re.match(email_regex, email):
            raise HTTPException(status_code=400, detail="邮箱格式不正确")
        
        # 验证必须是QQ邮箱
        if not email.endswith('@qq.com'):
            raise HTTPException(status_code=400, detail="本系统仅支持QQ邮箱，请输入QQ邮箱地址")
    
    # 检查用户名是否已存在
    if username in user_data:
        raise HTTPException(status_code=400, detail="用户名已存在")
    
    # 创建新用户，分配唯一的user_id
    user_data[username] = {
        "password": password,  # 实际应用中应该加密存储
        "user_id": str(uuid.uuid4()),  # 为每个用户分配唯一的UUID
        "devices": [],
        "email": email,
        "registerTime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "preferences": {
            "theme": "light",
            "notification": True,
            "sound": False,
            "defaultPage": "/dashboard"
        }
    }
    
    # 保存到文件
    save_user_data(user_data)
    
    return {
        "message": "注册成功，请登录",
        "username": username
    }

def parse_user_agent(user_agent: str) -> dict:
    """解析User-Agent字符串，提取设备和浏览器信息"""
    if not user_agent:
        return {"device": "Unknown Device", "browser": "Unknown Browser"}
    
    user_agent_lower = user_agent.lower()
    
    # 检测操作系统
    device = "Unknown Device"
    if "windows" in user_agent_lower:
        if "phone" in user_agent_lower:
            device = "Windows Phone"
        else:
            device = "Windows PC"
    elif "mac" in user_agent_lower or "darwin" in user_agent_lower:
        if "iphone" in user_agent_lower or "ipad" in user_agent_lower:
            device = "iPhone" if "iphone" in user_agent_lower else "iPad"
        else:
            device = "Mac"
    elif "linux" in user_agent_lower:
        device = "Linux"
    elif "android" in user_agent_lower:
        device = "Android"
    elif "iphone" in user_agent_lower:
        device = "iPhone"
    elif "ipad" in user_agent_lower:
        device = "iPad"
    
    # 检测浏览器
    browser = "Unknown Browser"
    if "edg" in user_agent_lower:
        browser = "Edge"
    elif "chrome" in user_agent_lower and "edg" not in user_agent_lower:
        # 提取Chrome版本号
        import re
        chrome_match = re.search(r'chrome/(\d+)', user_agent_lower)
        if chrome_match:
            browser = f"Chrome {chrome_match.group(1)}"
        else:
            browser = "Chrome"
    elif "firefox" in user_agent_lower:
        firefox_match = re.search(r'firefox/(\d+)', user_agent_lower)
        if firefox_match:
            browser = f"Firefox {firefox_match.group(1)}"
        else:
            browser = "Firefox"
    elif "safari" in user_agent_lower and "chrome" not in user_agent_lower:
        safari_match = re.search(r'version/(\d+)', user_agent_lower)
        if safari_match:
            browser = f"Safari {safari_match.group(1)}"
        else:
            browser = "Safari"
    elif "opera" in user_agent_lower or "opr" in user_agent_lower:
        browser = "Opera"
    
    return {"device": device, "browser": browser}

@app.post("/api/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest, http_request: Request):
    """用户登录"""
    username = request.username.strip()
    password = request.password
    
    # 重新加载用户数据，确保使用最新的密码
    global user_data
    user_data = load_user_data()
    
    # 重要：登录时不使用映射关系，只允许当前有效的用户名登录
    # 如果用户名已迁移，旧用户名不能登录（但可以注册）
    # 这样确保修改用户名后，旧用户名不能登录，但可以被其他人注册
    if username not in user_data:
        # 检查是否有映射关系（仅用于提示用户）
        mapping = load_username_mapping()
        if username in mapping:
            new_username = mapping[username]
            if new_username in user_data:
                raise HTTPException(
                    status_code=401, 
                    detail=f"用户名已更改，请使用新用户名 '{new_username}' 登录"
                )
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    
    # 验证用户名和密码（只验证当前有效的用户名）
    if user_data[username]["password"] == password:
        # 从请求头获取User-Agent（FastAPI的headers是大小写不敏感的，但使用小写更安全）
        # 尝试多种可能的键名
        user_agent = (
            http_request.headers.get("user-agent") or 
            http_request.headers.get("User-Agent") or 
            http_request.headers.get("USER-AGENT") or
            ""
        )
        
        # 调试：打印User-Agent（仅在开发环境）
        if not user_agent:
            print(f"[WARNING] 未获取到User-Agent，请求头: {list(http_request.headers.keys())}")
        else:
            print(f"[DEBUG] 获取到User-Agent: {user_agent[:100]}...")  # 只打印前100个字符
        
        # 从请求头获取客户端IP（优先使用X-Forwarded-For，然后是X-Real-IP，最后是client.host）
        client_ip = (
            http_request.headers.get("x-forwarded-for", "").split(",")[0].strip() or
            http_request.headers.get("x-real-ip", "") or
            (http_request.client.host if http_request.client else "127.0.0.1")
        )
        
        # 解析User-Agent获取设备和浏览器信息
        device_browser = parse_user_agent(user_agent)
        
        print(f"[DEBUG] 解析结果 - 设备: {device_browser['device']}, 浏览器: {device_browser['browser']}, IP: {client_ip}")
        
        device_info = {
            "device": device_browser["device"],
            "browser": device_browser["browser"],
            "ip": client_ip,
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "current": True
        }
        
        # 更新设备列表（只保留最近5台设备）
        devices = user_data[username].get("devices", [])
        # 将当前设备标记为非当前
        for d in devices:
            d["current"] = False
        devices.insert(0, device_info)
        user_data[username]["devices"] = devices[:5]
        
        # 更新最后登录时间
        user_data[username]["lastLogin"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 保存设备信息
        save_user_data(user_data)
        
        # token中包含用户名信息，格式: token-timestamp-username
        # 注意：这里使用 actual_username（可能是映射后的新用户名），确保 token 中包含正确的用户名
        token = f"token-{datetime.now().timestamp()}-{username}"
        
        # 登录成功后，确保用户状态已正确加载（使用映射后的用户名）
        # 这样可以确保即使用户输入的是旧用户名，也能正确加载数据
        try:
            user_state = get_user_state(username, check_auto_start=False)
            print(f"[登录] 用户 {username} 登录成功，已加载状态: {len(user_state.emails_cache)} 封邮件, {len(user_state.history)} 条历史记录")
        except Exception as e:
            print(f"[登录] 加载用户状态时出错: {e}")
            import traceback
            traceback.print_exc()
        
        # 登录成功后，检查 autoStart 设置，如果为 True 且系统未运行，自动启动监控
        # 这样，开启"自动运行"后，用户重新登录时也会自动启动监控
        
        return LoginResponse(
            token=token,
            username=username
        )
    raise HTTPException(status_code=401, detail="用户名或密码错误")

@app.post("/api/auth/change-password")
async def change_password(request: ChangePasswordRequest, current_username: str = Depends(get_username_from_request)):
    """修改密码"""
    username = current_username
    
    if username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 验证旧密码
    if user_data[username]["password"] != request.oldPassword:
        raise HTTPException(status_code=400, detail="当前密码错误")
    
    # 验证新密码长度
    if len(request.newPassword) < 6:
        raise HTTPException(status_code=400, detail="新密码长度至少6位")
    
    # 更新密码
    user_data[username]["password"] = request.newPassword
    # 保存到文件
    save_user_data(user_data)
    
    return {"message": "密码修改成功"}

@app.post("/api/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """忘记密码 - 验证用户名和邮箱"""
    username = request.username.strip()
    email = request.email.strip()
    
    # 重新加载用户数据
    global user_data
    user_data = load_user_data()
    
    # 验证用户名是否存在
    if username not in user_data:
        raise HTTPException(status_code=404, detail="用户名不存在")
    
    # 验证邮箱格式
    email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    if not re.match(email_regex, email):
        raise HTTPException(status_code=400, detail="邮箱格式不正确")
    
    # 验证必须是QQ邮箱
    if not email.endswith('@qq.com'):
        raise HTTPException(status_code=400, detail="本系统仅支持QQ邮箱，请输入QQ邮箱地址")
    
    # 验证用户名和邮箱是否匹配
    user_email = user_data[username].get("email", "").strip()
    if not user_email:
        raise HTTPException(status_code=400, detail="该用户未绑定邮箱，无法重置密码")
    
    if user_email.lower() != email.lower():
        raise HTTPException(status_code=400, detail="用户名和邮箱不匹配")
    
    return {
        "message": "验证成功",
        "username": username
    }

@app.post("/api/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    """重置密码"""
    username = request.username.strip()
    email = request.email.strip()
    new_password = request.newPassword
    
    # 重新加载用户数据
    global user_data
    user_data = load_user_data()
    
    # 验证用户名是否存在
    if username not in user_data:
        raise HTTPException(status_code=404, detail="用户名不存在")
    
    # 验证邮箱格式
    email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
    if not re.match(email_regex, email):
        raise HTTPException(status_code=400, detail="邮箱格式不正确")
    
    # 验证必须是QQ邮箱
    if not email.endswith('@qq.com'):
        raise HTTPException(status_code=400, detail="本系统仅支持QQ邮箱，请输入QQ邮箱地址")
    
    # 再次验证用户名和邮箱是否匹配（防止直接调用重置接口）
    user_email = user_data[username].get("email", "").strip()
    if not user_email:
        raise HTTPException(status_code=400, detail="该用户未绑定邮箱，无法重置密码")
    
    if user_email.lower() != email.lower():
        raise HTTPException(status_code=400, detail="用户名和邮箱不匹配")
    
    # 验证新密码长度
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="新密码长度至少6位")
    
    if len(new_password) > 50:
        raise HTTPException(status_code=400, detail="新密码长度不能超过50位")
    
    # 更新密码
    user_data[username]["password"] = new_password
    # 保存到文件
    save_user_data(user_data)
    
    print(f"🔐 [密码重置] 用户 {username} 通过忘记密码功能重置了密码")
    
    return {
        "message": "密码重置成功，请使用新密码登录",
        "username": username
    }

@app.post("/api/auth/update-profile")
async def update_profile(request: UpdateProfileRequest, current_username: str = Depends(get_username_from_request)):
    """更新用户资料（用户名）"""
    username = current_username
    new_username = request.username.strip()
    
    print(f"更新用户资料: 当前用户名={username}, 新用户名={new_username}")
    
    if not new_username:
        raise HTTPException(status_code=400, detail="用户名不能为空")
    
    if len(new_username) < 2:
        raise HTTPException(status_code=400, detail="用户名长度至少2位")
    
    if len(new_username) > 20:
        raise HTTPException(status_code=400, detail="用户名长度不能超过20位")
    
    # 重新加载用户数据，确保使用最新数据
    global user_data
    user_data = load_user_data()
    
    print(f"当前用户数据: {list(user_data.keys())}")
    
    if username not in user_data:
        raise HTTPException(status_code=404, detail=f"用户不存在: {username}")
    
    # 如果用户名改变，只需要更新用户名，user_id保持不变
    # 因为数据文件现在使用user_id命名，所以不需要迁移数据文件
    if new_username != username:
        # 检查新用户名是否已存在
        if new_username in user_data:
            raise HTTPException(status_code=400, detail="用户名已存在")
        
        print(f"更新用户名: {username} -> {new_username}")
        
        # 获取用户的user_id（保持不变）
        user_id = user_data[username].get("user_id")
        if not user_id:
            # 如果旧用户没有user_id，为其生成一个
            user_id = str(uuid.uuid4())
            print(f"为用户 {username} 生成新的 user_id: {user_id}")
        
        # 1. 迁移用户基本信息（深拷贝，确保设备列表也被复制）
        import copy
        user_data[new_username] = copy.deepcopy(user_data[username])
        # 确保user_id保持不变
        user_data[new_username]["user_id"] = user_id
        print(f"✓ 已创建新用户名数据: {new_username}")
        print(f"   检查新用户名是否在数据中: {new_username in user_data}")
        
        # 删除旧用户数据
        if username in user_data:
            del user_data[username]
            print(f"✓ 已删除旧用户名数据: {username}")
        else:
            print(f"⚠️ 警告: 旧用户名 {username} 不在数据中，可能已被删除")
        
        print(f"✓ 用户基本信息已更新（user_id保持不变: {user_id}）")
        print(f"   当前内存中的用户数据: {list(user_data.keys())}")
        
        # 2. 记录用户名映射关系（重要：用于token验证和登录时的用户名解析）
        mapping = load_username_mapping()
        mapping[username] = new_username  # 记录：旧用户名 -> 新用户名
        save_username_mapping(mapping)
        print(f"✓ 用户名映射关系已记录: {username} -> {new_username}")
        
        print(f"✓ 用户名更新完成: {username} -> {new_username} (user_id: {user_id})")
        print(f"   更新后的用户数据（保存前）: {list(user_data.keys())}")
    
    # 先保存用户数据到文件（重要：必须在调用其他函数之前保存）
    print(f"   保存前的用户数据: {list(user_data.keys())}")
    save_user_data(user_data)
    print(f"✓ 用户数据已保存到文件")
    
    # 3. 迁移系统状态（user_states字典）- 在保存user_data之后进行
    if new_username != username:
        global user_states
        if username in user_states:
            # 更新字典键，但不需要迁移数据文件（因为使用user_id命名）
            user_state = user_states[username]
            user_state.username = new_username  # 更新状态对象中的用户名
            user_states[new_username] = user_state
            del user_states[username]
            # 保存状态到文件（使用user_id，文件名不变）
            # 注意：此时user_data已保存，get_user_id_by_username会从文件加载，但新用户名已经在文件中了
            save_user_email_data(new_username, user_state)
            print(f"✓ 系统状态已更新")
    
    # 验证保存是否成功（重新加载数据）
    saved_data = load_user_data()
    print(f"   重新加载后的用户数据: {list(saved_data.keys())}")
    if new_username not in saved_data:
        print(f"❌ 错误: 保存后验证失败，新用户名 {new_username} 不在数据中")
        print(f"   当前数据中的用户名: {list(saved_data.keys())}")
        # 检查是否有映射关系
        mapping = load_username_mapping()
        if username in mapping:
            print(f"   映射关系: {username} -> {mapping[username]}")
        raise HTTPException(status_code=500, detail="保存用户数据失败，请重试")
    
    return {
        "message": "资料更新成功，请重新登录",
        "username": new_username
    }

@app.get("/api/auth/profile")
async def get_profile(current_username: str = Depends(get_username_from_request)):
    """获取当前用户资料"""
    username = current_username
    
    # 重新加载用户数据，确保使用最新数据
    global user_data
    user_data = load_user_data()
    
    if username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 返回用户信息（不包含密码）
    user_info = user_data[username].copy()
    if "password" in user_info:
        del user_info["password"]
    
    # 从用户数据中获取真实信息
    email = user_info.get("email", "")
    register_time = user_info.get("registerTime", "")
    last_login = user_info.get("lastLogin", "")
    avatar = user_info.get("avatar", "")  # 头像（如果存储在用户数据中）
    
    # 如果没有注册时间，使用默认值（兼容旧数据）
    if not register_time:
        register_time = "2024-01-01"
    
    # 如果没有最后登录时间，使用当前时间
    if not last_login:
        last_login = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 如果没有偏好设置，初始化默认值（兼容旧数据）
    if "preferences" not in user_info:
        user_info["preferences"] = {
            "theme": "light",
            "notification": True,
            "sound": False,
            "defaultPage": "/dashboard"
        }
        # 保存更新后的用户数据
        user_data[username] = user_info
        save_user_data(user_data)
    
    return {
        "username": username,
        "email": email,
        "role": "系统管理员",  # 角色可以根据需要从用户数据中获取
        "registerTime": register_time,
        "lastLogin": last_login,
        "avatar": avatar  # 返回头像（如果有）
    }

@app.get("/api/auth/devices")
async def get_devices(current_username: str = Depends(get_username_from_request)):
    """获取登录设备列表"""
    username = current_username
    
    # 重新加载用户数据，确保使用最新数据
    global user_data
    user_data = load_user_data()
    
    if username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    all_devices = user_data[username].get("devices", [])
    
    # 只返回当前活跃的设备（current: true）
    active_devices = [d for d in all_devices if d.get("current", False)]
    
    # 如果没有活跃设备，返回默认数据
    if not active_devices:
        active_devices = [
            {
                "device": "Windows PC",
                "browser": "Chrome 120",
                "ip": "127.0.0.1",
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "current": True
            }
        ]
    
    return {
        "devices": active_devices,
        "activeCount": len(active_devices)
    }

@app.get("/api/auth/preferences")
async def get_preferences(current_username: str = Depends(get_username_from_request)):
    """获取用户偏好设置"""
    username = current_username
    
    # 重新加载用户数据，确保使用最新数据
    global user_data
    user_data = load_user_data()
    
    if username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 获取用户偏好设置，如果没有则返回默认值
    preferences = user_data[username].get("preferences", {
        "theme": "light",
        "notification": True,
        "sound": False,
        "defaultPage": "/dashboard"
    })
    
    return preferences

@app.post("/api/auth/preferences")
async def save_preferences(request: PreferencesRequest, current_username: str = Depends(get_username_from_request)):
    """保存用户偏好设置"""
    username = current_username
    
    # 重新加载用户数据，确保使用最新数据
    global user_data
    user_data = load_user_data()
    
    if username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    # 更新用户偏好设置
    if "preferences" not in user_data[username]:
        user_data[username]["preferences"] = {}
    
    # 只更新提供的字段
    if request.theme is not None:
        user_data[username]["preferences"]["theme"] = request.theme
    if request.notification is not None:
        user_data[username]["preferences"]["notification"] = request.notification
    if request.sound is not None:
        user_data[username]["preferences"]["sound"] = request.sound
    if request.defaultPage is not None:
        user_data[username]["preferences"]["defaultPage"] = request.defaultPage
    
    # 保存到文件
    save_user_data(user_data)
    
    # 记录操作：修改偏好设置
    user_state = get_user_state(current_username)
    user_state.add_activity('warning', '修改了个人偏好设置', 'Setting')
    
    return {
        "message": "偏好设置已保存",
        "preferences": user_data[username]["preferences"]
    }

@app.post("/api/auth/devices/{device_id}/logout")
async def logout_device(device_id: int, current_username: str = Depends(get_username_from_request)):
    """下线指定设备"""
    username = current_username
    
    if username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    devices = user_data[username].get("devices", [])
    
    if device_id < 0 or device_id >= len(devices):
        raise HTTPException(status_code=404, detail="设备不存在")
    
    # 不能下线当前设备
    if devices[device_id]["current"]:
        raise HTTPException(status_code=400, detail="不能下线当前设备")
    
    # 移除设备
    devices.pop(device_id)
    user_data[username]["devices"] = devices
    # 保存设备信息
    save_user_data(user_data)
    
    return {"message": "设备已下线"}

# ==================== 邮件API ====================

@app.get("/api/emails")
async def get_emails(status: Optional[str] = None, category: Optional[str] = None, current_username: str = Depends(get_username_from_request)):
    """获取邮件列表（即使没有配置邮箱也返回空列表，不阻止用户查看页面）"""
    try:
        user_state = get_user_state(current_username)
        emails = user_state.emails_cache.copy()
        
        # 筛选
        if status:
            if status == "pending":
                emails = [e for e in emails if e.get('status') == 'pending']
            elif status == "processed":
                emails = [e for e in emails if e.get('status') != 'pending']
                
        if category:
            emails = [e for e in emails if e.get('category') == category]
        
        # 只返回真实数据，不使用模拟数据
        # 如果缓存为空，返回空数组
        
        return {"emails": emails, "total": len(emails)}
    except Exception as e:
        # 即使出错也返回空列表，不阻止用户查看页面
        print(f"获取邮件列表失败: {e}")
        return {"emails": [], "total": 0}

@app.get("/api/emails/{email_id:path}")
async def get_email_detail(email_id: str, current_username: str = Depends(get_username_from_request)):
    """获取邮件详情"""
    email_id = unquote(email_id)
    user_state = get_user_state(current_username)
    for email in user_state.emails_cache:
        if email.get('id') == email_id:
            return email
    raise HTTPException(status_code=404, detail="邮件不存在")

@app.post("/api/emails/{email_id:path}/process")
async def process_email(email_id: str, current_username: str = Depends(get_username_from_request)):
    """处理单封邮件（异步后台处理，通过WebSocket通知完成）"""
    email_id = unquote(email_id)
    user_state = get_user_state(current_username)
    
    # 查找邮件
    email = None
    for e in user_state.emails_cache:
        if e.get('id') == email_id:
            email = e
            break
    
    if not email:
        raise HTTPException(status_code=404, detail="邮件不存在")
    
    # 使用用户锁保护状态检查和更新（防止并发处理同一封邮件）
    user_lock = get_user_lock(current_username)
    # 使用显式 acquire/release，避免 with 语句在某些环境触发缩进相关问题
    user_lock.acquire()
    try:
        # 如果已经在处理中，返回提示
        if email.get('status') == 'processing':
            return {
                "success": False,
                "message": "邮件正在处理中，请稍候",
                "status": "processing"
            }
        
        # 更新状态为处理中
        email['status'] = 'processing'
    finally:
        try:
            user_lock.release()
        except Exception:
            pass
    
    # 通过WebSocket通知前端邮件开始处理
    await ws_manager.broadcast({
        "type": "email_process_started",
        "email_id": email_id,
        "message": "开始处理邮件"
    })
    
    def process_email_sync():
        """同步处理邮件的函数（在线程池中执行，避免阻塞事件循环）"""
        from src.nodes import Nodes
        from src.state import Email
        
        # 捕获外层作用域的 email_id（避免作用域冲突）
        task_email_id = email_id
        
        # 重新获取用户状态（确保使用最新的数据）
        task_user_state = get_user_state(current_username)
        user_lock = get_user_lock(current_username)
        
        # 辅助函数：检查并处理终止
        def check_and_handle_stop(checkpoint_name):
            """检查是否被终止，如果是则清除标记并返回True"""
            # 添加调试日志
            print(f"🔍 [检查点] 邮件 {task_email_id} 在{checkpoint_name}检查终止标志: stop_processing={task_user_state.stop_processing}, stopped_email_ids={task_email_id in task_user_state.stopped_email_ids}")
            
            # 检查全局停止标志（批量处理终止）
            if task_user_state.stop_processing:
                print(f"⏹️ [批量处理终止] 邮件 {task_email_id} 在{checkpoint_name}被终止（全局停止标志）")
                with user_lock:
                    # 查找邮件并恢复状态
                    for e in task_user_state.emails_cache:
                        if e.get('id') == task_email_id:
                            e['status'] = 'pending'
                            e['processing'] = False
                            break
                    save_user_email_data(current_username, task_user_state)
                print(f"⏹️ [批量处理终止] 已恢复邮件 {task_email_id} 的状态")
                
                # 发送WebSocket通知（真正终止成功）
                asyncio.run_coroutine_threadsafe(
                    ws_manager.broadcast({
                        "type": "email_process_stopped",
                        "email_id": task_email_id,
                        "message": "已终止处理"
                    }),
                    websocket_event_loop
                )
                
                return True
            
            # 检查单封邮件停止标志
            if task_email_id in task_user_state.stopped_email_ids:
                print(f"⏹️ [单封邮件处理] 邮件 {task_email_id} 在{checkpoint_name}被终止")
                with user_lock:
                    # 查找邮件并恢复状态
                    for e in task_user_state.emails_cache:
                        if e.get('id') == task_email_id:
                            e['status'] = 'pending'
                            e['processing'] = False
                            break
                    # 主动清除终止标记
                    task_user_state.stopped_email_ids.discard(task_email_id)
                    save_user_email_data(current_username, task_user_state)
                print(f"⏹️ [单封邮件处理] 已清除邮件 {task_email_id} 的终止标记")
                
                # 发送WebSocket通知（真正终止成功）
                asyncio.run_coroutine_threadsafe(
                    ws_manager.broadcast({
                        "type": "email_process_stopped",
                        "email_id": task_email_id,
                        "message": "已终止处理"
                    }),
                    websocket_event_loop
                )
                
                return True
            
            return False
        
        # 检查点1：处理开始前
        if check_and_handle_stop("处理开始前"):
            return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
        
        # 根据ID重新查找邮件
        task_email = None
        for e in task_user_state.emails_cache:
            if e.get('id') == task_email_id:
                task_email = e
                break
        
        if not task_email:
            print(f"邮件不存在: {task_email_id}")
            return None
        
        # 确保邮件状态是处理中
        if task_email.get('status') != 'processing':
            print(f"邮件状态不是处理中，跳过: {task_email.get('status')}")
            return None
        
        try:
            # 获取当前用户的邮箱配置
            email_address, auth_code = get_user_email_config(current_username)
            # 获取当前用户的设置（包括API密钥和模型配置）
            user_settings = get_user_settings(current_username)
            reply_model = user_settings.get("replyModel", user_settings.get("model", "moonshotai/Kimi-K2-Thinking"))
            embedding_model = user_settings.get("embeddingModel", "Qwen/Qwen3-Embedding-4B")
            # 根据选择的模型自动获取API密钥和base URL
            models_config = get_models_config(current_username, reply_model, embedding_model)
            nodes = Nodes(
                email_address=email_address, 
                auth_code=auth_code, 
                api_key=models_config["apiKey"],  # 如果为None，则使用系统默认API（环境变量）
                reply_model=reply_model,
                embedding_model=embedding_model,
                signature=user_settings.get("signature"),
                greeting=user_settings.get("greeting"),
                closing=user_settings.get("closing"),
                reply_api_base=models_config["replyApiBaseUrl"],
                embedding_api_base=models_config["embeddingApiBaseUrl"]
            )
            
            # 创建Email对象
            email_obj = Email(
                id=task_email.get('id', ''),
                threadId=task_email.get('threadId', ''),
                messageId=task_email.get('messageId', ''),
                references=task_email.get('references', ''),
                sender=task_email.get('sender', ''),
                subject=task_email.get('subject', ''),
                body=task_email.get('body', ''),
                imap_id=task_email.get('imap_id', b'')
            )
            
            # 构建状态
            state = {
                "emails": [email_obj],
                "current_email": email_obj,
                "email_category": None,
                "rag_queries": [],
                "retrieved_documents": "",
                "generated_email": "",
                "sendable": False,
                "trials": 0,
                "writer_messages": []
            }
            
            # 分类名称映射
            category_names = {
                'product_enquiry': '产品咨询',
                'customer_complaint': '客户投诉',
                'customer_feedback': '客户反馈',
                'unrelated': '无关邮件'
            }
            
            # 1. 分类邮件（同步阻塞操作）
            print(f"[邮件分类] 正在分类邮件:")
            print(f"  - 主题: {task_email.get('subject', '')}")
            print(f"  - 发件人: {task_email.get('sender', '')}")
            print(f"  - 内容预览: {task_email.get('body', '')[:200]}...")
            
            categorize_result = nodes.categorize_email(state)
            state.update(categorize_result)
            category = state.get('email_category', 'product_enquiry')
            task_email['category'] = category
            # 同步紧急程度信息（从Email对象获取）
            if 'emails' in state and len(state['emails']) > 0:
                email_obj = state['emails'][0]
                if hasattr(email_obj, 'urgency_level'):
                    task_email['urgency_level'] = email_obj.urgency_level
                if hasattr(email_obj, 'urgency_keywords'):
                    task_email['urgency_keywords'] = email_obj.urgency_keywords
            
            print(f"[邮件分类] 分类结果: {category}")
            print(f"  - 邮件ID: {task_email.get('id', '')}")
            print(f"  - 主题: {task_email.get('subject', '')}")
            
            # 检查点2：分类后
            if check_and_handle_stop("分类后"):
                return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
            
            # 如果分类为无关邮件，但邮件主题或内容包含投诉相关关键词，记录警告
            if category == 'unrelated':
                complaint_keywords = ['投诉', '不满', '差评', '退款', '问题严重', '态度差', '客户投诉']
                email_text = (task_email.get('subject', '') + ' ' + task_email.get('body', '')).lower()
                has_complaint_keyword = any(keyword in email_text for keyword in complaint_keywords)
                if has_complaint_keyword:
                    print(f"[邮件分类警告] 邮件被分类为无关邮件，但包含投诉关键词！")
                    print(f"  - 邮件ID: {task_email.get('id', '')}")
                    print(f"  - 主题: {task_email.get('subject', '')}")
                    print(f"  - 内容: {task_email.get('body', '')[:500]}")
            
            # 2. 检查是否是无关邮件
            if category == 'unrelated':
                task_email['status'] = 'skipped'
                task_email['reply'] = '无关邮件，已跳过'
                
                # 标记为已读
                imap_id = task_email.get('imap_id')
                if imap_id:
                    try:
                        nodes.email_tools.mark_email_as_read(imap_id)
                    except:
                        pass
                
                print(f"[邮件处理] 跳过无关邮件:")
                print(f"  - 邮件ID: {task_email.get('id', '')}")
                print(f"  - 主题: {task_email.get('subject', '')}")
                print(f"  - 分类: {category}")
                
                # 更新统计
                task_user_state.stats['pending'] = max(0, task_user_state.stats['pending'] - 1)
                
                # 添加到历史记录
                task_user_state.history.insert(0, {
                    **task_email,
                    'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                
                # 自动保存数据
                save_user_email_data(current_username, task_user_state)
                
                # 为无关邮件生成原始邮件摘要（异步，不阻塞）
                task_email_id = task_email.get('id')
                email_body = task_email.get('body', '')
                if email_body:
                    print(f"🚀 [摘要触发] 准备为无关邮件 {task_email_id} 生成原始邮件摘要...")
                    # 无关邮件的回复内容是"无关邮件，已跳过"，很短，不需要生成摘要
                    # 只生成原始邮件摘要
                    generate_email_summaries_async(
                        current_username,
                        task_email_id,
                        email_body,
                        ''  # 不生成回复内容摘要
                    )
                
                return {
                    "status": "skipped",
                    "message": "无关邮件，已跳过",
                    "category": category,
                    "reply": "无关邮件，已跳过"  # 包含回复内容
                }
            
            # 3. 根据邮件类型进行RAG查询（同步阻塞操作）
            # 除了无关邮件，其他类型都进行RAG检索
            if category != 'unrelated':
                # 检查点3：RAG查询前
                if check_and_handle_stop("RAG查询前"):
                    return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
                
                print(f"正在进行RAG查询（类型: {category}）...")
                rag_query_result = nodes.construct_rag_queries(state)
                state.update(rag_query_result)
                
                # 发送 WebSocket 通知：显示生成的 RAG 查询问题
                rag_queries = state.get('rag_queries', [])
                if rag_queries:
                    asyncio.run_coroutine_threadsafe(
                        ws_manager.broadcast({
                            "type": "rag_queries_generated",
                            "email_id": task_email_id,
                            "queries": rag_queries,
                            "count": len(rag_queries)
                        }),
                        websocket_event_loop
                    )
                
                rag_result = nodes.retrieve_from_rag(state)
                state.update(rag_result)
            else:
                state['retrieved_documents'] = ""
            
            # 检查点4：RAG查询后/编写回复前
            if check_and_handle_stop("RAG查询后"):
                return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
            
            # 4. 编写回复邮件（同步阻塞操作）
            print("正在编写回复邮件...")
            
            # 检查点5：开始编写回复前
            if check_and_handle_stop("开始编写回复前"):
                return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
            
            max_trials = 3
            for trial in range(max_trials):
                # 检查点6：每次重试前
                if check_and_handle_stop(f"编写回复循环中（第{trial+1}次尝试）"):
                    return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
                
                write_result = nodes.write_draft_email(state)
                state.update(write_result)
                
                # 检查点7：验证前
                if check_and_handle_stop(f"验证前（第{trial+1}次尝试）"):
                    return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
                
                # 5. 验证邮件（同步阻塞操作）
                verify_result = nodes.verify_generated_email(state)
                state.update(verify_result)
                
                # 检查点7.5：验证后
                if check_and_handle_stop(f"验证后（第{trial+1}次尝试）"):
                    return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
                
                if state.get('sendable', False):
                    break
            
            # 6. 获取生成的回复
            generated_reply = state.get('generated_email', '')
            
            # 检查点8：保存回复前
            if check_and_handle_stop("保存回复前"):
                return {'status': 'cancelled', 'message': '处理已终止', 'reply': None}
            
            task_email['reply'] = generated_reply
            task_email['status'] = 'processed'
            task_email['rag_queries'] = state.get('rag_queries', [])  # 保存 RAG 查询问题
            
            # 7. 检查是否自动发送（根据用户设置）
            auto_send = user_settings.get("autoSend", False)
            if auto_send and generated_reply:
                # 自动发送回复（带速率限制）
                try:
                    result, message = send_reply_with_rate_limit(
                        current_username,
                        nodes.email_tools,
                        email_obj,
                        generated_reply,
                        task_email
                    )
                    if result:
                        # 发送成功，状态已在 send_reply_with_rate_limit 中更新
                        sender_name = task_email.get('sender', '').split('@')[0] if '@' in task_email.get('sender', '') else task_email.get('sender', '未知')
                        task_user_state.add_activity('primary', f'自动发送回复给: {sender_name}', 'Message')
                        # 统计已在 send_reply_with_rate_limit 中更新（task_user_state 和 user_state 是同一个对象）
                    else:
                        # 发送失败或达到限制，保持 processed 状态
                        task_email['status'] = 'processed'
                        print(f"⚠️ 自动发送回复失败或达到限制: {task_email.get('subject', '')} - {message}")
                except Exception as send_err:
                    print(f"❌ 自动发送回复时出错: {send_err}")
                    # 发送失败，保持 processed 状态，用户可以手动发送
                    task_email['status'] = 'processed'
            
            # 8. 标记QQ邮箱中的邮件为已读
            imap_id = task_email.get('imap_id')
            if imap_id:
                try:
                    nodes.email_tools.mark_email_as_read(imap_id)
                    print(f"邮件已标记为已读: {task_email.get('subject', '')}")
                except Exception as mark_err:
                    print(f"标记已读失败: {mark_err}")
            
            # 9. 更新统计
            task_user_state.stats['processed'] += 1
            task_user_state.stats['pending'] = max(0, task_user_state.stats['pending'] - 1)
            
            # 10. 添加到历史记录
            task_user_state.history.insert(0, {
                **task_email,
                'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # 11. 记录操作（如果还没有记录）
            if not auto_send or not generated_reply or task_email.get('status') != 'sent':
                # 确保 category_names 已定义
                if 'category_names' not in locals():
                    category_names = {
                        'product_enquiry': '产品咨询',
                        'customer_complaint': '客户投诉',
                        'customer_feedback': '客户反馈',
                        'unrelated': '无关邮件'
                    }
            category_label = category_names.get(category, category or '未分类')
            task_user_state.add_activity('success', f'处理了邮件: {category_label}', 'CircleCheck')
            
            # 自动保存数据
            save_user_email_data(current_username, task_user_state)
            
            print(f"邮件处理完成: {task_email.get('subject', '')}")
            
            # 12. 异步生成摘要（不阻塞主流程）
            task_email_id = task_email.get('id')
            email_body = task_email.get('body', '')
            has_body = bool(email_body)
            has_reply = bool(generated_reply)
            
            print(f"🔍 [摘要检查] 邮件 {task_email_id}: body存在={has_body}, reply存在={has_reply}")
            
            if has_body or has_reply:
                print(f"🚀 [摘要触发] 准备为邮件 {task_email_id} 生成摘要...")
                generate_email_summaries_async(
                    current_username,
                    task_email_id,
                    email_body,
                    generated_reply or ''
                )
            else:
                print(f"⚠️ [摘要跳过] 邮件 {task_email_id} 没有body和reply，跳过摘要生成")
            
            return {
                "status": "processed",
                "message": f"{category_names.get(category, category)} - 处理成功",  # 只返回简洁的成功消息，不包含"邮件已标记为已读"等详细信息
                "category": category,
                "reply": generated_reply,  # 包含生成的回复内容
                "rag_queries": state.get('rag_queries', [])  # 包含生成的 RAG 查询问题
            }
            
        except Exception as e:
            task_email['status'] = 'failed'
            task_user_state.stats['failed'] += 1
            print(f"处理邮件错误: {e}")
            import traceback
            traceback.print_exc()
            
            # 处理失败时，也添加到历史记录中，这样即使从缓存中删除，统计数据也不会丢失
            task_user_state.history.insert(0, {
                **task_email,
                'status': 'failed',  # 明确设置为 'failed'
                'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            print(f"DEBUG [process_email_sync]: 处理失败，已添加到历史记录，ID: {task_email.get('id')}, Status: failed")
            
            # 自动保存数据
            save_user_email_data(current_username, task_user_state)
            
            return {
                "status": "failed",
                "message": f"处理失败: {str(e)}",
                "reply": None  # 失败时没有回复内容
            }
    
    async def process_email_task():
        """异步包装函数，在线程池中执行同步阻塞的AI操作"""
        try:
            # 获取用户设置的单封邮件并发数量
            user_settings = get_user_settings(current_username)
            single_email_concurrency = user_settings.get("singleEmailConcurrency", 4)
            # 限制在合理范围内（2-20）
            single_email_concurrency = max(2, min(20, int(single_email_concurrency)))
            
            # 获取或创建单封邮件处理线程池
            email_pool = get_or_create_single_email_thread_pool(single_email_concurrency)
            
            # 在线程池中执行同步阻塞的AI操作，避免阻塞事件循环
            # 使用单封邮件处理线程池，避免占用主线程池
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(email_pool, process_email_sync)
            
            if result:
                # 如果状态是 cancelled，说明已被终止，不发送完成通知
                if result.get('status') == 'cancelled':
                    print(f"[WebSocket发送] 邮件 {email_id} 已被终止，跳过发送完成通知")
                    return
                
                # 通过 WebSocket 通知前端
                ws_message = {
                    "type": "email_process_complete",
                    "email_id": email_id,
                    "message": result.get("message", ""),
                    "category": result.get("category"),
                    "status": result.get("status"),
                    "reply": result.get("reply"),  # 包含生成的回复内容
                    "rag_queries": result.get("rag_queries", [])  # 包含 RAG 查询问题
                }
                print(f"[WebSocket发送] 准备发送 email_process_complete 消息:")
                print(f"  - 邮件ID: {email_id}")
                print(f"  - 状态: {result.get('status')}")
                print(f"  - 分类: {result.get('category')}")
                print(f"  - 消息: {result.get('message', '')}")
                print(f"  - RAG查询: {len(result.get('rag_queries', []))} 个")
                print(f"  - 当前连接数: {len(ws_manager.active_connections)}")
                
                await ws_manager.broadcast(ws_message)
                print(f"[WebSocket发送] 已发送 email_process_complete 消息")
        except Exception as e:
            print(f"异步任务错误: {e}")
            import traceback
            traceback.print_exc()
            # 通知前端处理失败
            await ws_manager.broadcast({
                "type": "email_process_complete",
                "email_id": email_id,
                "message": f"处理失败: {str(e)}",
                "status": "failed"
            })
    
    # 在后台运行异步任务
    asyncio.create_task(process_email_task())
    
    return {
        "success": True,
        "message": "邮件处理已开始，处理完成后会通知您",
        "status": "processing"
    }


class ReRetrieveRequest(BaseModel):
    """重新检索请求"""
    queries: List[str] = Field(..., description="用户修改后的查询问题列表")


@app.post("/api/emails/{email_id:path}/re-retrieve")
async def re_retrieve_email(
    email_id: str, 
    request: ReRetrieveRequest,
    current_username: str = Depends(get_username_from_request)
):
    """使用用户修改的查询问题重新检索并生成回复"""
    email_id = unquote(email_id)
    
    print(f"🔄 [重新检索] 用户 {current_username} 请求对邮件 {email_id} 重新检索")
    print(f"   修改后的查询问题: {request.queries}")
    
    user_state = get_user_state(current_username)
    user_lock = get_user_lock(current_username)
    
    # 查找邮件
    email = None
    with user_lock:
        for e in user_state.emails_cache:
            if e.get('id') == email_id:
                email = e
                break
    
    if not email:
        raise HTTPException(status_code=404, detail="邮件不存在")
    
    def re_retrieve_sync():
        """同步执行重新检索"""
        from src.nodes import Nodes
        from src.state import Email
        
        try:
            # 获取用户设置
            user_settings = get_user_settings(current_username)
            
            # 获取模型配置（包括 API key 和 base URL）
            reply_model = user_settings.get('replyModel', 'moonshotai/Kimi-K2-Thinking')
            embedding_model = user_settings.get('embeddingModel', 'Qwen/Qwen3-Embedding-4B')
            
            print(f"🔍 [重新检索] 回复模型: {reply_model}")
            print(f"🔍 [重新检索] 嵌入模型: {embedding_model}")
            
            models_config = get_models_config(current_username, reply_model, embedding_model)
            api_key = models_config["apiKey"]
            reply_api_base = models_config["replyApiBaseUrl"]
            embedding_api_base = models_config["embeddingApiBaseUrl"]
            
            print(f"🔑 [重新检索] API密钥存在: {bool(api_key)}")
            print(f"🌐 [重新检索] Reply API Base: {reply_api_base}")
            print(f"🌐 [重新检索] Embedding API Base: {embedding_api_base}")
            
            if not api_key:
                import os
                env_key = os.getenv("SILICONFLOW_API_KEY")
                print(f"⚠️ [重新检索] 环境变量 SILICONFLOW_API_KEY 存在: {bool(env_key)}")
                if env_key:
                    print(f"⚠️ [重新检索] 环境变量值前10位: {env_key[:10]}...")
                raise ValueError("未找到 API 密钥，请在系统设置中配置或设置 SILICONFLOW_API_KEY 环境变量")
            
            # 创建 Nodes 实例
            nodes = Nodes(
                email_address=None,  # 重新检索不需要邮箱配置
                auth_code=None,
                api_key=api_key,
                reply_model=reply_model,
                embedding_model=embedding_model,
                signature=user_settings.get('signature'),
                greeting=user_settings.get('greeting'),
                closing=user_settings.get('closing'),
                reply_api_base=reply_api_base,
                embedding_api_base=embedding_api_base
            )
            
            # 创建邮件对象
            email_obj = Email(
                id=email.get('id', ''),
                threadId=email.get('threadId', ''),
                messageId=email.get('messageId', ''),
                references=email.get('references', ''),
                sender=email.get('sender', ''),
                subject=email.get('subject', ''),
                body=email.get('body', ''),
                imap_id=email.get('imap_id', b'')
            )
            
            # 使用用户提供的查询问题
            state = {
                "emails": [email_obj],
                "current_email": email_obj,
                "email_category": email.get('category', 'product_enquiry'),
                "rag_queries": request.queries,  # 使用用户修改的问题
                "retrieved_documents": "",
                "writer_messages": [],
                "generated_email": "",
                "sendable": False,
                "trials": 0
            }
            
            print(f"🔍 [重新检索] 使用 {len(request.queries)} 个查询问题进行检索...")
            
            # 执行 RAG 检索
            rag_result = nodes.retrieve_from_rag(state)
            state.update(rag_result)
            
            print(f"📄 [重新检索] 检索完成，结果长度: {len(state.get('retrieved_documents', ''))}")
            
            # 重新生成回复
            print(f"✍️ [重新检索] 正在重新生成回复...")
            write_result = nodes.write_draft_email(state)
            state.update(write_result)
            
            generated_reply = state.get('generated_email', '')
            print(f"✅ [重新检索] 回复生成完成，长度: {len(generated_reply)}")
            
            # 更新邮件数据
            with user_lock:
                email['reply'] = generated_reply
                email['rag_queries'] = request.queries
            
            # 保存数据
            save_user_email_data(current_username, user_state)
            
            return {
                "success": True,
                "reply": generated_reply,
                "rag_queries": request.queries,
                "retrieved_documents": state.get('retrieved_documents', '')[:500]  # 只返回前500字符
            }
            
        except Exception as e:
            print(f"❌ [重新检索] 失败: {e}")
            import traceback
            traceback.print_exc()
            return {
                "success": False,
                "error": str(e)
            }
    
    # 在线程池中执行同步操作
    import asyncio
    result = await asyncio.to_thread(re_retrieve_sync)
    
    if not result.get('success'):
        raise HTTPException(status_code=500, detail=result.get('error', '重新检索失败'))
    
    return result


@app.post("/api/emails/process-all")
async def process_all_emails(background_tasks: BackgroundTasks, current_username: str = Depends(get_username_from_request)):
    """处理所有待处理邮件（异步处理，通过WebSocket通知完成）"""
    user_state = get_user_state(current_username)
    # 使用用户锁保护状态检查和更新，避免与单封邮件处理冲突
    user_lock = get_user_lock(current_username)
    with user_lock:
        # 重置停止标志（确保之前的终止操作不会影响本次处理）
        user_state.stop_processing = False
        user_state.stopped_email_ids.clear()
        print(f"🔄 [批量处理] 重置停止标志，开始新的批量处理")
        
        # 获取待处理邮件列表（排除已经在处理中的邮件，避免重复处理）
        pending_emails = [
            e for e in user_state.emails_cache 
            if e.get('status') == 'pending'
        ]
    
    if not pending_emails:
        return {"message": "没有待处理的邮件", "count": 0}
    
        # 标记所有待处理邮件为处理中（原子操作，避免竞争条件）
    for email in pending_emails:
        email['status'] = 'processing'
    
    # 保存邮件ID列表，用于在后台任务中重新查找（因为邮件状态可能会变化）
    pending_email_ids = [e.get('id') for e in pending_emails]
    
    def process_all_sync():
        """同步处理所有邮件的函数（在线程池中执行，支持并发处理）"""
        from src.nodes import Nodes
        from src.state import Email
        
        # 重新获取用户状态（确保使用最新的数据）
        task_user_state = get_user_state(current_username)
        
        # 根据ID重新查找待处理的邮件（因为状态可能已经变化）
        emails_to_process = []
        for email_id in pending_email_ids:
            for e in task_user_state.emails_cache:
                if e.get('id') == email_id and e.get('status') == 'processing':
                    emails_to_process.append(e)
                    break
        
        if not emails_to_process:
            print("没有需要处理的邮件（可能已被其他操作处理）")
            return {
                "processed": 0,
                "skipped": 0,
                "failed": 0,
                "email_results": []
            }
        
        print(f"🚀 [并发处理] 开始处理 {len(emails_to_process)} 封邮件，使用线程池并发处理")
        
        # 分类名称映射
        category_names = {
            'product_enquiry': '产品咨询',
            'customer_complaint': '客户投诉',
            'customer_feedback': '客户反馈',
            'unrelated': '无关邮件'
        }
        
        # 线程安全的计数器（使用锁保护）
        user_lock = get_user_lock(current_username)
        processed_count = 0
        failed_count = 0
        skipped_count = 0
        # 收集每封邮件的处理结果，用于发送 WebSocket 通知
        email_results = []
        
        # 获取用户配置（所有邮件共享）
        try:
            email_address, auth_code = get_user_email_config(current_username)
            user_settings = get_user_settings(current_username)
            reply_model = user_settings.get("replyModel", user_settings.get("model", "moonshotai/Kimi-K2-Thinking"))
            embedding_model = user_settings.get("embeddingModel", "Qwen/Qwen3-Embedding-4B")
            models_config = get_models_config(current_username, reply_model, embedding_model)
            api_key = models_config["apiKey"]
            reply_api_base = models_config["replyApiBaseUrl"]
            embedding_api_base = models_config["embeddingApiBaseUrl"]
        except Exception as e:
            print(f"❌ [并发处理] 获取用户配置失败: {e}")
            import traceback
            traceback.print_exc()
            # 标记所有邮件为失败
            with user_lock:
                for email_id in pending_email_ids:
                    for e in task_user_state.emails_cache:
                        if e.get('id') == email_id and e.get('status') == 'processing':
                            e['status'] = 'failed'
                            break
            return {
                "processed": 0,
                "skipped": 0,
                "failed": len(emails_to_process),
                "email_results": []
            }
        
        def process_single_email(email):
            """处理单封邮件的函数（在线程池中并发执行）"""
            email_id = email.get('id', '')
            
            # 检查是否被终止
            if email_id in task_user_state.stopped_email_ids:
                print(f"⏹️ [并发处理] 邮件 {email_id} 已被终止，跳过处理")
                with user_lock:
                    email['status'] = 'pending'
                    email['processing'] = False
                return {
                    'email_id': email_id,
                    'status': 'cancelled',
                    'message': '处理已终止',
                    'reply': None
                }
            
            # 检查全局停止标志
            if task_user_state.stop_processing:
                print(f"⏹️ [并发处理] 检测到全局停止标志，跳过邮件 {email_id}")
                with user_lock:
                    email['status'] = 'pending'
                    email['processing'] = False
                
                # 发送WebSocket通知（单封邮件终止成功）
                asyncio.run_coroutine_threadsafe(
                    ws_manager.broadcast({
                        "type": "email_process_stopped",
                        "email_id": email_id,
                        "message": "已终止处理"
                    }),
                    websocket_event_loop
                )
                
                return {
                    'email_id': email_id,
                    'status': 'cancelled',
                    'message': '批量处理已终止',
                    'reply': None
                }
            
            try:
                print(f"📧 [并发处理] 开始处理邮件: {email.get('subject', '')[:50]}...")
                
                # 为每封邮件创建独立的Nodes实例（避免并发冲突）
                nodes = Nodes(
                    email_address=email_address, 
                    auth_code=auth_code, 
                    api_key=api_key,
                    reply_model=reply_model,
                    embedding_model=embedding_model,
                    signature=user_settings.get("signature"),
                    greeting=user_settings.get("greeting"),
                    closing=user_settings.get("closing"),
                    reply_api_base=reply_api_base,
                    embedding_api_base=embedding_api_base
                )
                
                # 创建Email对象
                email_obj = Email(
                    id=email.get('id', ''),
                    threadId=email.get('threadId', ''),
                    messageId=email.get('messageId', ''),
                    references=email.get('references', ''),
                    sender=email.get('sender', ''),
                    subject=email.get('subject', ''),
                    body=email.get('body', ''),
                    imap_id=email.get('imap_id', b'')
                )
                
                # 构建状态
                state = {
                    "emails": [email_obj],
                    "current_email": email_obj,
                    "email_category": None,
                    "rag_queries": [],
                    "retrieved_documents": "",
                    "generated_email": "",
                    "sendable": False,
                    "trials": 0,
                    "writer_messages": []
                }
                
                # 1. 分类邮件
                categorize_result = nodes.categorize_email(state)
                state.update(categorize_result)
                category = state.get('email_category', 'product_enquiry')
                
                # 获取分类标签
                category_label = category_names.get(category, category or '未分类')
                
                # 2. 检查是否是无关邮件
                if category == 'unrelated':
                    # 使用锁保护状态更新
                    with user_lock:
                        email['status'] = 'skipped'
                        email['category'] = category
                        email['reply'] = '无关邮件，已跳过'
                        
                        # 标记为已读
                        imap_id = email.get('imap_id')
                        if imap_id:
                            try:
                                nodes.email_tools.mark_email_as_read(imap_id)
                            except:
                                pass
                    
                    print(f"⏭️ [并发处理] 跳过无关邮件: {email.get('subject', '')[:50]}...")
                    
                    # 为无关邮件生成原始邮件摘要（异步，不阻塞）
                    email_body = email.get('body', '')
                    if email_body:
                        print(f"🚀 [摘要触发] 准备为无关邮件 {email_id} 生成原始邮件摘要...")
                        # 无关邮件的回复内容是"无关邮件，已跳过"，很短，不需要生成摘要
                        # 只生成原始邮件摘要
                        generate_email_summaries_async(
                            current_username,
                            email_id,
                            email_body,
                            ''  # 不生成回复内容摘要
                        )
                    
                    return {
                        'email_id': email_id,
                        'status': 'skipped',
                        'category': category,
                        'message': '无关邮件，已跳过',
                        'reply': '无关邮件，已跳过'  # 包含回复内容
                    }
                
                # 3. RAG查询
                if category != 'unrelated':
                    # 检查点：RAG查询前
                    if task_user_state.stop_processing:
                        print(f"⏹️ [批量处理终止] 邮件 {email_id} 在RAG查询前被终止")
                        with user_lock:
                            email['status'] = 'pending'
                            email['processing'] = False
                        asyncio.run_coroutine_threadsafe(
                            ws_manager.broadcast({
                                "type": "email_process_stopped",
                                "email_id": email_id,
                                "message": "已终止处理"
                            }),
                            websocket_event_loop
                        )
                        return {
                            'email_id': email_id,
                            'status': 'cancelled',
                            'message': '批量处理已终止',
                            'reply': None
                        }
                    
                    print(f"🔍 [并发处理] 正在进行RAG查询（类型: {category}）...")
                    rag_query_result = nodes.construct_rag_queries(state)
                    state.update(rag_query_result)
                    
                    # 发送通知：显示生成的 RAG 查询问题
                    rag_queries = state.get('rag_queries', [])
                    if rag_queries:
                        asyncio.run_coroutine_threadsafe(
                            ws_manager.broadcast({
                                "type": "rag_queries_generated",
                                "email_id": email_id,
                                "queries": rag_queries,
                                "count": len(rag_queries)
                            }),
                            websocket_event_loop
                        )
                    
                    rag_result = nodes.retrieve_from_rag(state)
                    state.update(rag_result)
                else:
                    state['retrieved_documents'] = ""
                
                # 检查点：RAG查询后
                if task_user_state.stop_processing:
                    print(f"⏹️ [批量处理终止] 邮件 {email_id} 在RAG查询后被终止")
                    with user_lock:
                        email['status'] = 'pending'
                        email['processing'] = False
                    asyncio.run_coroutine_threadsafe(
                        ws_manager.broadcast({
                            "type": "email_process_stopped",
                            "email_id": email_id,
                            "message": "已终止处理"
                        }),
                        websocket_event_loop
                    )
                    return {
                        'email_id': email_id,
                        'status': 'cancelled',
                        'message': '批量处理已终止',
                        'reply': None
                    }
                
                # 4. 编写回复邮件
                max_trials = 3
                for trial in range(max_trials):
                    # 检查点：每次重试前
                    if task_user_state.stop_processing:
                        print(f"⏹️ [批量处理终止] 邮件 {email_id} 在编写回复前被终止（第{trial+1}次尝试）")
                        with user_lock:
                            email['status'] = 'pending'
                            email['processing'] = False
                        asyncio.run_coroutine_threadsafe(
                            ws_manager.broadcast({
                                "type": "email_process_stopped",
                                "email_id": email_id,
                                "message": "已终止处理"
                            }),
                            websocket_event_loop
                        )
                        return {
                            'email_id': email_id,
                            'status': 'cancelled',
                            'message': '批量处理已终止',
                            'reply': None
                        }
                    
                    write_result = nodes.write_draft_email(state)
                    state.update(write_result)
                    
                    # 检查点：验证前
                    if task_user_state.stop_processing:
                        print(f"⏹️ [批量处理终止] 邮件 {email_id} 在验证前被终止（第{trial+1}次尝试）")
                        with user_lock:
                            email['status'] = 'pending'
                            email['processing'] = False
                        asyncio.run_coroutine_threadsafe(
                            ws_manager.broadcast({
                                "type": "email_process_stopped",
                                "email_id": email_id,
                                "message": "已终止处理"
                            }),
                            websocket_event_loop
                        )
                        return {
                            'email_id': email_id,
                            'status': 'cancelled',
                            'message': '批量处理已终止',
                            'reply': None
                        }
                    
                    verify_result = nodes.verify_generated_email(state)
                    state.update(verify_result)
                    
                    # 检查点：验证后
                    if task_user_state.stop_processing:
                        print(f"⏹️ [批量处理终止] 邮件 {email_id} 在验证后被终止（第{trial+1}次尝试）")
                        with user_lock:
                            email['status'] = 'pending'
                            email['processing'] = False
                        asyncio.run_coroutine_threadsafe(
                            ws_manager.broadcast({
                                "type": "email_process_stopped",
                                "email_id": email_id,
                                "message": "已终止处理"
                            }),
                            websocket_event_loop
                        )
                        return {
                            'email_id': email_id,
                            'status': 'cancelled',
                            'message': '批量处理已终止',
                            'reply': None
                        }
                    
                    if state.get('sendable', False):
                        break
                
                # 5. 获取生成的回复
                generated_reply = state.get('generated_email', '')
                
                # 6. 检查是否自动发送
                auto_send = user_settings.get("autoSend", False)
                final_status = 'processed'
                if auto_send and generated_reply:
                    class EmailObj:
                        def __init__(self, data):
                            self.sender = data.get('sender', '')
                            self.subject = data.get('subject', '')
                            self.messageId = data.get('messageId', '')
                            self.references = data.get('references', '')
                            self.imap_id = data.get('imap_id', b'')
                    
                    email_obj_for_send = EmailObj(email)
                    try:
                        result, message = send_reply_with_rate_limit(
                            current_username,
                            nodes.email_tools,
                            email_obj_for_send,
                            generated_reply,
                            email
                        )
                        if result:
                            final_status = 'sent'
                            sender_name = email.get('sender', '').split('@')[0] if '@' in email.get('sender', '') else email.get('sender', '未知')
                            # 使用锁保护状态更新
                            with user_lock:
                                task_user_state.add_activity('primary', f'自动发送回复给: {sender_name}', 'Message')
                    except Exception as send_err:
                        print(f"❌ [并发处理] 自动发送回复时出错: {send_err}")
                
                # 7. 标记为已读
                imap_id = email.get('imap_id')
                if imap_id:
                    try:
                        nodes.email_tools.mark_email_as_read(imap_id)
                    except:
                        pass
                
                # 8. 使用锁保护状态更新（关键：确保线程安全）
                with user_lock:
                    email['category'] = category
                    email['reply'] = generated_reply
                    email['status'] = final_status
                    task_user_state.stats['processed'] += 1
                    task_user_state.stats['pending'] = max(0, task_user_state.stats['pending'] - 1)
                    task_user_state.history.insert(0, {
                        **email,
                        'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    if not auto_send or not generated_reply or final_status != 'sent':
                        task_user_state.add_activity('success', f'处理了邮件: {category_label}', 'CircleCheck')
                    
                    # 异步生成摘要（不阻塞主流程）
                    email_id = email.get('id')
                    email_body = email.get('body', '')
                    has_body = bool(email_body)
                    has_reply = bool(generated_reply)
                    
                    print(f"🔍 [摘要检查] 邮件 {email_id}: body存在={has_body}, reply存在={has_reply}")
                    
                    if has_body or has_reply:
                        print(f"🚀 [摘要触发] 准备为邮件 {email_id} 生成摘要...")
                        generate_email_summaries_async(
                            current_username,
                            email_id,
                            email_body,
                            generated_reply or ''
                        )
                    else:
                        print(f"⚠️ [摘要跳过] 邮件 {email_id} 没有body和reply，跳过摘要生成")
                    
                print(f"✅ [并发处理] 邮件处理完成: {email.get('subject', '')[:50]}...")
                return {
                    'email_id': email_id,
                        'status': 'processed',
                        'category': category,
                        'message': f"{category_label} - 处理成功",
                        'reply': generated_reply  # 包含生成的回复内容
                    }
                
            except Exception as e:
                print(f"❌ [并发处理] 处理邮件错误: {email.get('subject', '')[:50]}... - {e}")
                import traceback
                traceback.print_exc()
                
                # 使用锁保护状态更新
                with user_lock:
                    email['status'] = 'failed'
                    task_user_state.stats['failed'] += 1
                    task_user_state.history.insert(0, {
                        **email,
                        'status': 'failed',
                        'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                return {
                    'email_id': email_id,
                        'status': 'failed',
                    'message': f"处理失败: {str(e)}",
                    'reply': None  # 失败时没有回复内容
                }
        
        # 将邮件分批，从用户设置中获取每批数量（默认4个）
        user_settings = get_user_settings(current_username)
        batch_size = user_settings.get("batchSize", 4)  # 默认每批4个
        # 限制 batch_size 在合理范围内（1-30，允许更高的并发）
        batch_size = max(1, min(30, int(batch_size)))
        
        # 使用独立的批量处理线程池（不会影响其他API请求）
        batch_pool = get_or_create_batch_thread_pool(batch_size)
        
        total_batches = (len(emails_to_process) + batch_size - 1) // batch_size
        
        print(f"📦 [并发处理] 将 {len(emails_to_process)} 封邮件分成 {total_batches} 批，每批最多 {batch_size} 封（批量线程池大小: {batch_pool._max_workers}，主线程池大小: {thread_pool._max_workers}）")
        
        # 分批并发处理
        for batch_idx in range(total_batches):
            start_idx = batch_idx * batch_size
            end_idx = min(start_idx + batch_size, len(emails_to_process))
            batch_emails = emails_to_process[start_idx:end_idx]
            
            print(f"🔄 [并发处理] 处理第 {batch_idx + 1}/{total_batches} 批，包含 {len(batch_emails)} 封邮件")
            
            # 提交批次内的所有邮件到批量处理线程池（独立线程池，不影响其他操作）
            future_to_email = {
                batch_pool.submit(process_single_email, email): email 
                for email in batch_emails
            }
            
            # 等待批次完成并收集结果
            for future in as_completed(future_to_email):
                email = future_to_email[future]
                try:
                    result = future.result()
                    email_results.append(result)
                    
                    # 更新计数器（使用锁保护）
                    with user_lock:
                        if result['status'] == 'processed':
                            processed_count += 1
                        elif result['status'] == 'skipped':
                            skipped_count += 1
                        elif result['status'] == 'failed':
                            failed_count += 1
                except Exception as e:
                    print(f"❌ [并发处理] 获取处理结果时出错: {e}")
                    import traceback
                    traceback.print_exc()
                    # 使用锁保护状态更新
                    with user_lock:
                        email['status'] = 'failed'
                        task_user_state.stats['failed'] += 1
                        failed_count += 1
                    email_results.append({
                        'email_id': email.get('id'),
                        'status': 'failed',
                        'message': f"处理异常: {str(e)}",
                        'reply': None  # 失败时没有回复内容
                    })
            
            print(f"✅ [并发处理] 第 {batch_idx + 1}/{total_batches} 批处理完成")
        
        # 自动保存数据（处理全部邮件完成后）
        with user_lock:
            save_user_email_data(current_username, task_user_state)
        
        print(f"🎉 [并发处理] 全部处理完成: {processed_count} 封成功, {skipped_count} 封跳过, {failed_count} 封失败")
        
        return {
            "processed": processed_count,
            "skipped": skipped_count,
            "failed": failed_count,
            "email_results": email_results
        }
    
    async def process_all_task():
        """异步包装函数，在线程池中执行同步阻塞的AI操作"""
        try:
            # 在线程池中执行同步阻塞的AI操作，避免阻塞事件循环
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(thread_pool, process_all_sync)
            
            # 处理全部邮件时，不发送单封邮件的 email_process_complete 消息
            # 只发送一条 process_all_complete 汇总消息，避免前端显示多条提示
            # 前端会通过刷新邮件列表来获取最新的邮件状态
            email_results = result.get('email_results', [])
            print(f"📊 [WebSocket发送] 处理全部邮件完成，跳过发送 {len(email_results)} 条单封邮件消息，只发送汇总消息")
            
            # 检查是否有邮件被终止（cancelled状态）
            cancelled_count = sum(1 for r in email_results if r.get('status') == 'cancelled')
            
            # 如果有邮件被终止，发送 process_all_stopped 消息
            if cancelled_count > 0:
                message = f"已终止批量处理: {result['processed']} 封成功, {result['skipped']} 封跳过, {cancelled_count} 封已终止, {result['failed']} 封失败"
                
                await ws_manager.broadcast({
                    "type": "process_all_stopped",
                    "message": message,
                    "processed": result['processed'],
                    "skipped": result['skipped'],
                    "cancelled": cancelled_count,
                    "failed": result['failed']
                })
            else:
                # 正常完成，发送 process_all_complete 消息
                message = f"处理完成: {result['processed']} 封成功, {result['skipped']} 封跳过, {result['failed']} 封失败"
                
                await ws_manager.broadcast({
                    "type": "process_all_complete",
                    "message": message,
                    "processed": result['processed'],
                    "skipped": result['skipped'],
                    "failed": result['failed']
                })
        except Exception as e:
            print(f"异步任务错误: {e}")
            import traceback
            traceback.print_exc()
            # 通知前端处理失败
            await ws_manager.broadcast({
                "type": "process_all_complete",
                "message": f"处理失败: {str(e)}",
                "processed": 0,
                "skipped": 0,
                "failed": 0
            })
    
    # 在后台运行异步任务
    asyncio.create_task(process_all_task())
    
    return {
        "message": f"正在处理 {len(pending_emails)} 封邮件，处理完成后会通知您",
        "count": len(pending_emails)
    }

@app.post("/api/emails/stop-process-all")
async def stop_process_all(current_username: str = Depends(get_username_from_request)):
    """终止全部邮件处理"""
    user_state = get_user_state(current_username)
    user_lock = get_user_lock(current_username)
    
    with user_lock:
        # 设置停止标志
        user_state.stop_processing = True
        
        # 将所有processing状态的邮件设置为stopping（正在终止）
        stopping_count = 0
        for email in user_state.emails_cache:
            if email.get('status') == 'processing':
                email['status'] = 'stopping'
                # processing 保持为 True，让按钮继续显示禁用状态
                stopping_count += 1
        
        # 保存数据
        save_user_email_data(current_username, user_state)
    
    # 通过WebSocket通知前端（状态为 stopping）
    await ws_manager.broadcast({
        "type": "process_all_stopping",
        "message": f"正在终止批量处理，{stopping_count} 封邮件正在终止...",
        "count": stopping_count
    })
    
    # 延迟重置停止标志
    # 增加延迟时间到5分钟（300秒），确保即使是最慢的AI调用也有足够时间检查标志
    async def reset_stop_flag():
        await asyncio.sleep(300)  # 5分钟
        with user_lock:
            user_state.stop_processing = False
            user_state.stopped_email_ids.clear()
            print(f"⏹️ [终止处理] 已重置全局停止标志（5分钟后）")
    
    asyncio.create_task(reset_stop_flag())
    
    print(f"⏹️ [终止处理] 已设置全局停止标志，{stopping_count} 封邮件正在终止（将在5分钟后自动重置）")
    
    return {
        "message": f"正在终止批量处理，{stopping_count} 封邮件正在终止...",
        "stopping": stopping_count
    }

@app.post("/api/emails/{email_id}/stop-process")
async def stop_process_email(email_id: str, current_username: str = Depends(get_username_from_request)):
    """终止单封邮件处理"""
    user_state = get_user_state(current_username)
    user_lock = get_user_lock(current_username)
    
    with user_lock:
        # 添加到终止列表
        user_state.stopped_email_ids.add(email_id)
        
        # 查找邮件，检查是否正在处理
        email_found = False
        for email in user_state.emails_cache:
            if email.get('id') == email_id:
                if email.get('status') == 'processing':
                    # 不立即更新状态为 pending，而是设置为 stopping（正在终止）
                    email['status'] = 'stopping'
                    # processing 保持为 True，让按钮继续显示禁用状态
                    email_found = True
                break
        
        if not email_found:
            return {"message": "邮件未找到或未在处理中", "success": False}
        
        # 保存数据
        save_user_email_data(current_username, user_state)
    
    # 通过WebSocket通知前端（状态为 stopping）
    await ws_manager.broadcast({
        "type": "email_process_stopping",
        "email_id": email_id,
        "message": "正在终止处理..."
    })
    
    # 延迟清除终止标记
    # 增加延迟时间到5分钟（300秒），确保即使是最慢的AI调用也有足够时间检查标志
    # 如果邮件处理真的很慢（比如RAG检索需要1-2分钟），标志也能保留
    async def clear_stop_flag():
        await asyncio.sleep(300)  # 5分钟
        with user_lock:
            user_state.stopped_email_ids.discard(email_id)
            print(f"⏹️ [终止处理] 已清除邮件 {email_id} 的终止标记（5分钟后）")
    
    asyncio.create_task(clear_stop_flag())
    
    print(f"⏹️ [终止处理] 已添加邮件 {email_id} 到终止列表（将在5分钟后自动清除）")
    
    return {
        "message": "正在终止处理...",
        "success": True
    }

@app.post("/api/emails/send")
async def send_reply(request: SendReplyRequest, current_username: str = Depends(get_username_from_request)):
    """发送邮件回复"""
    email_id = request.email_id
    user_state = get_user_state(current_username)
    for email in user_state.emails_cache:
        if email.get('id') == email_id:
            try:
                # 获取当前用户的邮箱配置
                email_address, auth_code = get_user_email_config(current_username)
                email_tools = QQEmailToolsClass(email_address=email_address, auth_code=auth_code)
                
                # 创建邮件对象
                class EmailObj:
                    def __init__(self, data):
                        self.sender = data.get('sender', '')
                        self.subject = data.get('subject', '')
                        self.messageId = data.get('messageId', '')
                        self.references = data.get('references', '')
                        self.imap_id = data.get('imap_id', b'')
                
                email_obj = EmailObj(email)
                result = email_tools.send_reply(email_obj, request.reply)
                
                if result:
                    email['reply'] = request.reply  # 更新为修改后的回复
                    email['status'] = 'sent'
                    
                    # 发送成功后标记为已读并从缓存移除
                    imap_id = email.get('imap_id')
                    if imap_id:
                        try:
                            email_tools.mark_email_as_read(imap_id)
                        except:
                            pass
                    
                    # 更新历史记录中对应邮件的回复内容（如果已存在）
                    # 通过邮件ID或主题+发件人匹配历史记录
                    email_id = email.get('id', '')
                    email_subject = email.get('subject', '')
                    email_sender = email.get('sender', '')
                    
                    print(f"DEBUG [send_reply]: 开始处理发送回复，邮件ID: {email_id}, 主题: {email_subject[:50]}")
                    print(f"DEBUG [send_reply]: 发送前 stats['sent'] = {user_state.stats.get('sent', 0)}")
                    print(f"DEBUG [send_reply]: history 中当前有 {len(user_state.history)} 条记录")
                    
                    history_updated = False
                    for history_record in user_state.history:
                        # 匹配条件：ID相同，或者主题和发件人都相同
                        if (history_record.get('id') == email_id or 
                            (history_record.get('subject') == email_subject and 
                             history_record.get('sender') == email_sender)):
                            # 更新历史记录中的回复内容为修改后的内容
                            print(f"DEBUG [send_reply]: 找到已存在的历史记录，ID: {history_record.get('id')}, 原Status: {history_record.get('status')}")
                            history_record['reply'] = request.reply
                            history_record['status'] = 'sent'
                            print(f"DEBUG [send_reply]: 更新历史记录，新Status: {history_record.get('status')}, ID: {history_record.get('id')}")
                            history_updated = True
                            break
                    else:
                        # 如果历史记录中不存在，则添加到历史记录（使用修改后的回复）
                        # 确保包含所有必要的字段，特别是 id 和 status
                        history_record = {
                            **email,
                            'reply': request.reply,  # 使用修改后的回复
                            'status': 'sent',  # 明确设置为 'sent'
                            'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        # 确保 id 字段存在
                        if not history_record.get('id'):
                            history_record['id'] = email_id
                        user_state.history.insert(0, history_record)
                        print(f"DEBUG [send_reply]: 添加新历史记录，ID: {history_record.get('id')}, Status: {history_record.get('status')}, 主题: {history_record.get('subject', '')[:50]}")
                        history_updated = True
                    
                    user_state.emails_cache.remove(email)
                    print(f"DEBUG [send_reply]: 已从 emails_cache 中移除邮件，当前缓存中有 {len(user_state.emails_cache)} 封邮件")
                    
                    # 更新发送回复数统计（先更新内存中的统计）
                    old_sent_count = user_state.stats.get('sent', 0)
                    user_state.stats['sent'] = old_sent_count + 1
                    print(f"DEBUG [send_reply]: 更新 stats['sent']: {old_sent_count} -> {user_state.stats['sent']}")
                    
                    # 记录操作：发送回复
                    sender_name = email.get('sender', '').split('@')[0] if '@' in email.get('sender', '') else email.get('sender', '未知')
                    user_state.add_activity('primary', f'发送回复给: {sender_name}', 'Message')
                    
                    # 自动保存数据
                    save_user_email_data(current_username, user_state)
                    print(f"DEBUG [send_reply]: 数据已保存，保存后 history 中有 {len(user_state.history)} 条记录")
                    print(f"DEBUG [send_reply]: 保存后，history 中 status='sent' 的记录数: {sum(1 for r in user_state.history if r.get('status') == 'sent')}")
                    
                    return {"message": "回复已发送", "success": True}
                else:
                    raise HTTPException(status_code=500, detail="发送失败")
                    
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"发送失败: {str(e)}")
    
    raise HTTPException(status_code=404, detail="邮件不存在")

@app.post("/api/emails/update-reply")
async def update_reply(request: UpdateReplyRequest, current_username: str = Depends(get_username_from_request)):
    """更新邮件回复内容（编辑后保存，不发送）"""
    email_id = request.email_id
    user_state = get_user_state(current_username)
    
    # 更新邮件缓存中的回复内容
    email_subject = None
    email_sender = None
    for email in user_state.emails_cache:
        if email.get('id') == email_id:
            email['reply'] = request.reply
            email_subject = email.get('subject', '')
            email_sender = email.get('sender', '')
            break
    
    # 更新历史记录中对应邮件的回复内容（如果已存在）
    if email_subject and email_sender:
        for history_record in user_state.history:
            # 匹配条件：ID相同，或者主题和发件人都相同
            if (history_record.get('id') == email_id or 
                (history_record.get('subject') == email_subject and 
                 history_record.get('sender') == email_sender)):
                # 更新历史记录中的回复内容为修改后的内容
                history_record['reply'] = request.reply
                break
    
    # 自动保存数据
    save_user_email_data(current_username, user_state)
    
    return {"message": "回复已更新", "success": True}

@app.delete("/api/emails/{email_id:path}")
async def delete_email(email_id: str, current_username: str = Depends(get_username_from_request)):
    """删除邮件（从缓存中移除）"""
    email_id = unquote(email_id)
    user_state = get_user_state(current_username)
    
    # 从邮件缓存中查找并删除
    email_found = False
    for email in user_state.emails_cache:
        if email.get('id') == email_id:
            email_status = email.get('status', '')
            email_subject = email.get('subject', '')
            user_state.emails_cache.remove(email)
            email_found = True
            
            # 更新统计（如果删除的是待处理邮件，减少待处理数）
            if email_status == 'pending':
                user_state.stats['pending'] = max(0, user_state.stats.get('pending', 0) - 1)
            
            print(f"DEBUG [delete_email]: 删除邮件，ID: {email_id}, Status: {email_status}, 主题: {email_subject[:50]}")
            print(f"DEBUG [delete_email]: 删除后，history 中 status='failed' 的记录数: {sum(1 for r in user_state.history if r.get('status') == 'failed')}")
            
            # 自动保存数据
            save_user_email_data(current_username, user_state)
            break
    
    if email_found:
        return {"message": "邮件已删除", "success": True}
    else:
        raise HTTPException(status_code=404, detail="邮件不存在")

@app.delete("/api/emails")
async def delete_all_emails(current_username: str = Depends(get_username_from_request)):
    """删除所有可删除的邮件（已处理、已跳过、处理失败的邮件），保留未处理的邮件"""
    user_state = get_user_state(current_username)
    
    # 定义可删除的状态
    deletable_statuses = ['processed', 'skipped', 'failed']
    
    # 统计并删除可删除的邮件
    deleted_emails = []
    emails_to_keep = []
    
    for email in user_state.emails_cache:
        email_status = email.get('status', '')
        # 只删除已处理、已跳过、处理失败的邮件
        if email_status in deletable_statuses:
            deleted_emails.append(email)
        else:
            # 保留未处理、处理中的邮件
            emails_to_keep.append(email)
    
    # 更新邮件缓存，只保留未处理的邮件
    user_state.emails_cache = emails_to_keep
    
    # 更新统计数据（只减少已处理的邮件数，不影响待处理数）
    deleted_count = len(deleted_emails)
    if deleted_count > 0:
        # 统计删除的已处理邮件数
        processed_deleted = sum(1 for e in deleted_emails if e.get('status') == 'processed')
        user_state.stats['processed'] = max(0, user_state.stats.get('processed', 0) - processed_deleted)
        
        # 统计删除的失败邮件数
        failed_deleted = sum(1 for e in deleted_emails if e.get('status') == 'failed')
        user_state.stats['failed'] = max(0, user_state.stats.get('failed', 0) - failed_deleted)
    
    # 自动保存数据
    save_user_email_data(current_username, user_state)
    
    print(f"DEBUG [delete_all_emails]: 用户 {current_username} 已删除可删除的邮件，共删除 {deleted_count} 封，保留 {len(emails_to_keep)} 封未处理邮件")
    
    return {
        "message": f"已删除 {deleted_count} 封可删除的邮件，保留 {len(emails_to_keep)} 封未处理邮件",
        "success": True,
        "count": deleted_count,
        "kept": len(emails_to_keep)
    }

@app.post("/api/emails/mark-read")
async def mark_email_read(request: MarkReadRequest, current_username: str = Depends(get_username_from_request)):
    """标记邮件为已读（同步到QQ邮箱）"""
    email_id = request.email_id
    user_state = get_user_state(current_username)
    for email in user_state.emails_cache:
        if email.get('id') == email_id:
            # 同步到QQ邮箱服务器
            imap_id = email.get('imap_id')
            if imap_id:
                try:
                    # 获取当前用户的邮箱配置
                    email_address, auth_code = get_user_email_config(current_username)
                    email_tools = QQEmailToolsClass(email_address=email_address, auth_code=auth_code)
                    email_tools.mark_email_as_read(imap_id)
                except Exception as e:
                    print(f"同步QQ邮箱已读状态失败: {e}")
            
            # 更新本地状态为已读，不立即移除
            # 等用户点击刷新时统一移除，这样可以看到移除了几封已读邮件
            email['status'] = 'read'
            user_state.stats['pending'] = max(0, user_state.stats['pending'] - 1)
            
            # 自动保存数据
            save_user_email_data(current_username, user_state)
            
            return {"message": "已标记为已读（已同步到QQ邮箱）", "success": True}
    raise HTTPException(status_code=404, detail="邮件不存在")

# ==================== 系统控制API ====================

@app.get("/api/system/status")
async def get_system_status(current_username: str = Depends(get_username_from_request)):
    """获取系统状态"""
    user_state = get_user_state(current_username)
    
    return {
        "running": user_state.is_running,
        "autoProcess": user_state.auto_process,
        "lastCheckTime": user_state.last_check_time,
        "checkInterval": user_state.check_interval,
        "emailCount": len(user_state.emails_cache),
        "pendingCount": len([e for e in user_state.emails_cache if e.get('status') == 'pending'])
    }

@app.post("/api/system/start")
async def start_system(current_username: str = Depends(get_username_from_request)):
    """启动邮件监控"""
    user_state = get_user_state(current_username)
    if not user_state.is_running:
        user_state.start_monitor()
        # 记录操作
        user_state.add_activity('success', '启动了邮件监控', 'VideoPlay')
    
    return {"message": "邮件监控已启动", "running": True}

@app.post("/api/system/stop")
async def stop_system(current_username: str = Depends(get_username_from_request)):
    """停止邮件监控"""
    user_state = get_user_state(current_username)
    if user_state.is_running:
        user_state.stop_monitor()
        # 停止监控时，同时关闭自动处理
        user_state.auto_process = False
        # 记录操作
        user_state.add_activity('warning', '停止了邮件监控', 'VideoPause')
    
    return {"message": "邮件监控已停止", "running": False, "autoProcess": False}

@app.post("/api/system/auto-process")
async def toggle_auto_process(enable: bool = True, current_username: str = Depends(get_username_from_request)):
    """开启/关闭自动处理"""
    user_state = get_user_state(current_username)
    user_state.auto_process = enable
    status = "已开启" if enable else "已关闭"
    return {"message": f"自动处理{status}", "autoProcess": enable}

@app.post("/api/system/refresh")
async def refresh_emails(current_username: str = Depends(get_username_from_request)):
    """手动刷新邮件（同步QQ邮箱状态，非阻塞）"""
    try:
        user_state = get_user_state(current_username)
        # 获取当前用户的邮箱配置
        try:
            email_address, auth_code = get_user_email_config(current_username)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        
        # 在线程池中执行同步阻塞的IMAP操作，避免阻塞事件循环
        def fetch_emails_sync():
            """同步获取邮件的函数（在线程池中执行）"""
            email_tools = QQEmailToolsClass(email_address=email_address, auth_code=auth_code)
            return email_tools.fetch_unanswered_emails(max_results=100)
        
        loop = asyncio.get_event_loop()
        emails = await loop.run_in_executor(thread_pool, fetch_emails_sync)
        print(f"DEBUG: 从QQ邮箱获取到 {len(emails)} 封邮件")
        
        # 获取当前QQ邮箱中未读邮件的ID列表
        current_unread_ids = {email_data.get('id', '') for email_data in emails}
        
        # 移除缓存中的邮件：
        # 1. 状态为 read 的邮件（用户点击了"已读"按钮）
        # 2. 状态为 sent 的邮件（已发送的邮件应该从列表中移除）
        # 3. 状态为 pending 且不在QQ邮箱未读列表中的邮件（在QQ邮箱中被手动标记已读）
        # 4. 发件人地址无效的邮件（格式错误，无法发送）
        # 保留：processed、skipped、failed 状态的邮件（但需要验证发件人地址）
        removed_count = 0
        emails_to_remove = []
        for cached_email in user_state.emails_cache:
            cached_id = cached_email.get('id', '')
            cached_status = cached_email.get('status', '')
            cached_sender = cached_email.get('sender', '').strip()
            
            # 检查发件人地址是否有效
            sender_valid = False
            if cached_sender:
                # 清理可能的引号
                cached_sender_clean = cached_sender.strip('"\'')
                # 如果包含 < >，提取邮箱地址部分
                if '<' in cached_sender_clean and '>' in cached_sender_clean:
                    try:
                        cached_sender_clean = cached_sender_clean.split('<')[1].split('>')[0].strip()
                    except (IndexError, AttributeError):
                        pass
                # 验证邮箱地址格式
                if cached_sender_clean and '@' in cached_sender_clean:
                    sender_valid = True
            
            # 移除已读状态的邮件
            if cached_status == 'read':
                emails_to_remove.append(cached_email)
            # 移除已发送状态的邮件
            elif cached_status == 'sent':
                emails_to_remove.append(cached_email)
            # 移除待处理但在QQ邮箱中已被标记已读的邮件
            elif cached_status == 'pending' and cached_id not in current_unread_ids:
                emails_to_remove.append(cached_email)
            # 移除发件人地址无效的邮件（无论状态如何）
            elif not sender_valid:
                print(f"⚠️ [刷新邮件] 发现无效发件人地址的邮件，移除: {cached_email.get('subject', '无主题')[:30]} (发件人: {repr(cached_email.get('sender', ''))})")
                emails_to_remove.append(cached_email)
        
        for email_to_remove in emails_to_remove:
            user_state.emails_cache.remove(email_to_remove)
            removed_count += 1
            print(f"移除邮件: {email_to_remove.get('subject', '')} (状态: {email_to_remove.get('status', '')})")
        
        # 添加新邮件
        new_count = 0
        new_emails_for_summary = []  # 收集需要生成摘要的新邮件
        # 获取缓存中所有邮件的ID集合（用于快速查找）
        cached_ids = {e.get('id', '') for e in user_state.emails_cache}
        
        for email_data in emails:
            email_id = email_data.get('id', '')
            # 检查邮件是否已经在缓存中（通过ID匹配）
            if email_id and email_id not in cached_ids:
                # 自动分类邮件
                subject = email_data.get('subject', '')
                body = email_data.get('body', '')
                category = auto_classify_email(subject, body)
                
                # 检测邮件紧急程度
                try:
                    urgency_level, urgency_keywords = analyze_email_urgency(subject, body)
                except Exception as e:
                    print(f"⚠️ 紧急程度检测失败: {str(e)}")
                    urgency_level = 'low'
                    urgency_keywords = []
                
                # 使用邮件的实际接收时间（如果存在），否则使用当前时间
                email_time = email_data.get('date', '') or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                # 如果时间格式不完整，补充秒数
                if len(email_time) < 19:  # 'YYYY-MM-DD HH:MM:SS' 应该是19个字符
                    email_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                new_email = {
                    **email_data,
                    'time': email_time,
                    'status': 'pending',
                    'category': category,
                    'reply': None,
                    'preview': body[:100] + '...' if body else '',
                    'urgency_level': urgency_level,
                    'urgency_keywords': urgency_keywords
                }
                
                user_state.emails_cache.append(new_email)
                new_count += 1
                
                # 收集需要生成摘要的新邮件
                new_emails_for_summary.append(new_email)
                
                # 判断是否是今天的邮件
                email_date = email_time[:10] if len(email_time) >= 10 else ''
                today = datetime.now().strftime('%Y-%m-%d')
                is_today = email_date == today
                
                if is_today:
                    user_state.stats['today_emails'] += 1
                    print(f"添加新邮件（今日）: {subject[:50]}... (ID: {email_id[:20]}..., 时间: {email_time}, 日期: {email_date})")
                else:
                    print(f"添加新邮件（非今日）: {subject[:50]}... (ID: {email_id[:20]}..., 时间: {email_time}, 日期: {email_date}, 今天: {today})")
                
                user_state.stats['pending'] += 1
            elif email_id:
                # 邮件已存在，但可能需要更新状态（如果之前是已读状态，现在QQ邮箱中又变成未读了）
                for cached_email in user_state.emails_cache:
                    if cached_email.get('id') == email_id:
                        # 如果缓存中的邮件状态是read，但QQ邮箱中还是未读，恢复为pending
                        if cached_email.get('status') == 'read':
                            cached_email['status'] = 'pending'
                            user_state.stats['pending'] += 1
                            print(f"恢复邮件状态为待处理: {cached_email.get('subject', '')[:50]}...")
                        break
        
        # 为新邮件生成原始邮件摘要（异步，不阻塞）
        if new_emails_for_summary:
            print(f"📝 [摘要生成] 开始为 {len(new_emails_for_summary)} 封新邮件生成原始邮件摘要")
            futures = []
            for email in new_emails_for_summary:
                # 在线程池中异步生成摘要，使用批量模式（不立即保存文件）
                future = summary_generation_pool.submit(
                    generate_body_summary_only,
                    email,
                    user_state,
                    current_username,
                    batch_mode=True  # 批量模式，不立即保存
                )
                futures.append(future)
            
            # 在后台等待所有摘要生成完成后统一保存
            def save_after_batch_complete():
                try:
                    # 等待所有任务完成
                    success_count = 0
                    timeout_count = 0
                    error_count = 0
                    
                    for i, future in enumerate(futures, 1):
                        try:
                            if future.result(timeout=150):  # 每个任务最多等待2.5分钟
                                success_count += 1
                        except TimeoutError:
                            timeout_count += 1
                            print(f"⏱️ [摘要生成] 批量任务 {i}/{len(futures)} 超时")
                        except Exception as e:
                            error_count += 1
                            print(f"⚠️ [摘要生成] 批量任务 {i}/{len(futures)} 失败: {type(e).__name__}")
                    
                    # 统一保存一次（即使有部分失败，只要有成功的就保存）
                    if success_count > 0:
                        # 使用用户锁保护批量保存（避免与邮件处理冲突）
                        user_lock = get_user_lock(current_username)
                        with user_lock:
                            save_user_email_data(current_username, user_state)
                            print(f"✅ [摘要生成] 批量保存完成 - 成功: {success_count}, 超时: {timeout_count}, 失败: {error_count}, 总计: {len(futures)}")
                    else:
                        print(f"⚠️ [摘要生成] 所有摘要生成均失败，跳过保存 - 超时: {timeout_count}, 失败: {error_count}")
                except Exception as e:
                    print(f"❌ [摘要生成] 批量保存过程出错: {e}")
                    import traceback
                    traceback.print_exc()
            
            # 提交保存任务到线程池
            summary_generation_pool.submit(save_after_batch_complete)
        
        user_state.last_check_time = datetime.now().isoformat()
        
        # 记录操作：刷新邮件列表
        if new_count > 0 or removed_count > 0:
            user_state.add_activity('info', f'刷新了邮件列表（新增{new_count}封，移除{removed_count}封）', 'Refresh')
        else:
            user_state.add_activity('info', '刷新了邮件列表', 'Refresh')
        
        # 计算当前总邮件数（用于提示信息）
        current_total = len(user_state.emails_cache)
        
        # 自动保存数据
        save_user_email_data(current_username, user_state)
        
        # 生成提示消息
        if new_count > 0 and removed_count > 0:
            message = f"刷新完成，发现 {new_count} 封新邮件，移除 {removed_count} 封已读邮件，当前共 {current_total} 封"
        elif new_count > 0:
            message = f"刷新完成，发现 {new_count} 封新邮件，当前共 {current_total} 封"
        elif removed_count > 0:
            message = f"刷新完成，移除 {removed_count} 封已读邮件，当前共 {current_total} 封"
        else:
            message = f"刷新完成，当前共 {current_total} 封邮件"
        
        return {
            "message": message,
            "newCount": new_count,
            "removedCount": removed_count,
            "totalCount": current_total
        }
    except HTTPException:
        # 重新抛出 HTTP 异常
        raise
    except ValueError as e:
        # 用户未配置邮箱等配置错误
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        # 其他错误
        print(f"刷新邮件失败: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"刷新失败: {str(e)}")

# ==================== 统计API ====================

@app.get("/api/stats")
async def get_stats(current_username: str = Depends(get_username_from_request)):
    """获取统计数据 - 基于真实数据计算，确保准确性"""
    user_state = get_user_state(current_username)
    
    # 计算今日邮件数（使用邮件ID去重，避免重复计算）
    today = datetime.now().strftime('%Y-%m-%d')
    today_email_ids = set()  # 使用集合去重
    
    print(f"DEBUG [get_stats]: 开始计算今日邮件数，今天日期: {today}")
    print(f"DEBUG [get_stats]: emails_cache 中有 {len(user_state.emails_cache)} 封邮件")
    print(f"DEBUG [get_stats]: history 中有 {len(user_state.history)} 条记录")
    
    # 从邮件缓存中统计今天的邮件
    emails_cache_today_count = 0
    emails_cache_today_ids = []
    for email in user_state.emails_cache:
        email_time = email.get('time', '')
        email_id = email.get('id', '')
        if email_id:
            # 使用精确的日期比较：检查时间字符串的前10个字符（YYYY-MM-DD）是否等于今天
            # 这样可以避免 "2024-01-1" 匹配到 "2024-01-10" 的问题
            email_date = email_time[:10] if len(email_time) >= 10 else ''
            if email_date == today:
                today_email_ids.add(email_id)
                emails_cache_today_count += 1
                emails_cache_today_ids.append(email_id[:20] if len(email_id) > 20 else email_id)
                print(f"DEBUG [get_stats]: 邮件缓存中发现今日邮件 - ID: {email_id[:20]}..., 时间: {email_time}, 日期: {email_date}")
            elif email_time:
                # 记录非今日邮件的时间，用于调试
                if emails_cache_today_count < 3:  # 只记录前3个，避免日志过多
                    print(f"DEBUG [get_stats]: 邮件缓存中非今日邮件 - ID: {email_id[:20]}..., 时间: {email_time}, 日期: {email_date}")
    
    print(f"DEBUG [get_stats]: 从 emails_cache 中找到 {emails_cache_today_count} 封今日邮件")
    if emails_cache_today_ids:
        print(f"DEBUG [get_stats]: emails_cache 中今日邮件的 ID 列表（前10个）: {emails_cache_today_ids[:10]}")
    
    # 从历史记录中统计今天的邮件（去重）
    history_today_count = 0
    history_today_ids = []
    for record in user_state.history:
        record_time = record.get('time', '') or record.get('processed_time', '')
        record_id = record.get('id', '')
        if record_id and record_time:
            # 使用精确的日期比较
            record_date = record_time[:10] if len(record_time) >= 10 else ''
            if record_date == today:
                if record_id not in today_email_ids:  # 避免重复计算
                    today_email_ids.add(record_id)
                    history_today_count += 1
                    history_today_ids.append(record_id[:20] if len(record_id) > 20 else record_id)
                    print(f"DEBUG [get_stats]: 历史记录中发现今日邮件 - ID: {record_id[:20]}..., 时间: {record_time}, 日期: {record_date}")
            elif record_time:
                # 记录非今日记录的时间，用于调试
                if history_today_count < 3:  # 只记录前3个，避免日志过多
                    print(f"DEBUG [get_stats]: 历史记录中非今日记录 - ID: {record_id[:20]}..., 时间: {record_time}, 日期: {record_date}")
    
    print(f"DEBUG [get_stats]: 从 history 中找到 {history_today_count} 条今日记录（去重后新增）")
    if history_today_ids:
        print(f"DEBUG [get_stats]: history 中今日记录的 ID 列表（前10个）: {history_today_ids[:10]}")
    
    today_emails_count = len(today_email_ids)
    print(f"DEBUG [get_stats]: 今日邮件总数（去重后）: {today_emails_count} (emails_cache: {emails_cache_today_count}, history新增: {history_today_count})")
    
    # 计算已处理数（同时从邮件缓存和历史记录中统计，去重）
    # 已处理包括：processed（已生成回复）、sent（已发送）、skipped（无关邮件已跳过）
    processed_email_ids = set()
    
    # 从邮件缓存中统计已处理的邮件
    for email in user_state.emails_cache:
        email_id = email.get('id', '')
        email_status = email.get('status', '')
        if email_id and email_status in ['processed', 'sent', 'skipped']:
            processed_email_ids.add(email_id)
    
    # 从历史记录中统计已处理的邮件（去重）
    for record in user_state.history:
        record_id = record.get('id', '')
        record_status = record.get('status', '')
        if record_id and record_status in ['success', 'processed', 'sent', 'skipped']:
            processed_email_ids.add(record_id)
    
    processed_count = len(processed_email_ids)
    
    # 计算待处理数（从邮件缓存中统计待处理的邮件）
    pending_count = len([e for e in user_state.emails_cache if e.get('status') == 'pending'])
    
    # 计算失败数（同时从邮件缓存和历史记录中统计，去重）
    failed_email_ids = set()
    
    # 从邮件缓存中统计失败的邮件
    for email in user_state.emails_cache:
        email_id = email.get('id', '')
        email_status = email.get('status', '')
        if email_id and email_status == 'failed':
            failed_email_ids.add(email_id)
    
    # 从历史记录中统计失败的邮件（去重）
    for record in user_state.history:
        record_id = record.get('id', '')
        record_status = record.get('status', '')
        if record_id and record_status == 'failed':
            failed_email_ids.add(record_id)
    
    failed_count = len(failed_email_ids)
    print(f"DEBUG [get_stats]: 计算失败数: emails_cache中有 {sum(1 for e in user_state.emails_cache if e.get('status') == 'failed')} 封, history中有 {sum(1 for r in user_state.history if r.get('status') == 'failed')} 封, 去重后共 {failed_count} 封")
    
    # 计算发送回复数（只统计实际发送的邮件，status为'sent'）
    print(f"DEBUG [get_stats]: 开始计算发送回复数")
    print(f"DEBUG [get_stats]: 当前内存中 stats['sent'] = {user_state.stats.get('sent', 0)}")
    print(f"DEBUG [get_stats]: emails_cache 中有 {len(user_state.emails_cache)} 封邮件")
    print(f"DEBUG [get_stats]: history 中有 {len(user_state.history)} 条记录")
    
    sent_email_ids = set()
    
    # 从邮件缓存中统计已发送的邮件
    emails_cache_sent_count = 0
    for email in user_state.emails_cache:
        email_id = email.get('id', '')
        email_status = email.get('status', '')
        if email_id and email_status == 'sent':
            sent_email_ids.add(email_id)
            emails_cache_sent_count += 1
    print(f"DEBUG [get_stats]: emails_cache 中 status='sent' 的邮件: {emails_cache_sent_count} 封")
    
    # 从历史记录中统计已发送的邮件（去重）
    history_sent_count = 0
    history_sent_ids = []
    for record in user_state.history:
        record_id = record.get('id', '')
        record_status = record.get('status', '')
        if record_id and record_status == 'sent':
            sent_email_ids.add(record_id)
            history_sent_count += 1
            history_sent_ids.append(record_id[:20] if len(record_id) > 20 else record_id)
        elif record_status == 'sent':
            # 如果 status 是 'sent' 但 id 为空，记录警告
            print(f"DEBUG [get_stats]: 警告：发现 status='sent' 但 id 为空的记录: {record.get('subject', '')[:50]}")
    
    print(f"DEBUG [get_stats]: history 中 status='sent' 的记录: {history_sent_count} 条")
    if history_sent_ids:
        print(f"DEBUG [get_stats]: history 中 sent 记录的 ID 列表（前10个）: {history_sent_ids[:10]}")
    
    sent_count = len(sent_email_ids)
    print(f"DEBUG [get_stats]: 去重后的发送回复数: {sent_count} (emails_cache: {emails_cache_sent_count}, history: {history_sent_count})")
    
    # 如果内存中的 stats['sent'] 更大，说明有刚刚发送的邮件还没被统计到，使用内存中的值
    sent_count_memory = user_state.stats.get('sent', 0)
    if sent_count_memory > sent_count:
        print(f"DEBUG [get_stats]: 使用内存中的 stats['sent'] = {sent_count_memory}（大于计算值 {sent_count}）")
        sent_count = sent_count_memory
    elif sent_count_memory < sent_count:
        print(f"DEBUG [get_stats]: 计算值 {sent_count} 大于内存值 {sent_count_memory}，使用计算值")
    else:
        print(f"DEBUG [get_stats]: 计算值与内存值一致: {sent_count}")
    
    # 计算本月处理数（从本月1号到今天的所有已处理邮件）
    # 已处理包括：processed（已生成回复）、sent（已发送）、skipped（无关邮件已跳过）
    now = datetime.now()
    current_month_start = datetime(now.year, now.month, 1).strftime('%Y-%m-%d')
    this_month_processed_ids = set()
    
    print(f"DEBUG [get_stats]: 开始计算本月处理数，本月起始日期: {current_month_start}")
    
    # 从邮件缓存中统计本月的已处理邮件
    for email in user_state.emails_cache:
        email_id = email.get('id', '')
        email_time = email.get('time', '')
        email_status = email.get('status', '')
        if email_id and email_status in ['processed', 'sent', 'skipped']:
            email_date = email_time[:10] if len(email_time) >= 10 else ''
            # 检查是否在本月（日期 >= 本月1号）
            if email_date >= current_month_start:
                this_month_processed_ids.add(email_id)
    
    # 从历史记录中统计本月的已处理邮件（去重）
    for record in user_state.history:
        record_id = record.get('id', '')
        record_time = record.get('time', '') or record.get('processed_time', '')
        record_status = record.get('status', '')
        if record_id and record_status in ['success', 'processed', 'sent', 'skipped']:
            record_date = record_time[:10] if len(record_time) >= 10 else ''
            # 检查是否在本月（日期 >= 本月1号）
            if record_date >= current_month_start:
                if record_id not in this_month_processed_ids:
                    this_month_processed_ids.add(record_id)
    
    this_month_processed_count = len(this_month_processed_ids)
    print(f"DEBUG [get_stats]: 本月处理数: {this_month_processed_count} (本月起始: {current_month_start})")
    
    # 同时更新内存中的统计数据，保持一致性
    user_state.stats['today_emails'] = today_emails_count
    user_state.stats['processed'] = processed_count
    user_state.stats['pending'] = pending_count
    user_state.stats['failed'] = failed_count
    user_state.stats['sent'] = sent_count
    print(f"DEBUG [get_stats]: 最终更新 stats['sent'] = {sent_count}")
    
    return {
        "todayEmails": today_emails_count,
        "processed": processed_count,
        "pending": pending_count,
        "failed": failed_count,
        "sentReplies": sent_count,  # 返回实际发送回复数
        "thisMonthProcessed": this_month_processed_count  # 返回本月处理数
    }

@app.get("/api/stats/category")
async def get_category_stats(current_username: str = Depends(get_username_from_request)):
    """获取分类统计 - 只统计今天的数据，确保用户隔离"""
    user_state = get_user_state(current_username)
    categories = {}
    
    # 获取今天的日期字符串
    today = datetime.now().strftime('%Y-%m-%d')
    
    # 统计邮件缓存中的分类（使用邮件ID去重，只统计今天的）
    processed_email_ids = set()  # 用于去重
    
    # 统计邮件缓存中今天的分类
    for email in user_state.emails_cache:
        email_id = email.get('id', '')
        email_time = email.get('time', '')
        cat = email.get('category', 'unknown')
        # 只统计今天的邮件
        email_date = email_time[:10] if len(email_time) >= 10 else ''
        if email_id and cat and cat != 'unknown' and email_date == today:
            if email_id not in processed_email_ids:
                processed_email_ids.add(email_id)
                categories[cat] = categories.get(cat, 0) + 1
    
    # 统计历史记录中今天的分类（去重）
    for record in user_state.history:
        record_id = record.get('id', '')
        record_time = record.get('time', '') or record.get('processed_time', '')
        cat = record.get('category', 'unknown')
        # 只统计今天的记录
        record_date = record_time[:10] if len(record_time) >= 10 else ''
        if record_id and cat and cat != 'unknown' and record_date == today:
            if record_id not in processed_email_ids:
                processed_email_ids.add(record_id)
                categories[cat] = categories.get(cat, 0) + 1
    
    # 只返回今天的数据
    return {"categories": categories}

@app.get("/api/stats/trend")
async def get_trend_stats(days: int = 7, current_username: str = Depends(get_username_from_request)):
    """获取趋势数据 - 使用真实数据，确保用户隔离"""
    user_state = get_user_state(current_username)
    
    # 从用户历史记录和邮件缓存中生成趋势数据（使用邮件ID去重）
    trend_data = []
    for i in range(days - 1, -1, -1):
        date = (datetime.now() - timedelta(days=i)).strftime('%m-%d')
        date_str = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        
        # 统计当天的邮件（使用集合去重）
        received_email_ids = set()
        processed_email_ids = set()
        
        # 统计邮件缓存中当天的邮件
        for email in user_state.emails_cache:
            email_id = email.get('id', '')
            email_time = email.get('time', '')
            if email_id and date_str in email_time:
                received_email_ids.add(email_id)
                if email.get('status') in ['processed', 'sent']:
                    processed_email_ids.add(email_id)
        
        # 统计历史记录中当天的邮件（去重）
        for record in user_state.history:
            record_id = record.get('id', '')
            record_time = record.get('time', '') or record.get('processed_time', '')
            if record_id and date_str in record_time:
                if record_id not in received_email_ids:
                    received_email_ids.add(record_id)
                if record.get('status') in ['success', 'processed', 'sent']:
                    processed_email_ids.add(record_id)
        
        # 只返回真实数据，不使用模拟数据
        trend_data.append({
            "date": date,
            "received": len(received_email_ids),
            "processed": len(processed_email_ids)
        })
    
    return {"trend": trend_data}

# ==================== 历史记录API ====================

@app.get("/api/history")
async def get_history(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    page_size: int = 10,
    current_username: str = Depends(get_username_from_request)
):
    """获取处理记录 - 使用真实数据，确保用户隔离"""
    user_state = get_user_state(current_username)
    records = user_state.history.copy()
    
    # 只返回真实数据，不使用模拟数据
    # 如果没有记录，返回空数组
    
    # 日期筛选（只比较日期部分，忽略时间）
    if start_date:
        records = [r for r in records if (r.get('time', '') and r.get('time', '')[:10] >= start_date) or 
                                          (r.get('processed_time', '') and r.get('processed_time', '')[:10] >= start_date)]
    if end_date:
        # 结束日期需要包含当天，所以比较日期部分
        records = [r for r in records if (r.get('time', '') and r.get('time', '')[:10] <= end_date) or 
                                          (r.get('processed_time', '') and r.get('processed_time', '')[:10] <= end_date)]
    
    # 分类筛选
    if category:
        records = [r for r in records if r.get('category') == category]
    
    # 状态筛选
    if status:
        # 状态映射：success/processed/sent -> success, failed -> failed, skipped -> skipped
        if status == 'success':
            records = [r for r in records if r.get('status') in ['success', 'processed', 'sent']]
        elif status == 'failed':
            records = [r for r in records if r.get('status') == 'failed']
        elif status == 'skipped':
            records = [r for r in records if r.get('status') == 'skipped']
        elif status == 'pending':
            records = [r for r in records if r.get('status') not in ['success', 'processed', 'sent', 'failed', 'skipped']]
    
    # 按时间倒序排序（最新的在前）
    records.sort(key=lambda x: x.get('time', '') or x.get('processed_time', ''), reverse=True)
    
    # 分页
    total = len(records)
    start = (page - 1) * page_size
    end = start + page_size
    records = records[start:end]
    
    return {
        "records": records,
        "total": total,
        "page": page,
        "pageSize": page_size
    }

@app.get("/api/history/export")
async def export_history(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    current_username: str = Depends(get_username_from_request)
):
    """导出处理记录为XLSX文件"""
    user_state = get_user_state(current_username)
    records = user_state.history.copy()
    
    print(f"📊 [导出XLSX] 开始导出，原始记录数: {len(records)}")
    print(f"📊 [导出XLSX] 筛选条件: start_date={start_date}, end_date={end_date}, category={category}, status={status}")
    
    # 应用相同的筛选逻辑（与get_history保持一致）
    # 日期筛选（只比较日期部分，忽略时间）
    if start_date:
        original_count = len(records)
        records = [r for r in records if (r.get('time', '') and r.get('time', '')[:10] >= start_date) or 
                                          (r.get('processed_time', '') and r.get('processed_time', '')[:10] >= start_date)]
        print(f"📊 [导出XLSX] 开始日期筛选后: {original_count} -> {len(records)}")
    if end_date:
        original_count = len(records)
        records = [r for r in records if (r.get('time', '') and r.get('time', '')[:10] <= end_date) or 
                                          (r.get('processed_time', '') and r.get('processed_time', '')[:10] <= end_date)]
        print(f"📊 [导出XLSX] 结束日期筛选后: {original_count} -> {len(records)}")
    
    # 分类筛选
    if category:
        original_count = len(records)
        records = [r for r in records if r.get('category') == category]
        print(f"📊 [导出XLSX] 分类筛选后: {original_count} -> {len(records)}")
    
    # 状态筛选
    if status:
        original_count = len(records)
        if status == 'success':
            records = [r for r in records if r.get('status') in ['success', 'processed', 'sent']]
        elif status == 'failed':
            records = [r for r in records if r.get('status') == 'failed']
        elif status == 'skipped':
            records = [r for r in records if r.get('status') == 'skipped']
        elif status == 'pending':
            records = [r for r in records if r.get('status') not in ['success', 'processed', 'sent', 'failed', 'skipped']]
        print(f"📊 [导出XLSX] 状态筛选后: {original_count} -> {len(records)}")
    
    # 按时间倒序排序（最新的在前）
    records.sort(key=lambda x: x.get('time', '') or x.get('processed_time', ''), reverse=True)
    
    print(f"📊 [导出XLSX] 最终记录数: {len(records)}")
    
    # 分类和状态映射
    category_names = {
        'product_enquiry': '产品咨询',
        'customer_complaint': '客户投诉',
        'customer_feedback': '客户反馈',
        'unrelated': '无关邮件'
    }
    
    status_names = {
        'success': '成功',
        'processed': '成功',
        'sent': '成功',
        'failed': '失败',
        'skipped': '跳过'
    }
    
    # 使用openpyxl生成XLSX文件
    if OPENPYXL_AVAILABLE:
        wb = Workbook()
        ws = wb.active
        ws.title = "处理记录"
        
        # 设置表头
        headers = ['时间', '发件人', '主题', '分类', '状态', '回复内容']
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        row_count = 0
        for record in records:
            # 解析时间字符串为datetime对象
            time_value = None
            time_str = record.get('time') or record.get('processed_time', '')
            if time_str:
                try:
                    time_str = str(time_str).strip()
                    # 尝试多种格式解析
                    parsed = False
                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M', '%Y/%m/%d']:
                        try:
                            time_value = datetime.strptime(time_str, fmt)
                            parsed = True
                            break
                        except:
                            continue
                    if not parsed:
                        time_value = None
                except Exception as e:
                    print(f"⚠️ [导出XLSX] 时间解析失败: {time_str}, 错误: {e}")
                    time_value = None
            
            sender = record.get('sender') or record.get('sender_email', '')
            subject = record.get('subject', '')
            category_val = record.get('category', '')
            category_label = category_names.get(category_val, category_val or '未分类')
            status_val = record.get('status', '')
            status_label = status_names.get(status_val, status_val or '未知')
            reply = record.get('reply', '')
            
            # 写入行数据
            row_data = [
                time_value if time_value else time_str,  # 如果是datetime对象，直接使用；否则使用字符串
                sender,
                subject,
                category_label,
                status_label,
                reply
            ]
            ws.append(row_data)
            
            # 设置时间列的格式（如果是datetime对象）
            if time_value:
                time_cell = ws.cell(row=row_count + 2, column=1)  # +2因为第一行是表头
                time_cell.number_format = 'yyyy-mm-dd hh:mm:ss'
            
            row_count += 1
        
        # 自动调整列宽
        column_widths = [20, 25, 40, 15, 10, 60]  # 时间、发件人、主题、分类、状态、回复内容
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # 设置文本自动换行（回复内容列）
        reply_col = ws.column_dimensions[get_column_letter(6)]  # 回复内容在第6列
        for row in range(2, row_count + 2):  # 从第2行开始（跳过表头）
            cell = ws.cell(row=row, column=6)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        xlsx_content = output.getvalue()
        output.close()
        
        print(f"📊 [导出XLSX] 已写入 {row_count} 行数据（不包括表头）")
        
        # 生成文件名
        filename = f"processing_records_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filename_utf8 = f"处理记录_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filename_encoded = quote(filename_utf8.encode('utf-8'))
        
        # 返回XLSX文件
        return Response(
            content=xlsx_content,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename_encoded}'
            }
        )
    else:
        # 如果没有openpyxl，回退到CSV格式
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['时间', '发件人', '主题', '分类', '状态', '回复内容'])
        
        # 写入数据
        row_count = 0
        for record in records:
            time_str = record.get('time') or record.get('processed_time', '')
            sender = record.get('sender') or record.get('sender_email', '')
            subject = record.get('subject', '')
            category_val = record.get('category', '')
            category_label = category_names.get(category_val, category_val or '未分类')
            status_val = record.get('status', '')
            status_label = status_names.get(status_val, status_val or '未知')
            reply = record.get('reply', '')
            
            writer.writerow([time_str, sender, subject, category_label, status_label, reply])
            row_count += 1
        
        csv_content = output.getvalue()
        output.close()
        
        filename = f"processing_records_{datetime.now().strftime('%Y%m%d')}.csv"
        filename_utf8 = f"处理记录_{datetime.now().strftime('%Y%m%d')}.csv"
        filename_encoded = quote(filename_utf8.encode('utf-8'))
        
        return Response(
            content=csv_content.encode('utf-8-sig'),
            media_type='text/csv; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename_encoded}'
            }
        )

@app.post("/api/history/clear")
async def clear_history(
    before_date: Optional[str] = None,
    current_username: str = Depends(get_username_from_request)
):
    """
    清理历史记录
    
    参数:
        before_date: 删除此日期之前的记录，格式 'YYYY-MM-DD'。如果不提供，则删除所有记录
    
    返回:
        deleted_count: 删除的记录数量
    """
    try:
        user_state = get_user_state(current_username)
        original_count = len(user_state.history)
        
        if original_count == 0:
            return {
                "success": True,
                "message": "没有历史记录",
                "deleted_count": 0,
                "remaining_count": 0
            }
        
        if before_date:
            # 删除指定日期之前的记录
            filtered_history = []
            deleted_count = 0
            
            for record in user_state.history:
                record_time = record.get('time') or record.get('processed_time', '')
                record_date = record_time[:10] if record_time else ''
                
                if record_date and record_date < before_date:
                    deleted_count += 1
                else:
                    filtered_history.append(record)
            
            user_state.history = filtered_history
        else:
            # 删除所有记录
            deleted_count = original_count
            user_state.history = []
        
        # 保存数据
        save_user_email_data(current_username, user_state)
        
        remaining_count = len(user_state.history)
        
        print(f"✅ [历史记录清理] 用户 {current_username}: 删除 {deleted_count} 条记录，保留 {remaining_count} 条")
        
        return {
            "success": True,
            "message": f"成功删除 {deleted_count} 条记录",
            "deleted_count": deleted_count,
            "remaining_count": remaining_count
        }
        
    except Exception as e:
        print(f"❌ [历史记录清理] 清理失败: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"清理失败: {str(e)}")

# ==================== 最近操作API ====================

@app.get("/api/activities")
async def get_activities(
    limit: int = 10,
    current_username: str = Depends(get_username_from_request)
):
    """获取用户最近操作记录"""
    user_state = get_user_state(current_username)
    activities = user_state.activities.copy()
    
    # 限制返回数量
    if limit > 0:
        activities = activities[:limit]
    
    # 如果没有操作记录，返回空数组
    return {"activities": activities}

# ==================== 知识库API ====================

# 导入索引构建模块
from src.knowledge_index import get_data_dir, build_index, get_db_path

# 知识库数据目录配置（可通过环境变量KNOWLEDGE_DATA_DIR配置）
KNOWLEDGE_DATA_DIR = os.getenv("KNOWLEDGE_DATA_DIR", "data")

@app.get("/api/knowledge/documents")
async def get_documents():
    """获取知识库文档列表"""
    # 读取data目录下的文件
    documents = []
    data_dir = get_data_dir()
    
    # 确保data目录存在
    if not os.path.exists(data_dir):
        os.makedirs(data_dir, exist_ok=True)
    
    if os.path.exists(data_dir):
        for filename in os.listdir(data_dir):
            filepath = os.path.join(data_dir, filename)
            if os.path.isfile(filepath):
                stat = os.stat(filepath)
                # 检查是否已索引（检查db目录中是否有对应的向量数据）
                indexed = os.path.exists("db") or os.path.exists("db_1024") or os.path.exists("db_2560") or os.path.exists("db_4096")
                documents.append({
                    "id": filename,
                    "name": filename,
                    "size": f"{stat.st_size / 1024:.1f}KB",
                    "updateTime": datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d'),
                    "indexed": indexed
                })
    
    # 按更新时间倒序排序
    documents.sort(key=lambda x: x['updateTime'], reverse=True)
    
    return {"documents": documents}

@app.post("/api/knowledge/upload")
async def upload_document(file: UploadFile = File(...), auto_index: bool = Query(False)):
    """上传文档到知识库"""
    try:
        # 确保data目录存在
        data_dir = get_data_dir()
        if not os.path.exists(data_dir):
            os.makedirs(data_dir, exist_ok=True)
        
        # 验证文件类型
        allowed_extensions = ['.txt', '.pdf', '.docx', '.md']
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            raise HTTPException(status_code=400, detail=f"不支持的文件类型: {file_ext}，支持的类型: {', '.join(allowed_extensions)}")
        
        # 保存文件
        filepath = os.path.join(data_dir, file.filename)
        with open(filepath, "wb") as f:
            content = await file.read()
            f.write(content)
        
        print(f"✅ [知识库] 文档上传成功: {file.filename}, 大小: {len(content)} 字节")
        
        result = {
            "message": f"文档 {file.filename} 上传成功",
            "filename": file.filename,
            "success": True
        }
        
        # 如果启用了自动索引，则自动重建索引
        if auto_index:
            print(f"🔄 [知识库] 自动重建索引...")
            try:
                # 在后台线程中执行索引构建（避免阻塞）
                def build_index_background():
                    index_result = build_index(specific_file=file.filename)
                    if index_result.get("success"):
                        print(f"✅ [知识库] 自动索引重建成功: {file.filename}")
                    else:
                        print(f"⚠️ [知识库] 自动索引重建失败: {index_result.get('error')}")
                
                # 使用线程池执行
                thread_pool.submit(build_index_background)
                result["message"] = f"文档 {file.filename} 上传成功，正在后台重建索引..."
                result["indexing"] = True
            except Exception as e:
                print(f"⚠️ [知识库] 自动索引重建失败: {e}")
                result["indexing"] = False
                result["index_error"] = str(e)
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ [知识库] 文档上传失败: {e}")
        raise HTTPException(status_code=500, detail=f"上传失败: {str(e)}")

@app.get("/api/knowledge/documents/{doc_id}/preview")
async def preview_document(doc_id: str):
    """预览文档内容"""
    try:
        # 防止路径遍历攻击（只检查路径遍历模式，允许文件名中包含点）
        if '/' in doc_id or '\\' in doc_id:
            raise HTTPException(status_code=400, detail="无效的文件名：不能包含路径分隔符")
        if '../' in doc_id or '..\\' in doc_id or doc_id.startswith('..'):
            raise HTTPException(status_code=400, detail="无效的文件名：不能包含路径遍历模式")
        
        filepath = os.path.join(get_data_dir(), doc_id)
        # 确保文件路径在数据目录内（防止路径遍历攻击）
        filepath = os.path.abspath(filepath)
        data_dir_abs = os.path.abspath(get_data_dir())
        if not filepath.startswith(data_dir_abs):
            raise HTTPException(status_code=400, detail="无效的文件路径")
        
        if not os.path.exists(filepath):
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 只支持文本文件预览
        if not doc_id.lower().endswith(('.txt', '.md')):
            return {
                "content": f"文件 {doc_id} 不是文本文件，无法预览。\n支持预览的文件类型: .txt, .md",
                "filename": doc_id,
                "previewable": False
            }
        
        # 读取文件内容（限制大小，避免读取过大文件）
        max_size = 1024 * 1024  # 1MB
        file_size = os.path.getsize(filepath)
        
        if file_size > max_size:
            return {
                "content": f"文件过大（{file_size / 1024:.1f}KB），无法预览。最大预览大小: 1MB",
                "filename": doc_id,
                "previewable": False
            }
        
        # 读取文件内容
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
        except UnicodeDecodeError:
            # 如果UTF-8解码失败，尝试其他编码
            try:
                with open(filepath, 'r', encoding='gbk') as f:
                    content = f.read()
            except:
                content = "文件编码不支持，无法预览"
        
        # 限制预览长度（前5000字符）
        if len(content) > 5000:
            content = content[:5000] + "\n\n... (内容过长，仅显示前5000字符)"
        
        return {
            "content": content,
            "filename": doc_id,
            "previewable": True
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ [知识库] 预览文档失败: {e}")
        raise HTTPException(status_code=500, detail=f"预览失败: {str(e)}")

@app.get("/api/knowledge/documents/{doc_id}/download")
async def download_document(doc_id: str):
    """下载文档"""
    try:
        # 防止路径遍历攻击（只检查路径遍历模式，允许文件名中包含点）
        if '/' in doc_id or '\\' in doc_id:
            raise HTTPException(status_code=400, detail="无效的文件名：不能包含路径分隔符")
        if '../' in doc_id or '..\\' in doc_id or doc_id.startswith('..'):
            raise HTTPException(status_code=400, detail="无效的文件名：不能包含路径遍历模式")
        
        filepath = os.path.join(get_data_dir(), doc_id)
        # 确保文件路径在数据目录内（防止路径遍历攻击）
        filepath = os.path.abspath(filepath)
        data_dir_abs = os.path.abspath(get_data_dir())
        if not filepath.startswith(data_dir_abs):
            raise HTTPException(status_code=400, detail="无效的文件路径")
        
        if not os.path.exists(filepath):
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 读取文件内容
        with open(filepath, "rb") as f:
            content = f.read()
        
        # 获取文件MIME类型
        import mimetypes
        mime_type, _ = mimetypes.guess_type(filepath)
        if not mime_type:
            mime_type = 'application/octet-stream'
        
        print(f"✅ [知识库] 文档下载: {doc_id}")
        
        from urllib.parse import quote
        filename_encoded = quote(doc_id.encode('utf-8'))
        
        return Response(
            content=content,
            media_type=mime_type,
            headers={
                'Content-Disposition': f'attachment; filename="{doc_id}"; filename*=UTF-8\'\'{filename_encoded}'
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ [知识库] 文档下载失败: {e}")
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")

@app.delete("/api/knowledge/documents/{doc_id}")
async def delete_document(doc_id: str):
    """删除文档"""
    try:
        # 防止路径遍历攻击（只检查路径遍历模式，允许文件名中包含点）
        # 检查是否包含路径分隔符或路径遍历模式
        if '/' in doc_id or '\\' in doc_id:
            raise HTTPException(status_code=400, detail="无效的文件名：不能包含路径分隔符")
        
        # 检查路径遍历攻击（../ 或 ..\）
        if '../' in doc_id or '..\\' in doc_id or doc_id.startswith('..'):
            raise HTTPException(status_code=400, detail="无效的文件名：不能包含路径遍历模式")
        
        filepath = os.path.join(get_data_dir(), doc_id)
        
        # 确保文件路径在数据目录内（防止路径遍历攻击）
        filepath = os.path.abspath(filepath)
        data_dir_abs = os.path.abspath(get_data_dir())
        if not filepath.startswith(data_dir_abs):
            raise HTTPException(status_code=400, detail="无效的文件路径")
        
        if not os.path.exists(filepath):
            raise HTTPException(status_code=404, detail="文档不存在")
        
        # 检查是否是文件（不是目录）
        if not os.path.isfile(filepath):
            raise HTTPException(status_code=400, detail="只能删除文件，不能删除目录")
        
        # 删除文件
        try:
            os.remove(filepath)
        except PermissionError as pe:
            raise HTTPException(status_code=403, detail=f"删除失败：文件被占用或没有权限。错误: {str(pe)}")
        except OSError as oe:
            raise HTTPException(status_code=500, detail=f"删除失败：{str(oe)}")
        
        print(f"✅ [知识库] 文档删除成功: {doc_id}")
        
        return {
            "message": f"文档 {doc_id} 删除成功",
            "success": True
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ [知识库] 文档删除失败: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")

@app.post("/api/knowledge/test/cancel")
async def cancel_rag_test(current_username: str = Depends(get_username_from_request)):
    """取消正在进行的RAG测试"""
    cancel_flag = get_rag_cancel_flag(current_username)
    cancel_flag.set()
    print(f"🚫 [RAG测试] 用户 {current_username} 请求取消检索")
    return {"success": True, "message": "取消请求已发送"}

@app.post("/api/knowledge/test")
async def test_rag(
    request: RAGTestRequest, 
    current_username: str = Depends(get_username_from_request),
    http_request: Request = None
):
    """测试RAG检索（使用与处理邮件相同的逻辑）"""
    import asyncio
    from src.nodes import Nodes
    from src.state import Email
    
    # 获取取消标志（按用户）
    cancel_flag = get_rag_cancel_flag(current_username)
    cancel_flag.clear()  # 开始新的检索前，清除之前的取消标志
    
    def run_rag_test_sync():
        """同步执行RAG测试的函数（在线程池中执行，避免阻塞事件循环）"""
        try:
            # 在关键步骤检查是否已取消
            if cancel_flag.is_set():
                print(f"🚫 [RAG测试] 检测到取消信号，停止检索")
                cancel_result = {
                    "question": request.question,
                    "answer": "检索已取消",
                    "success": False
                }
                # 通过 WebSocket 通知前端检索已取消
                try:
                    message = {
                        "type": "rag_test_complete",
                        "question": request.question,
                        "answer": "检索已取消",
                        "success": False,
                        "cancelled": True
                    }
                    import asyncio
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    loop.run_until_complete(ws_manager.broadcast(message))
                    loop.close()
                    print(f"📢 [RAG测试] 已通过 WebSocket 通知前端检索已取消")
                except Exception as ws_error:
                    print(f"⚠️ [RAG测试] WebSocket 通知失败: {ws_error}")
                return cancel_result
            print(f"🔍 [RAG测试] 开始测试，问题: {request.question}")
            
            # 获取用户设置，使用与处理邮件相同的逻辑
            user_settings = get_user_settings(current_username)
            reply_model = user_settings.get("replyModel", user_settings.get("model", "moonshotai/Kimi-K2-Thinking"))
            embedding_model = user_settings.get("embeddingModel", "Qwen/Qwen3-Embedding-4B")
            # 获取模型配置
            models_config = get_models_config(current_username, reply_model, embedding_model)
            
            # 使用Nodes类，与处理邮件时完全一致
            nodes = Nodes(
                email_address=None,  # RAG测试不需要邮箱配置
                auth_code=None,
                api_key=models_config["apiKey"],
                reply_model=reply_model,
                embedding_model=embedding_model,
                signature=user_settings.get("signature"),
                greeting=user_settings.get("greeting"),
                closing=user_settings.get("closing"),
                reply_api_base=models_config["replyApiBaseUrl"],
                embedding_api_base=models_config["embeddingApiBaseUrl"]
            )
            
            print(f"🔍 [RAG测试] 使用用户配置: replyModel={reply_model}, embeddingModel={embedding_model}")
            print(f"🔍 [RAG测试] Nodes初始化成功，开始检索...")
            
            # 创建一个模拟的Email对象用于RAG检索（与处理邮件时相同的方式）
            email_obj = Email(
                id="rag_test",
                threadId="",
                messageId="",
                references="",
                sender="test@example.com",
                subject="RAG测试",
                body=request.question,  # 使用问题作为邮件内容
                imap_id=b''
            )
            
            # 构建状态（与处理邮件时相同）
            state = {
                "emails": [email_obj],
                "current_email": email_obj,
                "email_category": None,
                "rag_queries": [],
                "retrieved_documents": "",
                "generated_email": "",
                "sendable": False,
                "trials": 0,
                "writer_messages": []
            }
            
            # 先分类邮件（用于选择不同的检索策略）
            # 使用nodes.categorize_email方法，需要传入state
            category_state = nodes.categorize_email(state)
            state.update(category_state)
            category = state.get("email_category", "product_enquiry")
            
            # 检查是否已取消
            if cancel_flag.is_set():
                print(f"🚫 [RAG测试] 在分类后检测到取消信号，停止检索")
                cancel_result = {
                    "question": request.question,
                    "answer": "检索已取消",
                    "success": False
                }
                # 通过 WebSocket 通知前端检索已取消
                try:
                    message = {
                        "type": "rag_test_complete",
                        "question": request.question,
                        "answer": "检索已取消",
                        "success": False,
                        "cancelled": True
                    }
                    import asyncio
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    loop.run_until_complete(ws_manager.broadcast(message))
                    loop.close()
                except Exception as ws_error:
                    print(f"⚠️ [RAG测试] WebSocket 通知失败: {ws_error}")
                return cancel_result
            
            # email_category已经在state中，retrieve_from_rag会从那里获取
            # 保持current_email为Email对象，不要转换为字典
            
            print(f"📋 [RAG测试] 邮件分类: {category}")
            
            # 构建RAG查询（与处理邮件时相同）
            rag_query_result = nodes.construct_rag_queries(state)
            state.update(rag_query_result)
            print(f"🔍 [RAG测试] 生成的查询: {state.get('rag_queries', [])}")
            
            # 检查是否已取消
            if cancel_flag.is_set():
                print(f"🚫 [RAG测试] 在构建查询后检测到取消信号，停止检索")
                cancel_result = {
                    "question": request.question,
                    "answer": "检索已取消",
                    "success": False
                }
                # 通过 WebSocket 通知前端检索已取消
                try:
                    message = {
                        "type": "rag_test_complete",
                        "question": request.question,
                        "answer": "检索已取消",
                        "success": False,
                        "cancelled": True
                    }
                    import asyncio
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    loop.run_until_complete(ws_manager.broadcast(message))
                    loop.close()
                except Exception as ws_error:
                    print(f"⚠️ [RAG测试] WebSocket 通知失败: {ws_error}")
                return cancel_result
            
            # 从RAG检索信息（与处理邮件时相同）
            rag_result = nodes.retrieve_from_rag(state)
            state.update(rag_result)
            
            # 获取检索结果
            retrieved_docs = state.get('retrieved_documents', '')
            print(f"✅ [RAG测试] 检索成功，结果长度: {len(retrieved_docs) if retrieved_docs else 0}")
            
            # retrieved_documents 已经是最终答案（由 rag_generator.invoke 生成），直接使用
            if retrieved_docs and retrieved_docs.strip():
                result = retrieved_docs.strip()
                print(f"📝 [RAG测试] 最终答案预览: {result[:100]}...")
            else:
                result = "未找到相关信息"
                print(f"⚠️ [RAG测试] 未找到相关信息")
            
            result_data = {
                "question": request.question,
                "answer": result,
                "success": True
            }
            
            # 通过 WebSocket 通知前端检索完成
            try:
                message = {
                    "type": "rag_test_complete",
                    "question": request.question,
                    "answer": result,
                    "success": True
                }
                # 在新的事件循环中发送（因为当前在线程池中）
                import asyncio
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                loop.run_until_complete(ws_manager.broadcast(message))
                loop.close()
                print(f"📢 [RAG测试] 已通过 WebSocket 通知前端检索完成")
            except Exception as ws_error:
                print(f"⚠️ [RAG测试] WebSocket 通知失败: {ws_error}")
            
            return result_data
        except Exception as e:
            print(f"❌ [RAG测试] 检索失败: {e}")
            import traceback
            traceback.print_exc()
            
            error_result = {
                "question": request.question,
                "answer": f"检索失败: {str(e)}",
                "success": False
            }
            
            # 通过 WebSocket 通知前端检索失败
            try:
                message = {
                    "type": "rag_test_complete",
                    "question": request.question,
                    "answer": f"检索失败: {str(e)}",
                    "success": False
                }
                import asyncio
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                loop.run_until_complete(ws_manager.broadcast(message))
                loop.close()
            except Exception as ws_error:
                print(f"⚠️ [RAG测试] WebSocket 通知失败: {ws_error}")
            
            return error_result
    
    # 在线程池中执行同步函数，避免阻塞事件循环（与处理邮件时相同）
    loop = asyncio.get_event_loop()
    try:
        # 使用 asyncio.wait_for 来支持取消
        # 如果客户端断开连接，await 会抛出 CancelledError
        result = await asyncio.wait_for(
            loop.run_in_executor(None, run_rag_test_sync),
            timeout=None  # 不设置超时，只用于支持取消
        )
        # 检索完成，清除取消标志
        clear_rag_cancel_flag(current_username)
        
        # 如果同步函数中已经发送了 WebSocket 通知，这里就不需要再发送了
        # 但为了确保前端能收到，这里也发送一次（作为备用）
        # 注意：同步函数中已经发送了通知，这里主要是确保通知到达
        return result
    except asyncio.CancelledError:
        # 客户端断开连接或请求被取消
        cancel_flag.set()  # 设置取消标志，让同步函数也能检测到
        print(f"🚫 [RAG测试] 客户端断开连接，停止检索")
        # 通过 WebSocket 通知前端检索已取消
        try:
            message = {
                "type": "rag_test_complete",
                "question": request.question,
                "answer": "检索已取消",
                "success": False,
                "cancelled": True
            }
            await ws_manager.broadcast(message)
        except Exception as ws_error:
            print(f"⚠️ [RAG测试] WebSocket 通知失败: {ws_error}")
        return {
            "question": request.question,
            "answer": "检索已取消",
            "success": False
        }
    except Exception as e:
        print(f"❌ [RAG测试] 异常: {e}")
        import traceback
        traceback.print_exc()
        return {
            "question": request.question,
            "answer": f"检索异常: {str(e)}",
            "success": False
        }
    except TimeoutError as e:
        print(f"⏱️ [RAG测试] 超时: {e}")
        return {
            "question": request.question,
            "answer": f"检索超时: {str(e)}\n\n可能的原因:\n1. 网络连接较慢\n2. API服务响应延迟\n3. 嵌入模型处理时间较长\n\n建议:\n1. 检查网络连接\n2. 稍后重试\n3. 如果问题持续，考虑使用本地嵌入模型",
            "success": False
        }
    except Exception as e:
        print(f"❌ [RAG测试] 检索失败: {e}")
        import traceback
        traceback.print_exc()
        
        # 提供更友好的错误信息
        error_msg = str(e)
        if "timeout" in error_msg.lower() or "timed out" in error_msg.lower():
            friendly_msg = f"请求超时: {error_msg}\n\n可能的原因:\n1. 网络连接较慢\n2. API服务响应延迟\n3. 嵌入模型处理时间较长\n\n建议:\n1. 检查网络连接\n2. 稍后重试\n3. 如果问题持续，考虑使用本地嵌入模型"
        elif "api key" in error_msg.lower() or "unauthorized" in error_msg.lower():
            friendly_msg = f"API密钥错误: {error_msg}\n\n请检查系统设置中的API密钥配置"
        elif "database" in error_msg.lower() or "vectorstore" in error_msg.lower():
            friendly_msg = f"向量数据库错误: {error_msg}\n\n请检查:\n1. 向量数据库是否存在\n2. 是否已运行 create_index.py 创建索引"
        else:
            friendly_msg = f"检索失败: {error_msg}\n\n请检查:\n1. 向量数据库是否存在\n2. 嵌入模型配置是否正确\n3. 知识库文档是否已索引\n4. 网络连接是否正常"
        
        return {
            "question": request.question,
            "answer": friendly_msg,
            "success": False
        }

@app.post("/api/knowledge/documents/{doc_id}/reindex")
async def reindex_document(doc_id: str, current_username: str = Depends(get_username_from_request)):
    """重建文档索引（单个文档或全部文档）"""
    try:
        # 防止路径遍历攻击
        if doc_id != "all" and ('..' in doc_id or '/' in doc_id or '\\' in doc_id):
            raise HTTPException(status_code=400, detail="无效的文件名")
        
        # 如果是重建全部索引
        if doc_id == "all":
            print(f"🔄 [知识库] 用户 {current_username} 请求重建全部索引...")
            specific_file = None
        else:
            filepath = os.path.join(get_data_dir(), doc_id)
            if not os.path.exists(filepath):
                raise HTTPException(status_code=404, detail="文档不存在")
            specific_file = doc_id
            print(f"🔄 [知识库] 用户 {current_username} 请求重建文档索引: {doc_id}")
        
        # 在后台线程中执行索引构建（避免阻塞API响应）
        def build_index_background():
            try:
                # 获取用户设置中的嵌入模型配置
                user_settings = get_user_settings(current_username)
                embedding_model = user_settings.get("embeddingModel") or os.getenv("EMBEDDING_MODEL", "Qwen/Qwen3-Embedding-4B")
                api_key = user_settings.get("apiKey") or os.getenv("SILICONFLOW_API_KEY")
                
                print(f"🔧 [索引重建] 使用嵌入模型: {embedding_model}")
                
                index_result = build_index(
                    embedding_model=embedding_model,
                    api_key=api_key,
                    specific_file=specific_file
                )
                
                if index_result.get("success"):
                    print(f"✅ [索引重建] 索引重建成功")
                    print(f"   - 数据库路径: {index_result.get('db_path')}")
                    print(f"   - 维度: {index_result.get('dimension')}")
                    print(f"   - 文档块数: {index_result.get('chunks')}")
                    print(f"   - 耗时: {index_result.get('elapsed_time', 0):.1f}秒")
                else:
                    print(f"❌ [索引重建] 索引重建失败: {index_result.get('error')}")
            except Exception as e:
                print(f"❌ [索引重建] 后台执行失败: {e}")
                import traceback
                traceback.print_exc()
        
        # 提交到线程池执行
        thread_pool.submit(build_index_background)
        
        if doc_id == "all":
            return {
                "message": "全部索引重建任务已启动，正在后台执行...",
                "success": True,
                "note": "索引重建可能需要1-2分钟，请稍后查看结果"
            }
        else:
            return {
                "message": f"文档 {doc_id} 索引重建任务已启动，正在后台执行...",
                "success": True,
                "note": "索引重建可能需要1-2分钟，请稍后查看结果"
            }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ [知识库] 重建索引失败: {e}")
        raise HTTPException(status_code=500, detail=f"重建索引失败: {str(e)}")

# ==================== 设置API ====================

# 默认设置（用于新用户初始化）
DEFAULT_SETTINGS = {
    "apiKey": "",
    "model": "moonshotai/Kimi-K2-Thinking",  # 保留用于向后兼容
    "replyModel": "moonshotai/Kimi-K2-Thinking",  # 回复大模型
    "embeddingModel": "Qwen/Qwen3-Embedding-4B",  # 嵌入大模型
    "interval": 15,
    "autoProcess": False,  # 监控运行时自动处理新邮件
    "autoSend": False,
    "batchSize": 4,  # 每批并发处理的邮件数量（1-30）
    "singleEmailConcurrency": 4,  # 单封邮件处理的并发数量（2-20）
    "signature": "Agentia 团队",
    "greeting": "尊敬的客户，您好！",
    "closing": "祝好！"
}

def get_user_settings(username: str) -> dict:
    """获取用户设置，如果不存在则返回默认值"""
    global user_data
    user_data = load_user_data()
    
    if username not in user_data:
        return DEFAULT_SETTINGS.copy()
    
    user_info = user_data[username]
    
    # 从用户数据中获取设置，如果不存在则使用默认值
    settings = user_info.get("settings", {})
    # 兼容旧配置：如果只有 model 字段，则同时设置 replyModel
    model = settings.get("model", DEFAULT_SETTINGS["model"])
    # 优先使用 replyModel，如果不存在则使用 model（兼容旧配置）
    reply_model = settings.get("replyModel")
    if not reply_model:  # 如果 replyModel 不存在或为空，使用 model
        reply_model = model
    print(f"[设置获取] 用户 {username} 的设置: model={model}, replyModel={settings.get('replyModel')}, 最终使用 reply_model={reply_model}")
    return {
        "apiKey": settings.get("apiKey", DEFAULT_SETTINGS["apiKey"]),
        "model": model,  # 保留用于向后兼容
        "replyModel": reply_model,
        "embeddingModel": settings.get("embeddingModel", DEFAULT_SETTINGS["embeddingModel"]),
        "interval": settings.get("interval", DEFAULT_SETTINGS["interval"]),
        "autoProcess": settings.get("autoProcess", DEFAULT_SETTINGS["autoProcess"]),  # 监控运行时自动处理新邮件
        "autoSend": settings.get("autoSend", DEFAULT_SETTINGS["autoSend"]),
        "batchSize": settings.get("batchSize", DEFAULT_SETTINGS["batchSize"]),
        "singleEmailConcurrency": settings.get("singleEmailConcurrency", DEFAULT_SETTINGS["singleEmailConcurrency"]),
        "signature": settings.get("signature", DEFAULT_SETTINGS["signature"]),
        "greeting": settings.get("greeting", DEFAULT_SETTINGS["greeting"]),
        "closing": settings.get("closing", DEFAULT_SETTINGS["closing"])
    }

def get_custom_models(username: str) -> List[dict]:
    """获取用户的自定义模型列表"""
    global user_data
    user_data = load_user_data()
    
    if username not in user_data:
        return []
    
    user_info = user_data[username]
    return user_info.get("settings", {}).get("customModels", [])

@app.get("/api/settings")
async def get_settings(current_username: str = Depends(get_username_from_request)):
    """获取用户设置（包括用户邮箱配置）"""
    global user_data
    user_data = load_user_data()
    
    if current_username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user_info = user_data[current_username]
    user_settings = get_user_settings(current_username)
    custom_models = get_custom_models(current_username)
    
    # 返回用户设置（包括邮箱配置）
    return {
        "email": user_info.get("email", ""),
        "authCode": user_info.get("emailAuthCode", ""),  # 注意：实际应用中应该加密存储
        "apiKey": user_settings["apiKey"],  # 保留用于向后兼容，但前端不再显示
        "model": user_settings.get("model", user_settings.get("replyModel", DEFAULT_SETTINGS["model"])),  # 兼容旧配置
        "replyModel": user_settings["replyModel"],
        "embeddingModel": user_settings["embeddingModel"],
        "interval": user_settings["interval"],
        "autoProcess": user_settings.get("autoProcess", False),  # 监控运行时自动处理新邮件
        "autoSend": user_settings["autoSend"],
        "batchSize": user_settings["batchSize"],
        "singleEmailConcurrency": user_settings.get("singleEmailConcurrency", DEFAULT_SETTINGS["singleEmailConcurrency"]),
        "signature": user_settings["signature"],
        "greeting": user_settings["greeting"],
        "closing": user_settings["closing"],
        "customModels": custom_models  # 返回自定义模型列表
    }

@app.post("/api/settings")
async def save_settings(settings: SettingsModel, background_tasks: BackgroundTasks, current_username: str = Depends(get_username_from_request)):
    """保存用户设置（包括用户邮箱配置）"""
    global user_data
    user_data = load_user_data()
    
    if current_username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user_info = user_data[current_username]
    
    # 保存用户邮箱配置（用户级别）
    if settings.email is not None:
        user_info["email"] = settings.email
    if settings.authCode is not None:
        user_info["emailAuthCode"] = settings.authCode  # 注意：实际应用中应该加密存储
    
    # 初始化用户设置字典（如果不存在）
    if "settings" not in user_info:
        user_info["settings"] = {}
    
    # 保存用户特定的设置（用户级别）
    # 注意：apiKey 保留用于向后兼容，但不再在前端显示
    # 系统默认模型使用环境变量中的 SILICONFLOW_API_KEY
    if settings.apiKey is not None:
        user_info["settings"]["apiKey"] = settings.apiKey
    if settings.model is not None:
        user_info["settings"]["model"] = settings.model  # 保留用于向后兼容
    if settings.replyModel is not None:
        user_info["settings"]["replyModel"] = settings.replyModel
    if settings.embeddingModel is not None:
        user_info["settings"]["embeddingModel"] = settings.embeddingModel
    if settings.interval is not None:
        user_info["settings"]["interval"] = settings.interval
        # 更新当前用户的检查间隔
        user_state = get_user_state(current_username)
        user_state.check_interval = settings.interval * 60
    if settings.autoProcess is not None:
        user_info["settings"]["autoProcess"] = settings.autoProcess
        # 同步更新用户状态中的 auto_process
        user_state = get_user_state(current_username)
        old_auto_process = user_state.auto_process
        user_state.auto_process = settings.autoProcess
        if settings.autoProcess:
            print(f"✅ [保存设置] 用户 {current_username} 的自动处理已开启")
            # 如果从关闭变为开启，立即触发一次处理（不等待下次监控循环）
            if not old_auto_process and settings.autoProcess:
                print(f"🚀 [保存设置] 检测到自动处理从关闭变为开启，立即触发一次处理...")
                # 在后台任务中执行，避免阻塞响应
                background_tasks.add_task(lambda: thread_pool.submit(user_state._auto_process_emails_async))
        else:
            print(f"❌ [保存设置] 用户 {current_username} 的自动处理已关闭")
        print(f"[保存设置] autoProcess 已保存到数据文件: {settings.autoProcess}")
        print(f"[保存设置] user_state.auto_process 已更新: {user_state.auto_process}")
    auto_send_enabled = False
    auto_send_disabled = False
    if settings.autoSend is not None:
        old_auto_send = user_info["settings"].get("autoSend", False)
        user_info["settings"]["autoSend"] = settings.autoSend
        auto_send_enabled = settings.autoSend
        
        # 如果 autoSend 从 True 变为 False，记录日志
        if old_auto_send and not settings.autoSend:
            auto_send_disabled = True
            print(f"ℹ️ [保存设置] 用户 {current_username} 关闭了自动发送")
            # 注意：自动发送线程会在下次循环时检查设置并跳过发送，无需手动停止线程
    if settings.singleEmailConcurrency is not None:
        # 限制 singleEmailConcurrency 在合理范围内（2-20）
        # 注意：设置过高的值可能导致内存和API调用压力增大
        concurrency = max(2, min(20, int(settings.singleEmailConcurrency)))
        user_info["settings"]["singleEmailConcurrency"] = concurrency
        if concurrency > 10:
            print(f"⚠️ [设置保存] 用户 {current_username} 设置了较高的单封邮件并发数量 ({concurrency})，请注意系统资源使用情况")
    if settings.batchSize is not None:
        # 限制 batchSize 在合理范围内（1-30，允许更高的并发）
        # 注意：设置过高的值可能导致内存和API调用压力增大
        batch_size = max(1, min(30, int(settings.batchSize)))
        user_info["settings"]["batchSize"] = batch_size
        if batch_size > 15:
            print(f"⚠️ [设置保存] 用户 {current_username} 设置了较高的批量并发数量 ({batch_size})，请注意系统资源使用情况")
    if settings.signature is not None:
        user_info["settings"]["signature"] = settings.signature
    if settings.greeting is not None:
        user_info["settings"]["greeting"] = settings.greeting
    if settings.closing is not None:
        user_info["settings"]["closing"] = settings.closing
    
    # 保存用户数据
    user_data[current_username] = user_info
    save_user_data(user_data)
    
    # 如果开启了自动发送，且监控正在运行，确保自动发送线程已启动
    # 注意：不再自动启动监控，让用户自己决定何时启动监控
    if auto_send_enabled:
        user_state = get_user_state(current_username)
        # 只有在监控已经运行的情况下，才启动自动发送线程
        if user_state.is_running:
            # 如果自动发送线程未启动，启动自动发送线程
            if user_state.auto_send_thread is None or not user_state.auto_send_thread.is_alive():
                print(f"🚀 [保存设置] 检测到自动发送已开启且监控正在运行，启动自动发送线程...")
                user_state.auto_send_thread = threading.Thread(target=user_state._auto_send_loop, daemon=True)
                user_state.auto_send_thread.start()
                print(f"✅ [保存设置] 自动发送线程已启动")
            
            # 在后台任务中执行自动发送，避免阻塞响应
            background_tasks.add_task(send_processed_emails_with_rate_limit, current_username)
        else:
            print(f"ℹ️ [保存设置] 自动发送已开启，但监控未运行。自动发送线程将在启动监控时自动启动。")
    
    return {"message": "设置已保存"}

@app.post("/api/settings/test-email")
async def test_email_connection(request: Optional[TestEmailRequest] = None, current_username: str = Depends(get_username_from_request)):
    """测试邮箱连接（优先使用请求中的邮箱配置，否则使用已保存的配置）"""
    try:
        # 优先使用请求中传入的邮箱配置（用于测试未保存的配置）
        email_address = None
        auth_code = None
        
        if request:
            email_address = request.email
            auth_code = request.authCode
        
        # 如果请求中没有提供，则从已保存的配置中获取
        if not email_address or not auth_code:
            try:
                saved_email, saved_auth = get_user_email_config(current_username)
                if not email_address:
                    email_address = saved_email
                if not auth_code:
                    auth_code = saved_auth
            except ValueError:
                # 如果已保存的配置也不存在，继续使用请求中的值（可能为空）
                pass
        
        if not email_address:
            return {"success": False, "message": "请先输入QQ邮箱地址"}
        if not auth_code:
            return {"success": False, "message": "请先输入授权码"}
        
        # 验证邮箱格式
        email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not re.match(email_regex, email_address):
            return {"success": False, "message": "请输入合法的邮箱地址格式"}
        
        # 验证必须是QQ邮箱
        if not email_address.endswith('@qq.com'):
            return {"success": False, "message": "请输入合法的QQ邮箱地址（必须以 @qq.com 结尾）"}
        
        email_tools = QQEmailToolsClass(email_address=email_address, auth_code=auth_code)
        # 尝试连接
        import imaplib
        try:
            mail = imaplib.IMAP4_SSL(email_tools.imap_server, email_tools.imap_port)
            mail.login(email_tools.email_address, email_tools.auth_code)
            mail.logout()
            return {"success": True, "message": "邮箱连接成功"}
        except imaplib.IMAP4.error as e:
            # 处理IMAP登录错误，转换为友好提示
            error_str = str(e)
            # QQ邮箱登录失败时，通常不会明确区分是邮箱错误还是授权码错误
            # 所以统一提示用户检查两者
            if 'Login fail' in error_str or 'Account is abnormal' in error_str:
                # 登录失败可能是邮箱地址错误、授权码错误、账号异常等多种原因
                return {"success": False, "message": "登录失败，请检查：\n1. 邮箱地址是否正确（确保是有效的QQ邮箱）\n2. 授权码是否正确\n3. 是否已开启IMAP/SMTP服务\n4. 账号是否异常或被限制"}
            elif 'password is incorrect' in error_str:
                # 只有在明确提示密码错误时才单独提示授权码错误
                return {"success": False, "message": "授权码错误，请检查授权码是否正确，或重新生成授权码"}
            elif 'service is not open' in error_str:
                return {"success": False, "message": "IMAP/SMTP服务未开启，请登录QQ邮箱网页版开启该服务"}
            elif 'login frequency limited' in error_str:
                return {"success": False, "message": "登录频率过高，请稍后再试"}
            elif 'system is busy' in error_str:
                return {"success": False, "message": "系统繁忙，请稍后再试"}
            else:
                # 其他登录错误，统一提示检查邮箱和授权码
                return {"success": False, "message": "登录失败，请检查邮箱地址和授权码是否正确"}
    except ValueError as e:
        return {"success": False, "message": str(e)}
    except Exception as e:
        # 处理其他异常
        error_str = str(e)
        if 'Login fail' in error_str or 'Account is abnormal' in error_str:
            return {"success": False, "message": "登录失败，请检查：\n1. 邮箱地址是否正确（确保是有效的QQ邮箱）\n2. 授权码是否正确\n3. 是否已开启IMAP/SMTP服务\n4. 账号是否异常或被限制"}
        return {"success": False, "message": f"连接失败：{error_str}"}

def generate_body_summary_only(email: dict, user_state, username: str, batch_mode: bool = False):
    """
    只生成原始邮件的摘要（用于新邮件获取后立即生成摘要）
    不生成回复内容的摘要，因为回复内容还没有生成
    
    Args:
        email: 邮件对象
        user_state: 用户状态
        username: 用户名
        batch_mode: 批量模式，如果为True，不立即保存文件，由调用者统一保存
    """
    try:
        email_id = email.get('id', '')
        body = email.get('body', '')
        
        if not body:
            print(f"⚠️ [摘要生成] 邮件 {email_id} 没有原始内容，跳过摘要生成")
            return False
        
        print(f"📝 [摘要生成] 开始为新邮件 {email_id} 生成原始邮件摘要...")
        
        # 获取用户设置
        user_settings = get_user_settings(username)
        reply_model = user_settings.get("replyModel", user_settings.get("model", DEFAULT_SETTINGS["replyModel"]))
        
        # 获取模型配置（API密钥和base URL）
        model_config = get_model_config(username, reply_model, "reply")
        api_key = model_config["apiKey"]
        api_base_url = model_config["apiBaseUrl"]
        
        if not api_key:
            print(f"⚠️ [摘要生成] 用户 {username} 未配置API密钥，跳过摘要生成")
            return False
        
        # 调用大模型生成摘要
        from langchain_openai import ChatOpenAI
        from langchain_core.prompts import ChatPromptTemplate
        
        llm = ChatOpenAI(
            model=reply_model,
            temperature=0.3,
            openai_api_key=api_key,
            openai_api_base=api_base_url,
            timeout=90,
            max_retries=2
        )
        
        summary_prompt = ChatPromptTemplate.from_messages([
            ("system", "你是一个专业的文本摘要助手。请将以下文本内容总结成一段简洁的摘要（50-100字），保留关键信息。"),
            ("user", "{text}")
        ])
        
        chain = summary_prompt | llm
        
        # 生成原始邮件摘要
        try:
            print(f"📝 [摘要生成] 正在生成原始邮件摘要，文本长度: {len(body)}")
            body_summary = chain.invoke({"text": body}).content
            print(f"✅ [摘要生成] 原始邮件摘要生成成功，长度: {len(body_summary)}")
            
            if body_summary:
                # 使用用户锁保护数据更新（避免与邮件处理冲突）
                user_lock = get_user_lock(username)
                with user_lock:
                    # 保存摘要到邮件记录
                    email['body_summary'] = body_summary
                    
                    # 同时更新 history 中的记录（如果存在）
                    for record in user_state.history:
                        if record.get('id') == email_id:
                            record['body_summary'] = body_summary
                            break
                    
                    # 如果不是批量模式，立即保存到文件
                    if not batch_mode:
                        save_user_email_data(username, user_state)
                        print(f"✅ [摘要生成] 原始邮件摘要已保存")
                
                # 通过 WebSocket 推送摘要已保存的消息
                try:
                    asyncio.run(ws_manager.send_message_to_user(
                        username,
                        {
                            "type": "summary_saved",
                            "email_id": email_id,
                            "body_summary": body_summary,
                            "reply_summary": None  # 新邮件还没有回复内容
                        }
                    ))
                    print(f"📤 [摘要生成] 已通过 WebSocket 推送原始邮件摘要: {email_id}")
                except Exception as ws_error:
                    print(f"⚠️ [摘要生成] WebSocket 推送失败: {ws_error}")
                
                return True
            else:
                print(f"⏱️ [摘要生成] 原始邮件摘要生成失败（返回空）: {email_id}")
                return False
        except Exception as e:
            print(f"❌ [摘要生成] 原始邮件摘要生成失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    except Exception as e:
        print(f"❌ [摘要生成] 生成原始邮件摘要时出错: {e}")
        import traceback
        traceback.print_exc()
        return False

def generate_email_summaries_async(username: str, email_id: str, body: str, reply: str):
    """
    异步生成邮件摘要（不阻塞主流程）
    在后台线程中生成摘要并保存到邮件记录中
    优化：并行生成两个摘要，添加超时和重试机制
    """
    def generate_and_save():
        try:
            print(f"📝 [摘要生成] 开始为邮件 {email_id} 生成摘要...")
            
            # 获取用户设置
            user_settings = get_user_settings(username)
            reply_model = user_settings.get("replyModel", user_settings.get("model", DEFAULT_SETTINGS["replyModel"]))
            
            # 获取模型配置（API密钥和base URL）
            model_config = get_model_config(username, reply_model, "reply")
            api_key = model_config["apiKey"]
            api_base_url = model_config["apiBaseUrl"]
            
            if not api_key:
                print(f"⚠️ [摘要生成] 用户 {username} 未配置API密钥，跳过摘要生成")
                return
            
            # 调用大模型生成摘要
            from langchain_openai import ChatOpenAI
            from langchain_core.prompts import ChatPromptTemplate
            import concurrent.futures
            
            llm = ChatOpenAI(
                model=reply_model,
                temperature=0.3,
                openai_api_key=api_key,
                openai_api_base=api_base_url,
                timeout=90,  # 增加超时时间到90秒（API调用）
                max_retries=2  # 最多重试2次
            )
            
            summary_prompt = ChatPromptTemplate.from_messages([
                ("system", "你是一个专业的文本摘要助手。请将以下文本内容总结成一段简洁的摘要（50-100字），保留关键信息。"),
                ("user", "{text}")
            ])
            
            chain = summary_prompt | llm
            
            # 并行生成两个摘要的函数
            def generate_body_summary():
                if not body:
                    return None
                try:
                    print(f"📝 [摘要生成] 正在生成原始邮件摘要，文本长度: {len(body)}")
                    result = chain.invoke({"text": body}).content
                    print(f"✅ [摘要生成] 原始邮件摘要生成成功，长度: {len(result)}")
                    return result
                except Exception as e:
                    print(f"⚠️ [摘要生成] 原始邮件摘要生成失败: {e}")
                    return None
            
            def generate_reply_summary():
                if not reply:
                    return None
                try:
                    print(f"📝 [摘要生成] 正在生成回复内容摘要，文本长度: {len(reply)}")
                    result = chain.invoke({"text": reply}).content
                    print(f"✅ [摘要生成] 回复内容摘要生成成功，长度: {len(result)}")
                    return result
                except Exception as e:
                    print(f"⚠️ [摘要生成] 回复内容摘要生成失败: {e}")
                    return None
            
            # 使用线程池并行执行两个摘要生成任务
            body_summary = None
            reply_summary = None
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
                # 提交两个任务
                body_future = executor.submit(generate_body_summary) if body else None
                reply_future = executor.submit(generate_reply_summary) if reply else None
                
                # 等待任务完成（不设置超时，让任务自然完成）
                # 这样可以避免超时后任务继续执行但结果丢失的问题
                try:
                    if body_future:
                        try:
                            body_summary = body_future.result()  # 不设置超时，等待任务完成
                        except Exception as e:
                            print(f"❌ [摘要生成] 原始邮件摘要生成异常: {e}")
                    
                    if reply_future:
                        try:
                            reply_summary = reply_future.result()  # 不设置超时，等待任务完成
                        except Exception as e:
                            print(f"❌ [摘要生成] 回复内容摘要生成异常: {e}")
                    
                except Exception as e:
                    print(f"⚠️ [摘要生成] 线程池执行异常: {e}")
            
            # 如果两个摘要都没有生成成功，直接返回
            if not body_summary and not reply_summary:
                print(f"⚠️ [摘要生成] 两个摘要都未生成成功，跳过保存: {email_id}")
                return
            
            # 保存摘要到邮件记录
            user_state = get_user_state(username, check_auto_start=False)
            
            # 使用用户锁保护数据更新和保存（避免与邮件处理冲突）
            user_lock = get_user_lock(username)
            with user_lock:
                # 更新邮件缓存中的摘要
                email_found_in_cache = False
                for email in user_state.emails_cache:
                    if email.get('id') == email_id:
                        if body_summary:
                            email['body_summary'] = body_summary
                        if reply_summary:
                            email['reply_summary'] = reply_summary
                        email_found_in_cache = True
                        print(f"✅ [摘要生成] 已更新邮件缓存中的摘要: {email_id}")
                        break
                
                if not email_found_in_cache:
                    print(f"⚠️ [摘要生成] 未在邮件缓存中找到邮件: {email_id}")
                
                # 更新历史记录中的摘要
                history_found = False
                for history_record in user_state.history:
                    if history_record.get('id') == email_id:
                        if body_summary:
                            history_record['body_summary'] = body_summary
                        if reply_summary:
                            history_record['reply_summary'] = reply_summary
                        history_found = True
                        print(f"✅ [摘要生成] 已更新历史记录中的摘要: {email_id}")
                        print(f"  - body_summary 长度: {len(body_summary) if body_summary else 0}")
                        print(f"  - reply_summary 长度: {len(reply_summary) if reply_summary else 0}")
                        break
                
                if not history_found:
                    print(f"⚠️ [摘要生成] 未在历史记录中找到邮件: {email_id}")
                    print(f"  - 历史记录总数: {len(user_state.history)}")
                    print(f"  - 历史记录ID列表: {[h.get('id') for h in user_state.history[:5]]}")
                
                # 保存数据（在锁保护下）
                save_user_email_data(username, user_state)
                print(f"✅ [摘要生成] 摘要已保存到邮件记录: {email_id}")
            
            # 尝试通过 WebSocket 推送给当前用户（若已连接）
            try:
                payload = {
                    "type": "summary_saved",
                    "email_id": email_id,
                    "body_summary": body_summary,
                    "reply_summary": reply_summary
                }
                print(f"📤 [摘要生成] 准备推送摘要更新:")
                print(f"  - 用户: {username}")
                print(f"  - 邮件ID: {email_id}")
                print(f"  - body_summary: {body_summary[:50] if body_summary else None}...")
                print(f"  - reply_summary: {reply_summary[:50] if reply_summary else None}...")
                print(f"  - websocket_event_loop 存在: {websocket_event_loop is not None}")
                
                if websocket_event_loop:
                    # 在事件循环中异步发送（线程安全）
                    future = asyncio.run_coroutine_threadsafe(
                        ws_manager.send_message_to_user(username, payload), 
                        websocket_event_loop
                    )
                    # 等待发送完成（最多等待5秒）
                    try:
                        future.result(timeout=5)
                        print(f"✅ [摘要生成] WebSocket 消息已成功发送给用户: {username}")
                    except Exception as send_err:
                        print(f"❌ [摘要生成] WebSocket 消息发送失败: {send_err}")
                else:
                    print("⚠️ [摘要生成] 未初始化 websocket_event_loop，无法推送消息")
            except Exception as e:
                print(f"⚠️ [摘要生成] 推送 WS 消息失败: {e}")
                import traceback
                traceback.print_exc()
            
        except Exception as e:
            print(f"❌ [摘要生成] 生成摘要时出错: {e}")
            import traceback
            traceback.print_exc()
    
    # 提交到摘要生成线程池（不阻塞主流程，且限制并发数量）
    summary_generation_pool.submit(generate_and_save)
    print(f"🚀 [摘要生成] 已提交摘要生成任务到线程池: {email_id}")

@app.post("/api/emails/summarize")
async def generate_summary(request: dict, current_username: str = Depends(get_username_from_request)):
    """使用回复大模型生成文本摘要（与系统设置中的回复大模型配置一致）"""
    text = request.get("text")
    if not text:
        raise HTTPException(status_code=400, detail="文本内容不能为空")
    
    try:
        # 获取用户设置（与系统设置界面同步）
        user_settings = get_user_settings(current_username)
        reply_model = user_settings.get("replyModel", user_settings.get("model", DEFAULT_SETTINGS["replyModel"]))
        
        # 根据选择的模型自动获取API密钥和base URL
        model_config = get_model_config(current_username, reply_model, "reply")
        api_key = model_config["apiKey"]
        api_base_url = model_config["apiBaseUrl"]
        
        if not api_key:
            raise HTTPException(status_code=400, detail="请先配置API密钥（系统默认模型需要配置环境变量SILICONFLOW_API_KEY，自定义模型需要在设置中配置API密钥）")
        
        # 调用大模型生成摘要
        from langchain_openai import ChatOpenAI
        from langchain_core.prompts import ChatPromptTemplate
        
        llm = ChatOpenAI(
            model=reply_model,
            temperature=0.3,  # 稍微高一点，让摘要更自然
            openai_api_key=api_key,
            openai_api_base=api_base_url
        )
        
        # 构建摘要提示词
        summary_prompt = ChatPromptTemplate.from_messages([
            ("system", "你是一个专业的文本摘要助手。请将以下文本内容总结成一段简洁的摘要（50-100字），保留关键信息。"),
            ("user", "{text}")
        ])
        
        chain = summary_prompt | llm
        summary = chain.invoke({"text": text}).content
        
        return {
            "success": True,
            "summary": summary
        }
    except Exception as e:
        print(f"生成摘要失败: {e}")
        import traceback
        traceback.print_exc()
        error_msg = str(e)
        if "Invalid API key" in error_msg or "Unauthorized" in error_msg:
            raise HTTPException(status_code=400, detail="API密钥无效，请检查密钥是否正确")
        elif "Model not found" in error_msg or "model" in error_msg.lower():
            raise HTTPException(status_code=400, detail="模型不存在或不可用，请检查模型配置")
        else:
            raise HTTPException(status_code=500, detail=f"生成摘要失败: {error_msg}")

def get_api_key_for_model(username: str, model_name: str, model_type: str = "reply") -> Optional[str]:
    """
    根据模型名称获取对应的API密钥
    如果是自定义模型，返回自定义模型的API；否则返回None（使用系统默认API）
    
    @param username: 用户名
    @param model_name: 模型名称
    @param model_type: 模型类型（'reply' 或 'embedding'）
    @return: API密钥，如果为None则使用系统默认API（环境变量中的SILICONFLOW_API_KEY）
    """
    custom_models = get_custom_models(username)
    for custom_model in custom_models:
        if custom_model.get("model") == model_name and custom_model.get("type") == model_type:
            return custom_model.get("apiKey")
    return None  # 返回None表示使用系统默认API

def get_model_config(username: str, model_name: str, model_type: str = "reply") -> dict:
    """
    根据模型名称获取模型配置（API密钥和base URL）
    
    @param username: 用户名
    @param model_name: 模型名称
    @param model_type: 模型类型（'reply' 或 'embedding'）
    @return: 包含apiKey和apiBaseUrl的字典
    """
    custom_models = get_custom_models(username)
    print(f"🔍 [模型配置] 查找模型: {model_name}, 类型: {model_type}")
    print(f"🔍 [模型配置] 用户 {username} 的自定义模型数量: {len(custom_models)}")
    
    for custom_model in custom_models:
        if custom_model.get("model") == model_name and custom_model.get("type") == model_type:
            provider = custom_model.get("provider", "硅基流动")
            custom_url = custom_model.get("apiBaseUrl")
            api_base_url = get_api_base_url(provider, custom_url)
            print(f"✅ [模型配置] 找到自定义模型: {model_name}")
            print(f"   - 服务商: {provider}")
            print(f"   - 自定义URL: {custom_url}")
            print(f"   - 最终API地址: {api_base_url}")
            return {
                "apiKey": custom_model.get("apiKey"),
                "apiBaseUrl": api_base_url
            }
    
    # 如果不是自定义模型，使用系统默认配置
    import os
    print(f"⚠️ [模型配置] 未找到自定义模型 {model_name}，使用系统默认配置")
    return {
        "apiKey": os.getenv("SILICONFLOW_API_KEY"),
        "apiBaseUrl": "https://api.siliconflow.cn/v1"
    }

def get_models_config(username: str, reply_model: str, embedding_model: str) -> dict:
    """
    同时获取回复模型和嵌入模型的配置
    
    @param username: 用户名
    @param reply_model: 回复模型名称
    @param embedding_model: 嵌入模型名称
    @return: 包含apiKey、replyApiBaseUrl、embeddingApiBaseUrl的字典
    """
    reply_config = get_model_config(username, reply_model, "reply")
    embedding_config = get_model_config(username, embedding_model, "embedding")
    
    # 优先使用回复模型的API密钥
    api_key = reply_config["apiKey"] or embedding_config["apiKey"]
    
    return {
        "apiKey": api_key,
        "replyApiBaseUrl": reply_config["apiBaseUrl"],
        "embeddingApiBaseUrl": embedding_config["apiBaseUrl"]
    }

def get_api_key_for_models(username: str, reply_model: str, embedding_model: str) -> Optional[str]:
    """
    根据回复模型和嵌入模型获取对应的API密钥
    优先使用回复模型的API，如果回复模型不是自定义模型，则使用嵌入模型的API
    如果都不是自定义模型，返回None（使用系统默认API）
    
    @param username: 用户名
    @param reply_model: 回复模型名称
    @param embedding_model: 嵌入模型名称
    @return: API密钥，如果为None则使用系统默认API（环境变量中的SILICONFLOW_API_KEY）
    """
    # 优先检查回复模型是否是自定义模型
    reply_api_key = get_api_key_for_model(username, reply_model, "reply")
    if reply_api_key:
        return reply_api_key
    
    # 检查嵌入模型是否是自定义模型
    embedding_api_key = get_api_key_for_model(username, embedding_model, "embedding")
    if embedding_api_key:
        return embedding_api_key
    
    # 如果都不是自定义模型，返回None表示使用系统默认API
    return None

@app.post("/api/settings/test-ai")
async def test_ai_connection(request: Optional[TestAIRequest] = None, current_username: str = Depends(get_username_from_request)):
    """测试AI连接（优先使用请求中的AI配置，否则使用已保存的配置）"""
    try:
        # 优先使用请求中传入的AI配置（用于测试未保存的配置）
        api_key = None
        reply_model = None
        embedding_model = None
        
        if request:
            api_key = request.apiKey  # 如果前端明确提供了API（自定义模型），则使用它
            reply_model = request.replyModel
            embedding_model = request.embeddingModel
        
        # 如果请求中没有提供，则从已保存的配置中获取
        if not reply_model or not embedding_model:
            user_settings = get_user_settings(current_username)
            if not reply_model:
                reply_model = user_settings.get("replyModel", user_settings.get("model", DEFAULT_SETTINGS["replyModel"]))
            if not embedding_model:
                embedding_model = user_settings.get("embeddingModel", DEFAULT_SETTINGS["embeddingModel"])
        
        # 根据选择的模型确定使用的API密钥和base URL
        # 需要分别获取回复模型和嵌入模型的配置，因为它们可能来自不同服务商
        reply_api_key = None
        embedding_api_key = None
        reply_api_base_url = request.replyApiBaseUrl if request and request.replyApiBaseUrl else "https://api.siliconflow.cn/v1"
        embedding_api_base_url = request.embeddingApiBaseUrl if request and request.embeddingApiBaseUrl else "https://api.siliconflow.cn/v1"
        
        # 获取回复模型的配置
        reply_config = get_model_config(current_username, reply_model, "reply")
        reply_api_key = reply_config["apiKey"]
        if not request or not request.replyApiBaseUrl:
            reply_api_base_url = reply_config["apiBaseUrl"]
        
        # 获取嵌入模型的配置
        embedding_config = get_model_config(current_username, embedding_model, "embedding")
        embedding_api_key = embedding_config["apiKey"]
        if not request or not request.embeddingApiBaseUrl:
            embedding_api_base_url = embedding_config["apiBaseUrl"]
        
        # 如果前端明确提供了API密钥（用于测试未保存的配置），使用前端提供的
        if request and request.apiKey:
            # 前端提供的API密钥用于回复模型
            reply_api_key = request.apiKey
            # 如果嵌入模型没有自己的API密钥，也使用这个
            if not embedding_api_key:
                embedding_api_key = request.apiKey
        
        if not reply_api_key:
            return {"success": False, "message": "请先配置回复模型的API密钥"}
        if not embedding_api_key:
            return {"success": False, "message": "请先配置嵌入模型的API密钥"}
        if not reply_model:
            return {"success": False, "message": "请先选择回复大模型"}
        if not embedding_model:
            return {"success": False, "message": "请先选择嵌入大模型"}
        
        print(f"🧪 [测试AI] 开始测试AI连接")
        print(f"   - 回复模型: {reply_model}")
        print(f"   - 嵌入模型: {embedding_model}")
        print(f"   - 回复模型API地址: {reply_api_base_url}")
        print(f"   - 嵌入模型API地址: {embedding_api_base_url}")
        print(f"   - 回复模型API密钥: {reply_api_key[:20]}...")
        print(f"   - 嵌入模型API密钥: {embedding_api_key[:20]}...")
        
        # 测试回复大模型
        from langchain_openai import ChatOpenAI
        print(f"🧪 [测试AI] 正在测试回复模型...")
        llm = ChatOpenAI(
            model=reply_model,
            temperature=0.1,
            openai_api_key=reply_api_key,
            openai_api_base=reply_api_base_url
        )
        test_result = llm.invoke("测试")
        print(f"✅ [测试AI] 回复模型测试成功")
        
        # 测试嵌入大模型
        from langchain_openai import OpenAIEmbeddings
        print(f"🧪 [测试AI] 正在测试嵌入模型...")
        embeddings = OpenAIEmbeddings(
            model=embedding_model,
            openai_api_key=embedding_api_key,
            openai_api_base=embedding_api_base_url,
            request_timeout=10
        )
        test_embedding = embeddings.embed_query("测试")
        print(f"✅ [测试AI] 嵌入模型测试成功")
        
        return {"success": True, "message": f"API连接成功！回复模型：{reply_model}，嵌入模型：{embedding_model}"}
    except Exception as e:
        error_msg = str(e)
        if "Invalid API key" in error_msg or "Unauthorized" in error_msg:
            return {"success": False, "message": "API密钥无效，请检查密钥是否正确"}
        elif "Model not found" in error_msg or "model" in error_msg.lower():
            return {"success": False, "message": f"模型配置错误：{error_msg}"}
        else:
            return {"success": False, "message": f"连接失败: {error_msg}"}

# ==================== 自定义模型管理API ====================

@app.post("/api/settings/models")
async def add_custom_model(model: CustomModelModel, current_username: str = Depends(get_username_from_request)):
    """添加自定义模型"""
    global user_data
    user_data = load_user_data()
    
    if current_username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user_info = user_data[current_username]
    
    # 初始化用户设置字典（如果不存在）
    if "settings" not in user_info:
        user_info["settings"] = {}
    if "customModels" not in user_info["settings"]:
        user_info["settings"]["customModels"] = []
    
    # 检查模型是否已存在
    existing_models = user_info["settings"]["customModels"]
    for existing_model in existing_models:
        if existing_model.get("model") == model.model and existing_model.get("type") == model.type:
            raise HTTPException(status_code=400, detail="该模型已存在")
    
    # 添加新模型
    new_model = {
        "id": str(uuid.uuid4()),
        "provider": model.provider,
        "model": model.model,
        "apiKey": model.apiKey,
        "type": model.type,
        "apiBaseUrl": model.apiBaseUrl  # 保存自定义API base URL
    }
    user_info["settings"]["customModels"].append(new_model)
    
    # 保存用户数据
    user_data[current_username] = user_info
    save_user_data(user_data)
    
    return {"success": True, "message": "模型添加成功", "model": new_model}

@app.delete("/api/settings/models/{model_id}")
async def delete_custom_model(model_id: str, current_username: str = Depends(get_username_from_request)):
    """删除自定义模型"""
    global user_data
    user_data = load_user_data()
    
    if current_username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user_info = user_data[current_username]
    
    # 初始化用户设置字典（如果不存在）
    if "settings" not in user_info:
        user_info["settings"] = {}
    if "customModels" not in user_info["settings"]:
        user_info["settings"]["customModels"] = []
    
    # 查找并删除模型
    custom_models = user_info["settings"]["customModels"]
    original_count = len(custom_models)
    user_info["settings"]["customModels"] = [
        m for m in custom_models if m.get("id") != model_id
    ]
    
    if len(user_info["settings"]["customModels"]) == original_count:
        raise HTTPException(status_code=404, detail="模型不存在")
    
    # 保存用户数据
    user_data[current_username] = user_info
    save_user_data(user_data)
    
    return {"success": True, "message": "模型删除成功"}

@app.get("/api/settings/models")
async def get_custom_models_api(current_username: str = Depends(get_username_from_request)):
    """获取用户的自定义模型列表"""
    custom_models = get_custom_models(current_username)
    return {"success": True, "models": custom_models}

# ==================== MailBot 智能助手 API ====================

class AiChatRequest(BaseModel):
    """AI聊天请求模型"""
    conversationId: Optional[str] = None
    message: str
    pageContext: Optional[dict] = None

# 存储会话历史（简单实现，生产环境建议使用Redis等）
ai_conversations: Dict[str, List[dict]] = {}

@app.post("/api/ai/chat")
async def ai_chat(request: AiChatRequest, current_username: str = Depends(get_username_from_request)):
    """
    MailBot 智能助手聊天接口
    使用 Qwen/Qwen2.5-7B-Instruct 模型
    """
    from langchain_openai import ChatOpenAI
    from langchain_core.messages import HumanMessage, AIMessage, SystemMessage
    
    try:
        # 获取或创建会话ID
        conversation_id = request.conversationId or f"conv_{uuid.uuid4().hex[:12]}"
        
        # 获取API密钥（优先使用用户配置，否则使用环境变量）
        user_data_local = load_user_data()
        api_key = None
        if current_username in user_data_local:
            user_settings = user_data_local[current_username].get("settings", {})
            api_key = user_settings.get("apiKey")
        
        if not api_key:
            api_key = os.getenv("SILICONFLOW_API_KEY")
        
        if not api_key:
            return {
                "conversationId": conversation_id,
                "answer": "❌ 未配置API密钥。请在系统设置中配置API密钥，或在 .env 文件中设置 SILICONFLOW_API_KEY。",
                "sources": []
            }
        
        # 初始化 LLM
        llm = ChatOpenAI(
            model="Qwen/Qwen2.5-7B-Instruct",
            temperature=0.7,
            openai_api_key=api_key,
            openai_api_base="https://api.siliconflow.cn/v1",
            max_tokens=2000
        )
        
        # 获取会话历史
        if conversation_id not in ai_conversations:
            ai_conversations[conversation_id] = []
        
        history = ai_conversations[conversation_id]
        
        # 构建系统提示词
        system_prompt = """你是 MailBot 智能助手，一个专业的邮件自动化系统AI助教。你的职责是**仅限于**帮助用户解答关于本邮件自动化系统的使用问题。

【你的业务范围 - 只回答以下相关问题】：
1. 邮箱账号接入配置（QQ邮箱授权码获取、IMAP/SMTP设置、邮箱连接测试等）
2. 邮件处理功能（自动处理、手动处理、邮件分类、回复生成等）
3. 知识库使用（上传文档、RAG检索、重建索引等）
4. 系统设置（API配置、模型选择、监控间隔、自动发送等）
5. 故障排查（处理失败、连接错误、API错误等）
6. 处理记录查询（历史记录、导出功能等）
7. 系统功能介绍（各页面功能说明、操作指南等）

【严格限制 - 以下问题一律礼貌拒绝】：
- 与本邮件自动化系统无关的任何问题
- 闲聊、娱乐、情感咨询等
- 编程开发问题（除非是关于本系统的配置）
- 其他软件或系统的使用问题
- 任何通用知识问答

【拒绝回答时的标准回复】：
当用户提问超出业务范围时，请礼貌回复：
"抱歉，您的问题超出了我的业务范围。我是 MailBot 智能助手，专门负责解答本邮件自动化系统的使用问题。如果您有关于邮箱配置、邮件处理、知识库使用、系统设置等方面的问题，我很乐意为您解答！😊"

【回答规范】：
- 使用简洁、专业、友好的语气
- 使用清晰的格式（列表、步骤等）
- 提供具体的操作指引
- 必要时给出示例
- 适当使用 emoji 增加亲和力"""

        # 构建消息列表
        messages = [SystemMessage(content=system_prompt)]
        
        # 添加历史消息（最多保留最近10轮对话）
        for msg in history[-20:]:
            if msg["role"] == "user":
                messages.append(HumanMessage(content=msg["content"]))
            else:
                messages.append(AIMessage(content=msg["content"]))
        
        # 添加当前用户消息
        messages.append(HumanMessage(content=request.message))
        
        # 调用 LLM
        response = await asyncio.to_thread(llm.invoke, messages)
        answer = response.content
        
        # 保存到会话历史
        history.append({"role": "user", "content": request.message})
        history.append({"role": "assistant", "content": answer})
        ai_conversations[conversation_id] = history
        
        return {
            "conversationId": conversation_id,
            "answer": answer,
            "sources": []
        }
        
    except Exception as e:
        error_msg = str(e)
        print(f"❌ [AI聊天] 错误: {error_msg}")
        
        # 返回友好的错误信息
        if "Invalid API key" in error_msg or "Unauthorized" in error_msg:
            friendly_error = "API密钥无效，请检查配置是否正确。"
        elif "rate limit" in error_msg.lower():
            friendly_error = "请求过于频繁，请稍后再试。"
        elif "timeout" in error_msg.lower():
            friendly_error = "请求超时，请稍后再试。"
        else:
            friendly_error = f"处理请求时出现错误：{error_msg}"
        
        return {
            "conversationId": request.conversationId or f"conv_{uuid.uuid4().hex[:12]}",
            "answer": f"❌ {friendly_error}",
            "sources": []
        }

@app.delete("/api/ai/history/{conversation_id}")
async def clear_ai_history(conversation_id: str, current_username: str = Depends(get_username_from_request)):
    """清除AI会话历史（内存中的当前会话）"""
    if conversation_id in ai_conversations:
        del ai_conversations[conversation_id]
    return {"success": True, "message": "会话历史已清除"}

@app.get("/api/ai/history/{conversation_id}")
async def get_ai_history(conversation_id: str, current_username: str = Depends(get_username_from_request)):
    """获取AI会话历史"""
    history = ai_conversations.get(conversation_id, [])
    return {"messages": history}

# ==================== 聊天记录持久化 API ====================

class SaveConversationRequest(BaseModel):
    """保存会话请求"""
    conversationId: str
    messages: List[dict]
    title: Optional[str] = None

@app.post("/api/ai/conversations/save")
async def save_conversation(request: SaveConversationRequest, current_username: str = Depends(get_username_from_request)):
    """
    保存当前会话到历史记录
    """
    global user_data
    user_data = load_user_data()
    
    if current_username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user_info = user_data[current_username]
    
    # 初始化聊天记录列表
    if "chatHistory" not in user_info:
        user_info["chatHistory"] = []
    
    # 如果没有消息，不保存
    if not request.messages or len(request.messages) == 0:
        return {"success": False, "message": "没有消息需要保存"}
    
    # 生成标题（取第一条用户消息的前20个字符）
    title = request.title
    if not title:
        for msg in request.messages:
            if msg.get("role") == "user":
                content = msg.get("content", "")
                title = content[:20] + ("..." if len(content) > 20 else "")
                break
        if not title:
            title = "未命名会话"
    
    # 创建会话记录
    conversation_record = {
        "id": request.conversationId,
        "title": title,
        "messages": request.messages,
        "messageCount": len(request.messages),
        "createdAt": datetime.now().isoformat(),
        "updatedAt": datetime.now().isoformat()
    }
    
    # 检查是否已存在同ID的记录，如果有则更新
    existing_index = None
    for i, conv in enumerate(user_info["chatHistory"]):
        if conv.get("id") == request.conversationId:
            existing_index = i
            break
    
    if existing_index is not None:
        # 更新现有记录
        conversation_record["createdAt"] = user_info["chatHistory"][existing_index].get("createdAt", datetime.now().isoformat())
        user_info["chatHistory"][existing_index] = conversation_record
    else:
        # 添加新记录到开头
        user_info["chatHistory"].insert(0, conversation_record)
    
    # 限制最多保存50条记录
    if len(user_info["chatHistory"]) > 50:
        user_info["chatHistory"] = user_info["chatHistory"][:50]
    
    # 保存用户数据
    user_data[current_username] = user_info
    save_user_data(user_data)
    
    # 清除内存中的会话
    if request.conversationId in ai_conversations:
        del ai_conversations[request.conversationId]
    
    return {"success": True, "message": "会话已保存", "conversation": conversation_record}

@app.get("/api/ai/conversations")
async def get_conversations(current_username: str = Depends(get_username_from_request)):
    """
    获取用户的所有聊天记录列表
    """
    user_data_local = load_user_data()
    
    if current_username not in user_data_local:
        return {"success": True, "conversations": []}
    
    chat_history = user_data_local[current_username].get("chatHistory", [])
    
    # 返回列表（不包含完整消息内容，只返回摘要信息）
    conversations_list = []
    for conv in chat_history:
        conversations_list.append({
            "id": conv.get("id"),
            "title": conv.get("title"),
            "messageCount": conv.get("messageCount", len(conv.get("messages", []))),
            "createdAt": conv.get("createdAt"),
            "updatedAt": conv.get("updatedAt")
        })
    
    return {"success": True, "conversations": conversations_list}

@app.get("/api/ai/conversations/{conversation_id}")
async def get_conversation_detail(conversation_id: str, current_username: str = Depends(get_username_from_request)):
    """
    获取单个聊天记录的详细内容
    """
    user_data_local = load_user_data()
    
    if current_username not in user_data_local:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    chat_history = user_data_local[current_username].get("chatHistory", [])
    
    for conv in chat_history:
        if conv.get("id") == conversation_id:
            return {"success": True, "conversation": conv}
    
    raise HTTPException(status_code=404, detail="聊天记录不存在")

@app.delete("/api/ai/conversations/{conversation_id}")
async def delete_conversation(conversation_id: str, current_username: str = Depends(get_username_from_request)):
    """
    删除单个聊天记录
    """
    global user_data
    user_data = load_user_data()
    
    if current_username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user_info = user_data[current_username]
    chat_history = user_info.get("chatHistory", [])
    
    # 查找并删除
    original_count = len(chat_history)
    user_info["chatHistory"] = [conv for conv in chat_history if conv.get("id") != conversation_id]
    
    if len(user_info["chatHistory"]) == original_count:
        raise HTTPException(status_code=404, detail="聊天记录不存在")
    
    # 保存
    user_data[current_username] = user_info
    save_user_data(user_data)
    
    return {"success": True, "message": "聊天记录已删除"}

@app.delete("/api/ai/conversations")
async def clear_all_conversations(current_username: str = Depends(get_username_from_request)):
    """
    清空所有聊天记录
    """
    global user_data
    user_data = load_user_data()
    
    if current_username not in user_data:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user_data[current_username]["chatHistory"] = []
    save_user_data(user_data)
    
    return {"success": True, "message": "所有聊天记录已清空"}

# ==================== 启动服务 ====================

def main():
    import uvicorn
    print("=" * 60)
    print("🚀 邮件自动化系统 - 后端API服务")
    print("=" * 60)
    print(f"📡 API地址: http://localhost:8000")
    print(f"📚 API文档: http://localhost:8000/docs")
    print("=" * 60)
    uvicorn.run(app, host="0.0.0.0", port=8000)

if __name__ == "__main__":
    main()

